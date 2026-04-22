"""
DAR Mobile App v2.0 - SNFOR SDN BHD
iPhone Web App — buka di Safari, guna kat site terus.

Workflow (Worker):
  1. Buka apps dari phone
  2. Ambik 3 gambar per lantern (full, serial label, issue)
  3. AI scan QR → extract serial, contract, delivery date (auto)
  4. Isi maklumat DAR (ticket, station dll)
  5. Generate & download DAR Excel

Run: python DAR_Mobile.py
Set GEMINI_API_KEY atau ANTHROPIC_API_KEY dalam run_mobile.bat
Akses dari phone: http://YOUR_PC_IP:5679
"""

import os, sys, copy as cp, json, threading, re, time, socket, subprocess
from http.server import HTTPServer, BaseHTTPRequestHandler
from urllib.parse import urlparse
import io, base64

def install(pkg):
    import subprocess
    subprocess.check_call([sys.executable, "-m", "pip", "install", pkg, "-q"])

for pkg in ["openpyxl", "Pillow", "qrcode", "zxing-cpp", "gspread", "google-auth"]:
    try:
        if pkg == "Pillow": __import__("PIL")
        elif pkg == "zxing-cpp": __import__("zxingcpp")
        elif pkg == "gspread": __import__("gspread")
        elif pkg == "google-auth": __import__("google.oauth2.service_account")
        else: __import__(pkg)
    except (ImportError, Exception):
        try:
            print(f"Installing {pkg}..."); install(pkg)
        except: pass

from openpyxl import load_workbook, Workbook
from openpyxl.cell import MergedCell
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.pagebreak import Break
from openpyxl.drawing.image import Image as XLImage
from openpyxl.drawing.spreadsheet_drawing import OneCellAnchor, AnchorMarker, TwoCellAnchor
from openpyxl.drawing.xdr import XDRPositiveSize2D
from openpyxl.utils.units import cm_to_EMU
from openpyxl.styles import Alignment, Font
from PIL import Image as PILImage

GEMINI_KEY    = os.environ.get('GEMINI_API_KEY', '')
ANTHROPIC_KEY = os.environ.get('ANTHROPIC_API_KEY', '')

TEMPLATE_CACHE = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'dar_template_cache.xlsx')

BLOCK_HEIGHT   = 25
BLOCKS_PER_PAGE = 3
FIRST_BLOCK_ROW = 9
SERIAL_COL     = 8
MAX_PRINT_COL  = 21

COL_WIDTHS = {
    1:4, 2:9, 3:9, 4:9, 5:9, 6:9, 7:9, 8:9, 9:9,
    10:9, 11:9, 12:9, 13:9, 14:9, 15:9, 16:9,
    17:4.57, 18:4.57, 19:9, 20:2, 21:7
}

# IMG_CFG: TwoCellAnchor coordinates (0-indexed col/row)
# Only 4 photos go into Excel: full, serial, issue, (qr handled separately)
# Block layout per 25 rows: header=row0, top_img=rows1-11, divider=row12, bot_img=rows13-23
# Full Lantern  B10:G20  → fc=1, fr_off=1,  tc=7,  tr_off=12
# Serial Label  H10:M20  → fc=7, fr_off=1,  tc=13, tr_off=12
# Issue Photo   B22:G32  → fc=1, fr_off=13, tc=7,  tr_off=24
# QR Code       H22:M32  → handled by add_qr_cell
IMG_CFG = {
    'full':   {'fc':1, 'fr_off':1,  'tc':7,  'tr_off':12},
    'serial': {'fc':7, 'fr_off':1,  'tc':13, 'tr_off':12},
    'issue':  {'fc':1, 'fr_off':13, 'tc':7,  'tr_off':24},
}
# Extra component photos (lampu_test, spd, driver) → compiled to separate JPG collage
COMPONENT_PHOTOS = ['lampu_test', 'spd', 'driver']

# ─── Helpers ───────────────────────────────────────────────────────────────────
def get_local_ip():
    try:
        s = socket.socket(socket.AF_INET, socket.SOCK_DGRAM)
        s.connect(("8.8.8.8", 80))
        ip = s.getsockname()[0]
        s.close()
        return ip
    except: return "localhost"

def safe_write(ws, row, col, val):
    if not val: return
    c = ws.cell(row=row, column=col)
    if not isinstance(c, MergedCell): c.value = val

def ai_active(): return bool(ANTHROPIC_KEY or GEMINI_KEY)
def ai_name():
    if ANTHROPIC_KEY: return 'Claude'
    if GEMINI_KEY: return 'Gemini'
    return 'None'

# ─── AI Vision ─────────────────────────────────────────────────────────────────
_last_ai = 0
def call_ai_vision(img_b64, prompt):
    global _last_ai
    elapsed = time.time() - _last_ai
    if elapsed < 4: time.sleep(4 - elapsed)
    import urllib.request
    if ANTHROPIC_KEY:
        payload = json.dumps({
            "model": "claude-haiku-4-5-20251001", "max_tokens": 400,
            "messages": [{"role":"user","content":[
                {"type":"image","source":{"type":"base64","media_type":"image/jpeg","data":img_b64}},
                {"type":"text","text":prompt}
            ]}]
        }).encode()
        req = urllib.request.Request("https://api.anthropic.com/v1/messages", data=payload,
            headers={"Content-Type":"application/json","x-api-key":ANTHROPIC_KEY,"anthropic-version":"2023-06-01"})
        with urllib.request.urlopen(req, timeout=25) as r:
            _last_ai = time.time()
            return json.loads(r.read())['content'][0]['text']
    if GEMINI_KEY:
        payload = json.dumps({"contents":[{"parts":[
            {"inline_data":{"mime_type":"image/jpeg","data":img_b64}},
            {"text":prompt}
        ]}],"generationConfig":{"maxOutputTokens":400,"temperature":0}}).encode()
        url = f"https://generativelanguage.googleapis.com/v1beta/models/gemini-2.0-flash:generateContent?key={GEMINI_KEY}"
        req = urllib.request.Request(url, data=payload, headers={"Content-Type":"application/json"})
        try:
            with urllib.request.urlopen(req, timeout=25) as r:
                _last_ai = time.time()
                return json.loads(r.read())['candidates'][0]['content']['parts'][0]['text']
        except urllib.error.HTTPError as e:
            if e.code == 429: raise Exception("RATE_LIMIT")
            raise
    raise Exception("No AI key")

def resize_img(img_bytes, max_size=1200, quality=82):
    pil = PILImage.open(io.BytesIO(img_bytes)).convert('RGB')
    pil.thumbnail((max_size, max_size), PILImage.LANCZOS)
    buf = io.BytesIO(); pil.save(buf, format='JPEG', quality=quality); buf.seek(0)
    return base64.b64encode(buf.read()).decode()

# ─── QR Scanner ────────────────────────────────────────────────────────────────
def decode_qr(img_bytes):
    try:
        import zxingcpp
        from PIL import ImageEnhance
        orig = PILImage.open(io.BytesIO(img_bytes)).convert('RGB')
        w, h = orig.size
        for attempt, img in [
            ('orig', orig),
            ('2x', orig.resize((w*2,h*2), PILImage.LANCZOS)),
            ('contrast', ImageEnhance.Contrast(ImageEnhance.Sharpness(
                orig.convert('L')).enhance(3.0)).enhance(2.5).convert('RGB')),
        ]:
            r = zxingcpp.read_barcodes(img)
            if r: return r[0].text
        return None
    except: return None

def parse_qr(qr_text):
    if not qr_text: return {}
    parts = [p.strip() for p in qr_text.split('|')]
    if len(parts) >= 6 and re.match(r'^TNB', parts[4], re.I) and re.match(r'^\d{4}$', parts[5] if len(parts)>5 else ''):
        parts[4] = f"{parts[4]}/{parts[5]}"; parts.pop(5)
    r = {}
    if len(parts) >= 4: r['defmodel']      = parts[3]
    if len(parts) >= 5: r['contract']      = re.sub(r'\s+', ' ', parts[4]).strip()
    if len(parts) >= 7: r['delivery_date'] = parts[6]
    if len(parts) >= 8: r['serial']        = parts[7]
    if r.get('contract') and '/' not in r['contract']: r.pop('contract')
    return {k:v for k,v in r.items() if v}

def extract_from_serial_photo(img_bytes):
    try:
        from PIL import ImageEnhance
        pil = PILImage.open(io.BytesIO(img_bytes)).convert('RGB')
        w, h = pil.size
        pil2 = pil.resize((min(w*2,2000), min(h*2,2000)), PILImage.LANCZOS)
        pil2 = ImageEnhance.Sharpness(pil2).enhance(3.0)
        pil2 = ImageEnhance.Contrast(pil2).enhance(2.0)
        buf = io.BytesIO(); pil2.save(buf, format='JPEG', quality=90); buf.seek(0)
        b64 = base64.b64encode(buf.read()).decode()
        prompt = """This is a KINGSUN brand TNB street light lantern label. Read ALL text very carefully.

Extract EXACTLY these 4 fields and return as JSON only:

1. "contract" — The TNB contract number. It starts with "TNB" and contains a "/" slash.
   Example: "TNB 1211/2023". Found at the TOP of the label text (first line).

2. "serial" — The full serial number after "No.:" on the label.
   It is a long code, example: "D4A3M26 IA12-0002 A-150/23-00110".
   Take the FULL string after "No.:" — do NOT truncate it.

3. "delivery_date" — The delivery date after "Delivery Date:" on the label.
   Format MM-YYYY or MM/YYYY, example: "08-2023".

4. "defmodel" — The Model Number from the label, after "Model No:" or "Model No.".
   Example: "RL151028B". This is NOT the product name.

Return ONLY valid JSON with these exact keys. Use null for any field not found.
Example output:
{"contract":"TNB 1211/2023","serial":"D4A3M26 IA12-0002 A-150/23-00110","delivery_date":"08-2023","defmodel":"RL151028B"}

JSON only, no explanation."""
        text = call_ai_vision(b64, prompt)
        text = re.sub(r'```json|```', '', text).strip()
        # Handle edge case where response has extra text before/after JSON
        m = re.search(r'\{.*\}', text, re.DOTALL)
        if m: text = m.group(0)
        result = json.loads(text)
        cleaned = {k:v for k,v in result.items() if v and str(v).lower() not in ['null','none','']}
        print(f"  Label AI extracted: {cleaned}")
        return cleaned
    except Exception as e:
        if "RATE_LIMIT" not in str(e): print(f"  AI extract error: {e}")
        return {}

def extract_from_email_img(img_bytes):
    try:
        # Upscale image for better OCR
        pil = PILImage.open(io.BytesIO(img_bytes)).convert('RGB')
        w, h = pil.size
        if w < 1000:
            pil = pil.resize((w*2, h*2), PILImage.LANCZOS)
        buf = io.BytesIO(); pil.save(buf, format='JPEG', quality=90); buf.seek(0)
        b64 = base64.b64encode(buf.read()).decode()

        prompt = """This is a screenshot of a TNB SMB complaint email. Read ALL text carefully.

Return JSON with EXACTLY these keys:
{
  "ticket": "29947",
  "station": "Kuantan",
  "totalqty": "159 Nos",
  "sitedate": "2022-10-18",
  "contract": "TNB/",
  "pic": "Mohd Zulkifli bin Zuhari"
}

Rules:
- ticket: number after "Ticket #" in the table
- station: value in "Station" column
- totalqty: number from "JUMLAH : XXX BIJI" line in the description text, add " Nos" after number
- sitedate: "Date Submitted" value, format YYYY-MM-DD
- contract: "Contract Number" column value, null if just "TNB/" incomplete
- pic: "Name" value in User Info section at bottom

Return ONLY the JSON object. No explanation."""

        text = call_ai_vision(b64, prompt)
        text = re.sub(r'```json|```', '', text).strip()
        result = json.loads(text)

        # Fix totalqty
        if result.get('totalqty'):
            nums = re.findall(r'\d+', str(result['totalqty']))
            if nums: result['totalqty'] = nums[0] + ' Nos'
        # Fix contract
        if result.get('contract'):
            c = str(result['contract']).strip()
            if len(c) < 6 or c in ['TNB/','TNB','/']:
                result['contract'] = None
        # Fix sitedate
        if result.get('sitedate'):
            raw = str(result['sitedate'])
            m = re.search(r'(\d{1,2})[/\-](\d{1,2})[/\-](\d{2,4})', raw)
            if m:
                d,mo,y = m.groups()
                if len(y)==2: y='20'+y
                result['sitedate'] = f"{y}-{mo.zfill(2)}-{d.zfill(2)}"

        cleaned = {k:v for k,v in result.items() if v and str(v).lower() not in ['null','none','']}
        print(f"  Email AI extracted: {cleaned}")
        return cleaned
    except Exception as e:
        if "RATE_LIMIT" not in str(e): print(f"  AI email error: {e}")
        return {}

# ─── Excel helpers ─────────────────────────────────────────────────────────────
def extract_rows(ws, s, e):
    rows = []
    for r in range(s, e+1):
        rc = {}
        for col in range(1, ws.max_column+1):
            c = ws.cell(row=r, column=col)
            if not isinstance(c, MergedCell):
                rc[col] = {'v':c.value,'font':cp.copy(c.font) if c.has_style else None,
                           'border':cp.copy(c.border) if c.has_style else None,
                           'fill':cp.copy(c.fill) if c.has_style else None,
                           'nf':c.number_format,'align':cp.copy(c.alignment) if c.has_style else None}
        rows.append({'cells':rc,'height':ws.row_dimensions[r].height})
    return rows

def extract_merges(ws, s, e, rel=False):
    merges = []
    for m in ws.merged_cells.ranges:
        if s <= m.min_row <= e and m.min_col <= MAX_PRINT_COL:
            mc = min(m.max_col, MAX_PRINT_COL)
            merges.append((m.min_row-s, m.min_col, m.max_row-s, mc) if rel
                          else (m.min_row, m.min_col, m.max_row, mc))
    return merges

def write_rows(ws, rows, base):
    for i, rd in enumerate(rows):
        r = base+i; ws.row_dimensions[r].height = rd['height']
        for col, cd in rd['cells'].items():
            if col > MAX_PRINT_COL: continue
            c = ws.cell(row=r, column=col); c.value = cd['v']
            if cd['font']:   c.font = cd['font']
            if cd['border']: c.border = cd['border']
            if cd['fill']:   c.fill = cd['fill']
            c.number_format = cd['nf']
            if cd['align']:  c.alignment = cd['align']

def write_merges(ws, merges, base=0):
    for (r1,c1,r2,c2) in merges:
        ws.merge_cells(start_row=base+r1, start_column=c1, end_row=base+r2, end_column=c2)

def add_image_cell(ws, img_bytes, cfg, block_row):
    """Place image snapped exactly to cell boundaries using TwoCellAnchor."""
    try:
        pil = PILImage.open(io.BytesIO(img_bytes)).convert('RGB')
        buf = io.BytesIO(); pil.save(buf, format='JPEG', quality=85); buf.seek(0)
        xl = XLImage(buf)
        # block_row is 1-indexed Excel row; convert to 0-indexed for anchor
        br0 = block_row - 1
        anchor = TwoCellAnchor()
        anchor.editAs = 'twoCell'
        anchor._from = AnchorMarker(col=cfg['fc'],          colOff=0,
                                    row=br0 + cfg['fr_off'], rowOff=0)
        anchor.to    = AnchorMarker(col=cfg['tc'],          colOff=0,
                                    row=br0 + cfg['tr_off'], rowOff=0)
        xl.anchor = anchor
        ws.add_image(xl)
    except Exception as e: print(f"  Image error ({cfg}): {e}")

def add_qr_cell(ws, img_bytes, block_row):
    """Place QR code snapped to H22:M32 cell range."""
    try:
        pil = PILImage.open(io.BytesIO(img_bytes)).convert('RGB')
        buf = io.BytesIO(); pil.save(buf, format='PNG'); buf.seek(0)
        xl = XLImage(buf)
        br0 = block_row - 1
        anchor = TwoCellAnchor()
        anchor.editAs = 'twoCell'
        anchor._from = AnchorMarker(col=7,  colOff=0, row=br0 + 13, rowOff=0)
        anchor.to    = AnchorMarker(col=13, colOff=0, row=br0 + 24, rowOff=0)
        xl.anchor = anchor
        ws.add_image(xl)
    except Exception as e: print(f"  QR image error: {e}")

def make_qr_img(text):
    try:
        import qrcode
        from PIL import ImageDraw, ImageFont
        qr = qrcode.QRCode(version=1, box_size=5, border=2)
        qr.add_data(text); qr.make(fit=True)
        qr_img = qr.make_image(fill_color='black', back_color='white').convert('RGB')
        sz = 163
        qr_img = qr_img.resize((sz, sz), PILImage.LANCZOS)
        th = 28; final = PILImage.new('RGB', (sz, sz+th), 'white')
        final.paste(qr_img, (0,0))
        draw = ImageDraw.Draw(final)
        try: font = ImageFont.truetype("arial.ttf", 7)
        except: font = ImageFont.load_default()
        lines = []; line = ""
        for w in text.split("|"):
            t = line+("|" if line else "")+w
            if len(t) > 28: lines.append(line); line = w
            else: line = t
        if line: lines.append(line)
        y = sz+2
        for l in lines[:3]: draw.text((2,y), l, fill='black', font=font); y+=9
        buf = io.BytesIO(); final.save(buf, format='PNG'); buf.seek(0)
        return buf.read()
    except: return None

def add_qr_cell(ws, img_bytes, block_row):
    try:
        pil = PILImage.open(io.BytesIO(img_bytes)).convert('RGB')
        sz = int(4.3*37.795); pil = pil.resize((sz,sz), PILImage.LANCZOS)
        buf = io.BytesIO(); pil.save(buf, format='PNG'); buf.seek(0)
        xl = XLImage(buf)
        marker = AnchorMarker(col=7, colOff=cm_to_EMU(0.1),
                              row=block_row+13-1, rowOff=cm_to_EMU(0.1))
        xl.anchor = OneCellAnchor(_from=marker,
                    ext=XDRPositiveSize2D(cm_to_EMU(4.3), cm_to_EMU(4.3)))
        ws.add_image(xl)
    except Exception as e: print(f"  QR image error: {e}")

def make_photos_xlsx(units_list, info):
    """
    Create an Excel file with all 6 photos per unit, one row per unit.
    Columns: Serial No | Full | Serial Label | Issue | Lampu Test | SPD | Driver
    Returns xlsx bytes or None if no photos found.
    """
    try:
        from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
        wb2 = Workbook()
        ws2 = wb2.active
        ws2.title = 'Photos'

        # Header row
        headers = ['No', 'Serial No', 'Full Lantern', 'Serial Label', 'Issue', 'Lampu Test', 'SPD', 'Driver']
        header_fill = PatternFill('solid', start_color='1A56DB')
        header_font = Font(bold=True, color='FFFFFF', size=11)
        for ci, h in enumerate(headers, 1):
            c = ws2.cell(row=1, column=ci, value=h)
            c.fill = header_fill
            c.font = header_font
            c.alignment = Alignment(horizontal='center', vertical='center')

        # Column widths
        ws2.column_dimensions['A'].width = 5   # No
        ws2.column_dimensions['B'].width = 22  # Serial
        for col_letter in ['C','D','E','F','G','H']:
            ws2.column_dimensions[col_letter].width = 22

        IMG_H_CM = 5.5
        IMG_W_CM = 5.5
        ROW_H_PT = 150  # ~5.3cm

        ALL_TYPES = ['full', 'serial', 'issue', 'lampu_test', 'spd', 'driver']
        any_photo = False

        for i, unit in enumerate(units_list):
            imgs = unit.get('imgs', {})
            ext  = unit.get('extracted', {})
            serial = ext.get('serial', f"Unit {i+1}")
            row = i + 2

            ws2.row_dimensions[row].height = ROW_H_PT

            # No & Serial
            ws2.cell(row=row, column=1, value=i+1).alignment = Alignment(horizontal='center', vertical='center')
            ws2.cell(row=row, column=2, value=serial).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

            # Photos — columns 3-8
            for ci, ptype in enumerate(ALL_TYPES):
                col = ci + 3
                if ptype in imgs and imgs[ptype]:
                    try:
                        pil = PILImage.open(io.BytesIO(imgs[ptype])).convert('RGB')
                        img_buf = io.BytesIO()
                        pil.save(img_buf, format='JPEG', quality=85)
                        img_buf.seek(0)
                        xl_img = XLImage(img_buf)

                        # Use TwoCellAnchor to snap to cell
                        anchor = TwoCellAnchor()
                        anchor.editAs = 'twoCell'
                        anchor._from = AnchorMarker(col=col-1, colOff=cm_to_EMU(0.05),
                                                    row=row-1,  rowOff=cm_to_EMU(0.05))
                        anchor.to    = AnchorMarker(col=col,   colOff=-cm_to_EMU(0.05),
                                                    row=row,    rowOff=-cm_to_EMU(0.05))
                        xl_img.anchor = anchor
                        ws2.add_image(xl_img)
                        any_photo = True
                    except Exception as e:
                        print(f"  Photos xlsx img error: {e}")
                        ws2.cell(row=row, column=col, value='N/A')

        if not any_photo:
            return None

        # Row 1 height
        ws2.row_dimensions[1].height = 25
        ws2.freeze_panes = 'A2'

        buf = io.BytesIO()
        wb2.save(buf)
        buf.seek(0)
        return buf.read()
    except Exception as e:
        print(f"  Photos xlsx error: {e}")
        return None


def generate_dar(tpl_data, units, new_serials, info, block_data):
    """
    units: list of dicts — {folder, imgs:{full,serial,issue,lampu_test,spd,driver bytes}, extracted:{...}}
    New template: Sheet2, header rows 1-8, block rows 9-33 (25 rows), footer rows 84-86
    """
    tpl = load_workbook(io.BytesIO(tpl_data))
    # Use Sheet2 if available (new template), else active sheet
    tws = tpl['Sheet2'] if 'Sheet2' in tpl.sheetnames else tpl.active

    hr = extract_rows(tws,1,8);  hm = extract_merges(tws,1,8)
    br = extract_rows(tws,9,33); bm = extract_merges(tws,9,33,rel=True)
    fr = extract_rows(tws,84,86);fm = extract_merges(tws,84,86,rel=True)

    wb = Workbook(); ws = wb.active; ws.title = 'DAR'

    # Exact column widths as specified: A=4, B-P=9, Q&R=4.57, S=9, T=2, U=7
    EXACT_COL_WIDTHS = {
        1:4,    2:9,    3:9,    4:9,    5:9,    6:9,    7:9,
        8:9,    9:9,    10:9,   11:9,   12:9,   13:9,   14:9,
        15:9,   16:9,   17:4.57,18:4.57,19:9,   20:2,   21:7
    }
    for col, w in EXACT_COL_WIDTHS.items():
        ws.column_dimensions[get_column_letter(col)].width = w

    ws.page_setup.orientation = 'portrait'; ws.page_setup.paperSize = 1
    ws.page_setup.scale = tws.page_setup.scale or 49
    for a in ['left','right','top','bottom','header','footer']:
        setattr(ws.page_margins, a, getattr(tws.page_margins, a))
    ws.sheet_view.view = 'normal'; ws.print_title_rows = '$1:$8'

    write_rows(ws, hr, 1); write_merges(ws, hm)

    # Header field mapping — exact pink box cells from template
    # Row 5: F5=Ticket, J5=Station, N5=PIC, S5=TotalQty
    # Row 6: F6=SiteDate, J6=Contract, N6=DefModel, S6=DefQty
    # Row 7: F7=DONo, J7=DODate, N7=NewModel, S7=Delivery
    hmap = {
        'ticket':   (5,6),   # F5
        'station':  (5,10),  # J5
        'pic':      (5,14),  # N5
        'totalqty': (5,19),  # S5
        'sitedate': (6,6),   # F6
        'contract': (6,10),  # J6
        'defmodel': (6,14),  # N6
        'defqty':   (6,19),  # S6
        'dono':     (7,6),   # F7
        'dodate':   (7,10),  # J7
        'newmodel': (7,14),  # N7
        'delivery': (7,19),  # S7
    }
    for key, (row, col) in hmap.items():
        val = info.get(key,'')
        if val: safe_write(ws, row, col, val)

    # Auto-fill from first unit extracted data
    for u in units:
        ext = u.get('extracted',{})
        if not info.get('contract') and ext.get('contract'): safe_write(ws, 6, 10, ext['contract'])
        if not info.get('defmodel') and ext.get('defmodel'): safe_write(ws, 6, 14, ext['defmodel'])
        if ext.get('contract') or ext.get('defmodel'): break

    total = len(units); cr = FIRST_BLOCK_ROW; bn = 0; lr = 0
    for i, unit in enumerate(units):
        write_rows(ws, br, cr); write_merges(ws, bm, cr)
        ext = unit.get('extracted', {})

        # Serial No — H9 in block (col 8), merged H9:M9
        display_serial = ext.get('serial') or unit.get('folder', str(i+1))
        _sc = ws.cell(row=cr, column=SERIAL_COL)
        if not isinstance(_sc, MergedCell): _sc.value = f'Serial No: {display_serial}'

        # Block number — A9 (col 1)
        ca = ws.cell(row=cr, column=1)
        if not isinstance(ca, MergedCell): ca.value = str(i+1); ca.number_format = '@'

        # New lantern serial — N29:U30 → row cr+20, col 14
        ns = new_serials[i] if new_serials and i < len(new_serials) else ''
        safe_write(ws, cr+22, 14, ns)

        # Delivery date — N19:U22 → row cr+10, col 14
        dd = ext.get('delivery_date') or info.get('dodate','')
        if dd:
            try:
                c = ws.cell(row=cr+10, column=14)
                if not isinstance(c, MergedCell):
                    c.value = f'Delivery Date\n{dd}'
                    c.alignment = Alignment(horizontal='center', vertical='top', wrap_text=True)
            except: pass

        # Per-block dropdowns
        bd = block_data[i] if i < len(block_data) else {}
        safe_write(ws, cr+7,  14, bd.get('cause',''))   # Cause of Problem N16:U17
        safe_write(ws, cr+17, 14, bd.get('action',''))  # Action N26:U27
        safe_write(ws, cr+2,  14, bd.get('rca',''))     # RCA N11:U12

        # Photos — all 6 types
        imgs = unit.get('imgs', {})
        for img_type, cfg in IMG_CFG.items():
            if img_type in imgs and imgs[img_type]:
                add_image_cell(ws, imgs[img_type], cfg, cr)

        # QR Code — H22:M32 slot
        qr_text = ext.get('qr_code')
        if not qr_text:
            for st in ['serial','full','issue']:
                if st in imgs:
                    qr_text = decode_qr(imgs[st])
                    if qr_text: break
        if qr_text:
            qr_img = make_qr_img(qr_text)
            if qr_img: add_qr_cell(ws, qr_img, cr)
            try:
                c = ws.cell(row=cr+13, column=8)
                if not isinstance(c, MergedCell): c.value = None
            except: pass

        cr += BLOCK_HEIGHT; bn += 1
        if bn == BLOCKS_PER_PAGE or i == total-1:
            write_rows(ws, fr, cr); write_merges(ws, fm, cr)
            lr = cr+len(fr)-1
            if i < total-1:
                ws.row_breaks.append(Break(id=lr, min=0, max=16383, man=True))
            cr += len(fr); bn = 0

    ws.print_area = f'$A$1:$U${lr}'

    # ─── Sheet 2: All Photos ───────────────────────────────────────────────────
    try:
        ws2 = wb.create_sheet(title='Photos')
        headers = ['No', 'Serial No', 'Full Lantern', 'Serial Label', 'Issue', 'Lampu Test', 'SPD', 'Driver']
        from openpyxl.styles import PatternFill as PF2
        hfill = PF2('solid', start_color='1A56DB')
        hfont = Font(bold=True, color='FFFFFF', size=11)
        for ci, h in enumerate(headers, 1):
            c = ws2.cell(row=1, column=ci, value=h)
            c.fill = hfill; c.font = hfont
            c.alignment = Alignment(horizontal='center', vertical='center')

        ws2.column_dimensions['A'].width = 5
        ws2.column_dimensions['B'].width = 24
        for ltr in ['C','D','E','F','G','H']:
            ws2.column_dimensions[ltr].width = 22
        ws2.row_dimensions[1].height = 25
        ws2.freeze_panes = 'A2'

        ALL_PH = ['full', 'serial', 'issue', 'lampu_test', 'spd', 'driver']
        for i, unit in enumerate(units):
            imgs2 = unit.get('imgs', {})
            ext2  = unit.get('extracted', {})
            serial2 = ext2.get('serial', f"Unit {i+1}")
            row2 = i + 2
            ws2.row_dimensions[row2].height = 150
            ws2.cell(row=row2, column=1, value=i+1).alignment = Alignment(horizontal='center', vertical='center')
            ws2.cell(row=row2, column=2, value=serial2).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

            for ci2, ptype in enumerate(ALL_PH):
                col2 = ci2 + 3
                if ptype in imgs2 and imgs2[ptype]:
                    try:
                        pil2 = PILImage.open(io.BytesIO(imgs2[ptype])).convert('RGB')
                        ibuf2 = io.BytesIO(); pil2.save(ibuf2, format='JPEG', quality=85); ibuf2.seek(0)
                        xli2 = XLImage(ibuf2)
                        anc2 = TwoCellAnchor(); anc2.editAs = 'twoCell'
                        anc2._from = AnchorMarker(col=col2-1, colOff=cm_to_EMU(0.05), row=row2-1, rowOff=cm_to_EMU(0.05))
                        anc2.to    = AnchorMarker(col=col2,   colOff=-cm_to_EMU(0.05), row=row2,   rowOff=-cm_to_EMU(0.05))
                        xli2.anchor = anc2
                        ws2.add_image(xli2)
                    except Exception as pe:
                        ws2.cell(row=row2, column=col2, value='N/A')
    except Exception as e:
        print(f"  Sheet2 photos error: {e}")

    buf = io.BytesIO(); wb.save(buf); buf.seek(0); return buf.read()

# ─── State ─────────────────────────────────────────────────────────────────────
import datetime

DAR_HISTORY_FILE = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'dar_history.json')

def load_history():
    try:
        if os.path.exists(DAR_HISTORY_FILE):
            with open(DAR_HISTORY_FILE,'r',encoding='utf-8') as f:
                return json.load(f)
    except: pass
    return []

def save_history(history):
    try:
        with open(DAR_HISTORY_FILE,'w',encoding='utf-8') as f:
            json.dump(history, f, ensure_ascii=False, indent=2)
    except Exception as e:
        print(f"  History save error: {e}")

def sync_to_gsheet(record):
    """Append one DAR record to Google Sheet."""
    try:
        import gspread
        from google.oauth2.service_account import Credentials

        # Load credentials from Secret File or env var
        creds_path = '/etc/secrets/credentials.json'
        creds_env  = os.environ.get('GOOGLE_CREDENTIALS','')
        sheet_id   = os.environ.get('GOOGLE_SHEET_ID','')

        if not sheet_id:
            print("  GSheet: GOOGLE_SHEET_ID not set"); return

        if os.path.exists(creds_path):
            creds = Credentials.from_service_account_file(
                creds_path,
                scopes=['https://www.googleapis.com/auth/spreadsheets']
            )
        elif creds_env:
            import tempfile
            with tempfile.NamedTemporaryFile(mode='w', suffix='.json', delete=False) as f:
                f.write(creds_env); tmp = f.name
            creds = Credentials.from_service_account_file(
                tmp,
                scopes=['https://www.googleapis.com/auth/spreadsheets']
            )
            os.unlink(tmp)
        else:
            print("  GSheet: No credentials found"); return

        gc = gspread.authorize(creds)
        sh = gc.open_by_key(sheet_id)
        ws = sh.sheet1

        # Add header if sheet is empty
        if ws.row_count == 0 or ws.cell(1,1).value != 'Date':
            ws.insert_row(['Date','Ticket','Station','Contract','Staff','Units','Filename'], 1)

        # Append record
        ws.append_row([
            record.get('date',''),
            record.get('ticket',''),
            record.get('station',''),
            record.get('contract',''),
            record.get('staff',''),
            record.get('units',''),
            record.get('filename',''),
        ])
        print(f"  ✓ Synced to Google Sheets: Ticket {record.get('ticket')}")
    except Exception as e:
        print(f"  GSheet sync error: {e}")

STATE = {
    'template': None,
    'units': [],
    'new_serials': [],
    'dar_info': {},
    'last_collage': None,
    'history': load_history(),  # list of {id, ticket, station, contract, staff, date, units, filename}
    'dar_files': {},             # id -> xlsx bytes (in-memory cache)
    'email_config': {
        'address': os.environ.get('OUTLOOK_EMAIL', ''),
        'password': os.environ.get('OUTLOOK_PASSWORD', ''),
    }
}

# ─── Bundled default template (New_DAR_Template.xlsx) ─────────────────────────
_BUNDLED_TPL_B64 = "UEsDBBQABgAIAAAAIQBysddwvAEAAMkHAAATAAgCW0NvbnRlbnRfVHlwZXNdLnhtbCCiBAIooAACAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAADMVdtO4zAQfUfaf4j8ukpcQEIINeVhF3hjK8F+gImniWl8kWeA9u934ha0qtKWqJXgJU5izzlnLp4ZXy9sm71CRONdKU6LkcjAVV4bV5fi7+NtfikyJOW0ar2DUiwBxfXkx8n4cRkAM7Z2WIqGKFxJiVUDVmHhAzjemfloFfFnrGVQ1VzVIM9GowtZeUfgKKcOQ0zGv2GmXlrKbhb8e6XkyTiR/Vqd66hKoUJoTaWIhcpXpzdIcj+bmQq0r14sQxcYIiiNDQDZtgjRMGN8ACJ2DIXs5XwOUG+QGtuJThv9NtP7u16T4OotLBFaHObaOnYFWyb3sTEBf3KAtzB0O9tjt7b7w0mPRkM2VZHuleUIy0Ur33ycP3k/L3aDDE1ASkRhlXHvunfwp8Mo03J6ZCGdfwl4oI6zb6Lj/It0EN9skOl5eEoSzJ4EIC1bwGOXYQLdx9yoCPqBuGfURxfwP/YeHTqqt06CXL8cHvc10EDew0v/c7yBp5B3KFfrJzJvMYdFBdzek+Uut3guTKMPyKMnwvCieu/AnXUeGAgiGfjowX297IOR59bBVQzdYNSge7hlGsSTfwAAAP//AwBQSwMEFAAGAAgAAAAhALVVMCP0AAAATAIAAAsACAJfcmVscy8ucmVscyCiBAIooAACAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAACskk1PwzAMhu9I/IfI99XdkBBCS3dBSLshVH6ASdwPtY2jJBvdvyccEFQagwNHf71+/Mrb3TyN6sgh9uI0rIsSFDsjtnethpf6cXUHKiZylkZxrOHEEXbV9dX2mUdKeSh2vY8qq7iooUvJ3yNG0/FEsRDPLlcaCROlHIYWPZmBWsZNWd5i+K4B1UJT7a2GsLc3oOqTz5t/15am6Q0/iDlM7NKZFchzYmfZrnzIbCH1+RpVU2g5abBinnI6InlfZGzA80SbvxP9fC1OnMhSIjQS+DLPR8cloPV/WrQ08cudecQ3CcOryPDJgosfqN4BAAD//wMAUEsDBBQABgAIAAAAIQB7DDjRPAMAAHoHAAAPAAAAeGwvd29ya2Jvb2sueG1spFVRb9owEH6ftP/gWZW2PYTECQkhKkwhFA1pm6qNrY+VSUxjNbEz2xSqav995wQoLQ/rNgROnDu+++7uO+f8w7au0B1TmksxwqTnYcRELgsubkb4+2LmxBhpQ0VBKynYCN8zjT+MX78630h1u5TyFgGA0CNcGtMkrqvzktVU92TDBFhWUtXUwFbduLpRjBa6ZMzUlet7XuTWlAvcISTqJRhyteI5m8p8XTNhOhDFKmqAvi55o/dodf4SuJqq23Xj5LJuAGLJK27uW1CM6jyZ3wip6LKCtLckRFsF3wh+xIPF30cC00momudKarkyPYB2O9In+RPPJeRJCbanNXgZUt9V7I7bHh5YqegfWUUHrOgRjHj/jUZAWq1WEijeP6KFB24+Hp+veMV+dNJFtGm+0Np2qsKootpcFNywYoQHsJUb9uSBWjeTNa/A6g9j38fu+CDnS4VA/azDWpRcX+10bp1AE2llmBLUsEwKAxLcpfS/cmuxs1KCuNFX9nPNFYOZAmlBmrDSPKFLfUlNidaq6oqnYdqKXiFz3av4HesJZtyiYPmAsJDS1XBQ5KE7ZfrWyMZN52gBNUAzKBny3SPt0tNB+Qv10tzWxYXCdOS7++dFghxUslfopVEI7ufTT9Clb/QOehZAkyCRdqbn0JX4+oGkXjoZ+KHj9dOh05+FEyfOsswhQRp504hkXub9gjRUlOSSrk25E4IFHeE+AJ6YPtPt3kK8ZM2LRwIP4UUQDSex76ThrO/0vSGEu8iIM50Fk6B/MUuzSfTLpmqPvB+cbfSjZOwWba+4KOQGchmCzu+f7Dat6YoXpoQjlgRxiFH37CPjNyXwJX0/gIc0N9DJBV2CMG0CvmU5wg/e7uPAdWoXz5nBp132tpade0SvPWiBZntFoh2OaEAc3yMxeheQQUDew9luj2Nbc4KRSmw0NS+IzfT4f9+sF3gcvIHewbudnxNv8Dh4B0feQSuYPbmCrbhghR03oHq02xG+3lai7l0qLsx1Cq8PO8w5rVo+lrWHx2+fJ/X2zVl6RpKz72dxdO4eYf4pwIKbCl5vLwsBAeJn6MexIBvgmdvjBC4dVTsp+9fm+DcAAAD//wMAUEsDBBQABgAIAAAAIQDVEz0JJwEAAFEEAAAaAAgBeGwvX3JlbHMvd29ya2Jvb2sueG1sLnJlbHMgogQBKKAAAQAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAC8lMtqwzAQRfeF/oPRvh7baZNQImdTCtm26QcIefwg1gON+vDfVzjFriF1NyYbwcyge49GGu32X6qNPtBRYzRnaZywCLU0RaMrzt6Oz3dbFpEXuhCt0chZh8T2+e3N7gVb4cMmqhtLUVDRxFntvX0EIFmjEhQbizpUSuOU8CF0FVghT6JCyJJkDe63BssnmtGh4MwdihWLjp0Nzv9rm7JsJD4Z+a5Q+wsW8GnciWpEH0SFq9BzNqQI+soqDsQMLsNs/oBRjXSGTOljaRScOYJ/uoE0mZ4SbOi10aP/Oaaf/Jx5duVOZHMw6ZVh0jmY9ZIwVAuHxat3YQRovKdJeg7mYVEY37Vh4obXSn08Z3+/pL0Pc4yjex9Cvw73AZOPIP8GAAD//wMAUEsDBBQABgAIAAAAIQAfMZZe9xwAAIy3AAAYAAAAeGwvd29ya3NoZWV0cy9zaGVldDEueG1stN1bc9tIdsDx91TlO6j0vpZIibqwbG9ZIgERNxK33U3eNDJtq2yZjiSPZ5LKd08D6EOcxr9X5ZlFtpKd2Z/6AuKgG81DNPn6r789fDn4dfv4dL/7+uZw8ur48GD79W73/v7rxzeHdRX85eLw4On59uv72y+7r9s3h79vnw7/+vbf/+31j93j56dP2+3zgWnh69Obw0/Pz9/mR0dPd5+2D7dPr3bftl/NXz7sHh9un83/fPx49PTtcXv7vq308OVoenx8dvRwe//1sGth/vgzbew+fLi/2y52d98ftl+fu0Yet19un83xP326//YkrT3c/UxzD7ePn79/+8vd7uGbaeKX+y/3z7+3jR4ePNzNVx+/7h5vf/liXvdvk9Pbu4PfHs3/Tc3/n0g3raOnh/u7x93T7sPzK9PyUXfMfPmXR5dHt3f7lvj6f6qZyenR4/bX+yaAfVPTP3dIk9m+rWnf2MmfbOxs31hzuh7n3+/fvzn8n2P7n7+Yf06a/zru/0v+9r+Hb1+318nm8cBcjNvs9sHEoGwut8nh0dvX7+9N9JtXfPC4/fDm8N1k/u7dxVnzl7bW3+63P57Uvx/8auDN4bfbj9srcwl+3jQnbPvj8OC/d7uH8u62CfC5ufL3/zNrrtovbw5nGtvuk9vfd9+fmw66Ks0w+GW3+9zIyry84+bIt1+2d80FefD0X3J8/wgWk+PTi9l5e5RH+8N8+7r/dznkoB0z5qW/3364/f7ludj9uNnef/z0bAbozJzJ5lKcv/99sX26M2PA9PlqOmte+93ui2nC/PfBw30zmM01fPtb+88f9++fP705PH11Pjm+PDk3jdx9f3rePfy98/ac7uuZyHf1zg4PbMXLn6o4OZeaZtKwNWevphezyezMHOFLfU4upar5lz/U6dSEqD3c5l9szenPvU45QVPzL7bm+Us1j7oT3MZucft8+/b14+7HgRm2pv7Tt9tmEpzMm8bMv/Svt4tmW0gi2J38fUz9ITUX0l3T+FXX4KwLcSPXkAVkuZcDc4U8mcZ+fXv8+uhXc73d2YYDVAohN5AVJILEkASSQjLIGrKB5FbMxbp/oRP3hRb7Is0gac5gCakgtZYjE+l9uM0A+X8Md9O6mXlUuCELyBISQELIDWQFiSAxJIGkkAyyhmwguRUzs+yje3IxCO++zD68kApSW7loh8blvkkn0ua+9f8Y6aZ1N9KQBWQJCSAh5AaygkSQGJJAUkgGWUM2kNzKeXsnawZpASkhFaS2ctG04wTzdBDMZjaevTq+dP5jphHMzs+f7u8+X+2a2693qj6Rqfpd04WZ+6f713DVSXeDbefuISyGsBxCMIRwCDdDWA0hGkI8hGQI6RCyIayHsBlCbk9IH9MhlEOohlB3cIJ4mhWFMzibeJ5cvrqcySqy+ae5Gn7qbjvbx7Bp1jSk7ydTd8K5skX6S/XaSnvRtXeYRSen5h/9xOU2s7RF2hVhWymw0jccWukbvrFi1kn7hk/dhlddEbN4lQkxstJfhzE6T2wZMy/uG565DadoOMMRr3HEG0/DZ27DOY6mQFellfZCaE9XhVq17twZ/Cacf+5i+fnB33TRXIH94O/ktI/DtZU+DotOZvqsnw+uE1ukbyaw0ncV2obVdWJFXyeDG+cKDUednPVdxbaZ/iJNrJhTvb9O+ttnG5gUlTLbVR+8NY544zkVk8GaNUfLBV5Eib4q1Kp1786FYlYavgulf8+klu0/f200rZr3X+rGYKU/IddWTvfXz6IT59qYDNa2S1tGXRxW+pZD27K6OKzoi2MymOVWaDmy0h9hbNtRV4fvmE8G0whaznDMaxzzxtfyYObLcTwF+io7OTP38v4Nw2Ciq9BOrY/HuWCaBJl+89fehv6lVUTTollF9AG86qQ/79dDWHTQv2lYDiHooI9UOISbYZXVEKJhlXgIybBKOoRsWGU9hM2wSj6EYghlB/19sxpC3QFXhWYIIHzT41cTvYg4NtfKH1xFNM2+OTybtu8uBmP2yv5Rz6CTwR3x2pbp47WALCEBJITcWDEXdL8yGdxyVqgVQWJIAkkhmRUTjn70DXpfd2VO+nBuIDmkgJSQClJb4YUxMfkkXBmTmVlf/uGLoW3KXA1qRFs61/f9yeD+fC2F+nomidheWOc9LUkBKRTqJ5GbnzmGFZuKSDEpIaU8hkxOTb+UWUup/pa2IeWkglSSKlIt1C7LnMl90iSBhrP7n7wIunyScxF0ZNYWslC/bjs0K4X+xS9IS1JACi11Ked2YXYj1Pe4YsWIFJMSUsoeMyFnwA/Wi2s2tSHlpIJUkipSLeQJeZMYGinkXY7JCXlHTsg7ckIOWk5AASm05IQcPa5YMSLFpISUssdMqL+G16QNKScVpJJUkWohT4CbDNFIAe6STU6AO3IC3JETYNByAgpIoSUnwOhxxYoRKSYlpJQ9Zpb0nN0dQy8blMmtmMlWZruCVJIqUi3kiW6TLhopul3myYluR050O3KiC1pOQAEptOREFz2uWDEixaSElLLHTMisjPZLtOngffhaCvXrkA0pJxWkklSRaksX7eXj3qSbrM9IIe8SSE7IO3JCbtNM+iYNWjaf1TZv+tVbc1JoyQk5elyxYkSKSQkpZY+ZpTM9ou1B6FUYKJe29JhGqZKlKlIt5BnTTbZmpADbDJdeinfkBNimh3SAQcsJKCCFlpwAo8cVK0akmJSQUvaYCelV2HTwtnTNpjaknFSQSlJFqoU8IW/yLSOF3CZ8dMg7ckJuEzw65KBl84jDcEyDQlvKCTl6XLGtiBSTElLKHjMhvQqzB6HHNChnxYJUkipSLeQJcJOQGSnANoOkA2zzP/qdVUfOfRq0nIACUmjJCTB6XLFiRIpJCSllj5kls9CV9dRaSN+Vu+MyCyYplUupPhVQCPXnq2SpilQLdY816ackmkd6xgqwTRrpAHfkjOCOnACDlu1hDe7KKBXaUk6A0eOKbUWkmJSQUvaYCTkLsUGWfi2FdMjtkeqQgwpWLEkVqbbkWYg1j2CNFPK2KTdlJtQvUq4tmWf35OJekJakgBQK9VnHG/a4YsWIFJMSUsoeM/a4ZsUNKScVpJJUkWohTtrtw23DSXuEbLl9aE4vuy1d6I9ppoMPkq6lUD8FLkhLUkAKhfpRcyNk3oL174SGH7bbQuoajEgxKSGlpMzS5LifmtdSTI15Uk4qSCWpItVyJvjmyzzsMtqYZ7qsbf3N4YXzUckwTW4LqZz4grQkBaRQSKXJf+YYVmwqIsWkhJTyGDJL+g2alFKLOVJOKkglqSLVQp55YbyUmvnYs/3kTN3rLel7vSV9ryctSQEplOb7j9pu2OOKFSNSTEpIKXvMLOnFnCWVU4PkUk29/yaVQurzUVItxI/CpuPl1NqmBvd15tRsKSe8zKmxVEAKLemlnJD6FIQVI1JMSkgpe8yEnKXc4KGDtRTS07o9N2opx1IFqSRVpNqSbyk3Xk6teRS++Sxcj2jm1GwpJ+TMqbFUQAotOSFnTo0VI1JMSkgpe8zkZaucmpTSUzZzaixVkEpSRaqFPFP2eDm1KXNqlpwpmzk1W0rFfEkKSKE0r6ds5tRYMSLFpISUssfM0oWTUxs8xrGWenpM2yPVYxpUsGJJqki1HJZnqTZeTm3KnJolJ+TMqdlSTsiZU2OpUJrXIWdOjRUjUkxKSCl7zIRUTo20IeWkglSSKlIt5BnT4+XUpsypWXICzJyaLeUEmDk1lgqleR1g5tRYMSLFpISUssfMkp6zu2PQq7Ch5LaW/mSTVJIqUi3kie54CbUpE2qWnOgyoWZLOdFFqYClQmleR5cJNVaMSDEpIaXsMbN04azChg+fST09YzOhxlIFqSRVpFoOizP2yXgJtbYpdxVmSYfckl6FkZakgBRK8yrk7HHFihEpJiWklD1m8rLViJZSahVGykkFqSRVpFqIY/pkvOfL2qYGAebzZbaUE+CulB7TLBWQQkt6mS2k3lmxYkSKSQkpZY+ZpUt1S2bFDSknFaSSVJFqIU+Ax0uPndj0WJ+aurLUbX1vn/C7tnTaT18L0pIUkEJpXqXE2eOKFSNSTEpIKXvMWGpN2pByUkEqSRWpFmrPvbtF1Jf6+uO7zk66tJd5Qy6fZFxZMh+37J8OJS1IS1JACkk3pBUpIsWkhJSSMtKatCHlcr5U4otUkipSLeQZvb7E1wifd5zYJ8u67QHD7YXyVyfpPdwfIIXUBgHSkhSQQtKN0Iu7BFgvIsWkhJSSMiH9BnsyXK7ZQnqzACknFaSSVJFqIeZJT8ZLmrVNDe7mNlX04pYBW09/FkJakgJSKKQ+CxF66RhWbCoixaSElPIYMjk1eknHh9WkYj9B5KSCVJIqUi3kmTPGS6w1m5AHmVNLzpqdiTVbylnS8WE1lgqleb1mZ2KNFSNSTEpIKXvMhJxxP9wywKY2pJxUkEpSRaqFPCEfL7Fm5jGEnA+r2VLOKp4Pq7FUQAotOat4JtZYMSLFpISUssdMSK/i+bAaS+WkglSSKlIt5AnweIm1EybWLDljmok1W8oZ00yssVQozesxzcQaK0akmJSQUvaYWdJzNhJrKJNb0Yk1UkmqSLWQJ7rjJdbM7n8MX6S5rm0pZ/gii7ZkqYAUWnKGLxNrrBiRYlJCStljJuQk1oZbBqSQSqyRclJBKkkVqbbk+XjTnPqxnlppm3JXapb0gLakQ05akgJSKM2rAc0eV6wYkWJSQkrZYyYvW41oKaUSa6ScVJBKUkWqhTimT8dLrLVNDQLMxJotZY5o/ygiaUkKSKElPaaFVGKNFSNSTEpIKXvMhJyPN4dbBtjUhpSTClJJqki1kCfk46XazBd1DKdxS86Y5sZNW0pdBUtSQAqleT2muXGTFSNSTEpIKXvMhNQqjLQh5aSCVJIqUi3kCfB4T5mZ9CgCzI2btpQzprlxk6UCUmjJGdPcuMmKESkmJaSUPWaW9FNmQuquTMqF1JYBIbVlgKUqUi3ELQOn4z1n1jY1mLT5nJkt5QSYz5mxVEAKLTkB5t5NVoxIMSkhpewxE3IWYsMtA1JIh5zPmbFUQSpJFam25FuIjZcya75rbZAtEVIpdEt6ywBpSQpIoZD6fIQ9rlgxIsWkhJSyx4w9rllxQ8pJBakkVaRayDNp+9JhI6TQT5kbs/TylgEppLYMkJakgBQKqS0DQi9uGbCF9JYBUkxKSCkps+RsGZBieszzOTSWKkglqSLVcib4VMPpeOmytqnBNN/ljV7eMmDr6TQ5aUkKSKGQSpNbevEYVmwqIsWkhJTyGDI5NfoNGlNqUlGlyUkFqSRVpFrIMy+Ml1I7ZUrNkrNaZ0rNlnJW60ypsVQozevVOlNqrBiRYlJCStljZslZzCGnZgv1F0Eu1dQnp6RSSG0ZINVCni/bHS+ndsqcmiUnvHxYzZZywsuH1VgqlOZ1eJlTY8WIFJMSUsoeMyFnKTfcMiCF9LTOh9VYqiCVpIpUW/Is5ZofSxhp92fblDutW9Iht6RX76QlKSCF0rwKOXtcsWJEikkJKWWPmbxsNWVLKZVTI+WkglSSKlItxCl7Nl5OrW1qEGDm1GwpJ8B8WI2lAlJoSb89E1I5NVaMSDEpIaXsMbP08pYBqafGNCknFaSSVJFqOSwu1Zrvnh1rTDOn1rZufkdFfUuDJSfk/DI0lgpIoTSvxzRzaqwYkWJSQkrZYyakcmqkDSknFaSSVJFqIc+YHi+nNmNOzZITYCTQFraUvk+TAlIozesAM6fGihEpJiWklD1mlvScjS9DQ5nciv5kk1SSKlIt5InueAm1mX1kTe3is+REF9mzhS3lRBelApYKpXkdXSbUWDEixaSElLLHzNLLWwaknp6xmVBjqYJUkipSLYflmbHHS6iZ36AZJtQsOSHHLs2FLeWEHKUClgqleR1ybtxkxYgUkxJSyh4zedl6RPP5Mqmo3jiTClJJqki1kGdMj/d82Yw5NEtOgPl8mS3lBJjPl7FUKM3rAPP5MlaMSDEpIaXsMbOktwyw4oaUkwpSSapItZAnwOOlx2b2aTL1VLklvWXAkvlcZv/RNWlJCkihNK9S4uxxxYoRKSYlpJQ9Ziy1Jm1IOakglaSKVAtxy8DMl/r641sG2mbM7wTo4Hb5H71lwJZStCAtSQEpJN2QVqSIFJMSUkrKSGvShpTL+VKJL1JJqki1kGf0+hJfI3zeMbNZsG7LwOCLkK7kry9uGZBCassAaUkKSCHpRujFLQOsF5FiUkJKSZmQ8+jwcMuALWQe95T5bkPKSQWpJFWkWoh5UvNTO2O9wW6bcnMqll7+lQEppKZ/0pIUkEIh9VnIzxzDik1FpJiUkFIeQyanRi3ppJRKrJFyUkEqSRWpFuKc0fwWzUhZlrapwUXAxJotpbMspCUpIIWWdGJNSCXWWDEixaSElLLHTMgZ98MtA2xqQ8pJBakkVaRayBPy8RJr5kdmhm/TLOlVvCUn5EyssVRACqV5tYpnjytWjEgxKSGl7DETUok10oaUkwpSSapItZAnwOMl1sw3JyHAfFjNlnICzIfVWCoghZacMc3EGitGpJiUkFL2mFnSczYSayiTW9GJNVJJqki1kCe64yXWzphYs+QMXybWbCn9JpwUkEJpXg9fJtZYMSLFpISUssdM6MUtA1JIJdZIOakglaSKVFvyfLx5Nl5irW1qcJNGmuvalnIGNLJoS5YKSKElZ0AzscaKESkmJaSUPWbysvWIZmJNKqrEGqkglaSKVAt5xvR4ibUzJtYsOWOaiTVbyhnTTKyxVCjN6zHNxBorRqSYlJBS9pgJvbhlgE1tSDmpIJWkilQLeUI+XqrtjBs3LTkhxy7NhS3lhBylApYKpXkdcm7cZMWIFJMSUsoeMyG9CuNTZiyVkwpSSapItZAnwOM9ZXbGp8wsOQHmU2a2lBNgPmXGUqE0rwPMp8xYMSLFpISUssfMkn7KTEjflW3iUX2noZRSWwaE1JYBlqpItRC3DJhvRB7trTOfM2tbdx9QsOTclbl3k6UCUijN6wDzOTNWjEgxKSGl7DETenHLgBTSIedzZixVkEpSRaoteRZi5rHpsULeNuUuxITUlgFLessAaUkKSKGQ+nyEPa5YMSLFpISUsseMPa5ZcUPKSQWpJFWkWoiT9rkvHTZCCr1tdxDtLjf28pYBW+9CbRkgLUkBKRRSWwaEXtwyYAuZkyY564gUkxJSSsosOVsGpJga86ScVJBKUkWq5UzwqYYmgz1ShrRtanAVdImwl7cM2Hp6ywBpSQpIoZBKk1t6ecsAm4pIMSkhpTyGTE6NeoMmpVSanJSTClJJqki1kGdeGC+lZqI4TKlZ0os5S/peT1qSAlIozat7PXtcsWJEikkJKWWPmSW9mLPUh3wDyaWa+uSUVAqpLQOkWogfhZ2Pl1NrmxqMcWS4rm0pJ7xIsy1ZKiCFlnSCRUh9CsKKESkmJaSUPWZCL24ZkEJ6WufDaixVkEpSRaot+ZZy4+XUzLjFiGZOzZZyQs6cGksFpNCSE3Lm1FgxIsWkhJSyx0xetp6ymVOTiiqnRipIJaki1UKeKXu8nJo5zwgwMlzXtpQTYCTQliwVkEJLToCZU2PFiBSTElLKHjNLL28ZkHp6THN3J0sVpJJUkWo5LM9Sbbyc2jlzapbMFbf/UlRLTsiRQFuyVEAKpXl9l2ZOjRUjUkxKSCl7zIRUTo20IeWkglSSKlIt5BnT4+XUzplTs+QEmDk1W0rn1EgBKZTmdYCZU2PFiBSTElLKHjNLes7Gxk2Uya3oTzZJJaki1UKe6I6XUDtnQs2SE11kzxa2lBNdlApYKpTmdXSZUGPFiBSTElLKHjNLL28ZkHp6xmZCjaUKUkmqSLUcFmdsszAb681125S78LakQ25Jz9ikJSkghdK8Cjl7XLFiRIpJCSllj5m8bDWipZR640zKSQWpJFWkWohjuklv9dkTcz/7ZH7OcjJ7dWlWinffn553Dzfb+48tHh78Njm9vZu//32xfbrbfjUFj1+ZX5F6+7r9evl3bVODAPP5MlvKCTA3brJUQAot6VWYkHpnxYoRKSYlpJQ9Zpb0lgFW3JByUkEqSRWpFvIE2EmP/WsBtk+TqafKTeKzWXnrLQOW9JYB0pIUkEJpXqXE2eOKFSNSTEpIKXvMWGpN2pByUkEqSRWpFuKWgQsn9WUD7Nsy8Pzp/u7z1a4Z396hfNKP5C4FprcPtL2YmuqzD9KCtCQFpJB0Q1qRIlJMSkgpKSOtSRtSbsn8OLK88yhIJaki1UKekewkwbq3Hs2P+HXTdLB7fLjtYtvM1hfHr44v9X9MZcznf+CK6NJAl+bK6X/uePBju1cXtlA/P1wLqc9HhJxM1OB53KUtNDnuT2mwt/bUtPecUBrTx3Uy+HbZGx7Xat+Wmb/2L+hk8BWW0b5UvxSL99a/ysR7FIOvVkt5FJlQM/J+fXt6NpkOvxiXlTbezgabPnKe96LvTC7Skq1XpFqov6X+zZL5PUFp6++kf5D+g/SfpHfvHHN+U6X5EMSzXvnXLu4uK3Zp0iL9tTD4Le+rtuM3h+arq/bZBdKCtCQFpJB0Q1qRIlJMSkgpKSOtSRtSTipIJaki1UL8FMCsMsaPfZcLu3R2Jg2+Sueq7XgQe5s5bI+ynYsWttSlfsTtZPDbLEs2FZBC0g1pRYpIMSkhpaSMtCZtSDmpIJWkilQLqavh6OnTdvu8uH2+ffv6Yfv4cXu9/fLl6eBu9715S2JOv+KDx+2HN4fvLs7m16YlM3L3Fbq/LCfzTTu5D9zsWJs3m8hY42oyX3hrTI7nabcCGLRlHqOZN480sC3z3mHerOw8fzk5mzc/OsC/mMvM1PG9lsXZfOnzbDZft3P34Lg25/PC18PNmXkl3uO9OpnNm2+25FFF5/PY11Y2Ma+w+0xx0Pu7i5l5Hb62gvN56G3r1JxH8xVE7D05m6feV35mzpV5AtBzfs/McZkdGp74msg3Wxj5l5tT0003/w8jPDP9mE29vgifztfmbuaL8KmJsO/1m8/qzPXi+8vV+cScmnb5MziCq+l03vwmG/sx64V5sxzgX96dnM7feY/6yvyl+eUn33kzx2aek/acnbP5yheD8ty8GF/MZmaAebxsQubreWr+YH7flj0HZ/PQ27O5wLxHej5f+a8vc02Yr5j2vOqJ6Xviv77NNekf9VNzpszy3HOFTU0Up74o3phJx3flLWbzpa+ld7P5tXcEz+Yr78gyw9fni/P50tfOu8v5O+9FlTej2vfSzCQY/pNJcL72BSk5N3ONryUzCn2nIp/OK9/0a1KrZjrxTQ5XZjprnq3zhNUM6OZrKD1Dw0wC77wTs3nnMG/eA3jCalprvvbV08/MXArmWzs8g+Z8Om9SDr6/TMyJ8Z3jemrmDV8v5+bVmI9mPP2fz9f+i9cc18Q7BZ2YAPunk9m88NYwp7/7+GE4N5o5q3lwx3NYZhzUU//dz7TWZQgx05o65iurfME0L8Z7bzK/7WtejO8CMG8f5827Q8+seTYvfKGMZvPYe/87m8e+8uZdl7ksvTOpucSafb+e0JvZt/nNO88lZi6XJifmu/jMTdM7l9+YO0Pz25mefsz5T73zkHlDay5L3xHcmBVA803TniFj6jRv9z39nJox7j22hanTvK03dY76ldzb199uP27T28eP91+fDr5sP7SJ5rPp5eV0Mp1dXpxOZ5PL5rHvxy4xffyq+Z+z45PLk/Pj4/OL8+Z3IJ5335r09Mnk9PLseF+3+Wzrl92zyWr/kz9+2t6+3z42f/S0+WG3e/5nfzQvoTnqcvv8/dvB093tl615v2BSHLvHe5Mpv32+3319c/ht9/j8eHv/bA59fv/+zeHj6n07wt8/3v64//qx1zZcRz92j5/bte7b/wMAAP//AwBQSwMEFAAGAAgAAAAhAIJhZszbDAAA7GcAABgAAAB4bC93b3Jrc2hlZXRzL3NoZWV0Mi54bWycnVtvG8kRhd8D5D8QfJc4VcO5CZIWXmuN7EOCINdnmhpJhEmNQtI3BPnvqWpF7K5YB9BZYL2WreOj1nA+NYv8OLz86dtuO/sy7g+b6fFqLufVfDY+rqfbzeP91fzvf/tw1s9nh+Pq8Xa1nR7Hq/n38TD/6fr3v7v8Ou0/HR7G8TizhsfD1fzheHy6WCwO64dxtzqcT0/jo33mbtrvVkf74/5+cXjaj6vb9I9224VWVbvYrTaP8+eGi/1bOqa7u816vJnWn3fj4/G5ZD9uV0db/+Fh83R4adut31K3W+0/fX46W0+7J6v4uNlujt9T6Xy2W1/8ev847Vcft/Z9f5Plaj37trf/1H7VL18m/f0PX2m3We+nw3R3PLfmxfOaf/z2h8WwWK1PTT9+/2+qkeViP37Z+A2Yq/S3LUmaU5fmsvo3lrWnMj9c+4vPm9ur+b/fVcsPN133/kxuqv5sKcNwNgy/fDj7Weqqff/uRm5+7v8zv75M58mf9zM7Gcc/rXZ2G/zVTzedL64vbzd26/t3PNuPd1fzd3Lxrq4q/0z6V//YjF8PxcczP1k/TtMn/8SvtojK+8ftuPbTZray376M78ft9mr+S2fn+79Sq31ohYtTY/nxS/uHdHrbKm/Hu9Xn7fEv09c/jJv7h6Ox1Ng37WfNxe33m/GwttPVvvC5Nt66nrZWYf+f7TbOnZ1uq2/p96+b2+ODfVSfa99I01p+tv58OE67f/7vM2lVzwVpbTer4+r6cj99ndkZZE2Hp5XzKBf28esLsK/s2XcWsDUd7GB8ua6Hy8UX+xbX9su6ToV2Hry90MKnwmX1eqGdC28vtHAulNcLl0yhhXOhvl5oR/ztK7RwLqxfL2yZQgvnwuXrhX6SvvlWtnAubF4v9J/yby60cC5sXy8cmEIL58Lu9UKxnYk4sy2dK3tQycFS0rIEtAiFi6dPq2wAL0IB4+lcCYgRChlP50rAjP+0I26ekpoGUCMUNp7OqwTcCAWOp3MlIEcodDydKwE7QsHj6VwJ6FGKHk/nSkCPUvR4OleivYbbbEp6WkCPUvR4+rTKFtCjFD2ezpWAHt/kiV22pKcF9ChFj6fzKgE9StHj6VwJ6FGKHk/nSkCPUvR4OlcCemqKHk/nSkBPTdHj6VwJ6Kkpejx9quzQfTXuzlpJTwfoqSl6PJ1XCeipKXo8nSsBPTVFj6dzJaCnpujxdK4E9NQUPZ7OlYCemqLH07kS0LOk6PF0rgT0LCl6PJ0rAT1Lih5Pnyp7QM+SosfTuRLNOtywU9LTA3qWFD2ezqsE9CwpejydKwE9S4oeT+dKQM+SosfTuRLQs6To8XSuBPQ0FD2ezpWAnoaix9O5EtDTUPR4+lQ5AHoaih5P50pAT0PR4+lciR4r4B4sKOkZAD0NRY+n8yoBPQ1Fj6dzJaCnoejxdK4E9DQUPZ7OlYCelqLH07kS0NNS9Hg6VwJ6WooeT58qpQL4tBQ+ni46AT8txY+ni04AUEsB5OmiEz3exj3gVhIkFUCopRDydLFOwFBLMeTpohNA1FIQebroBBR1FEWeLjoBRh2FkaeLTsBRR3Hk6dwpgKOO4sjTRSfgqKM48nTRCTjqKI48XXQCjjqKI08Xneiha+6x68CRAI46iiNPF+sEHHUUR54uOgFHPcWRp4tOwFFPceTpohNw1FMceTp3KuCopzjydNEJOOopjjxddAKOeoojTxedgKOe4sjTRSfgqKc48nTRiZ4F4p4GChwp4KinOPJ0sU7A0UBx5OmiE3A0UBx5uugEHA0UR57OnTXgaKA48nTRCTgaKI48XXQCjgaKI08XnYCjgeLI00Un4GigOPJ00Qk4GiiOPF10oidUKY6GwFENn1OlQLJZI6wUPa1aUSiJx4vvHz2zWlEwicdzK5IRpKJwSvGiFT2/WlFA2cwR1oqeYq0opMTjxVrRs6wVBZV4vGhFT7RWFFbi8aIVPddaUWCJx4tW9HRrRaElHi9aEVuksODxohUqCxxbEtiC1gKpLQRvQaC4QJoLQV2QBrFFygvBXpAGsUX6Cx7PtxY0GEiFITgM0iC2SIshaAzSILZIkSGYDOapIcWGYyvIDNIgtjidQYLPIA1iizMaJCgN0kAjiNu3NOxbSGsQzmtI8Xy+IrNBOLUhxYtWxBZnN4jHi1a0b3GCgwTDQVrEFuc4SJAcpEVscZqDBM9BWsQWZzpIUB2kRWxxsoNJu+HWQmxxvoME4UFaxBanPEhwHgRJD1Jz9wk9ns9X5D0IJz6keNGK2OLcBwnyg3SILU5/kOA/SIfY4gwICQqEdIgtToKQYEFIh9jiPAgJIoR0iC1OhZDgQkiH2OJsCAk6hHSILU6IkGBECFIihHMiUjxT0EOflZu3lmHeQmKEcGZEihdrRWxxcoQEO0J6xBbnR0gQJKRHbHGKhARHQnrEFmdJSNAkpEdscaKETS3lT+0escW5Eja1hFbEFqdL2NRStiJhQjhjIsXz+TogtjhpwqaWsFaoi3OPZTThPiEyJ2wM4Sz08FjGgNji7AmbWsIRQGxxAoVNLaEVscU5FDa1hFbEFqdRSPAoZEBscSaFBJXCXjYGJk5OppBgU9hrElErd58w+BRaIbY4ocKmluLW0gqxxSkVNrWEVvhqDI6ttmRLkVVhYwhFbPAqtEJscWKFBLNCK8QWp1bY1BKOK2KLkytsagmtiC1Or5DgV2iF2OIECwmGhSLDwsYQ6hwIjoUKYouTLGxqKY+rILY4zcKmltCK2OJEC5taQit8sRPHVlfuW4pcCxtDuFur3LdUEFucbmFTSzgCiC1OuLCpJbQitjjlQoJzoYLY4qQLCdaFIutCOO0ixU/3CVURW5x4YVNLeVwVscWpFza1hFbEFidf2NQSWhFbnH5hU0toha8l5NjqA1vIwBBOwUjx4hxAbHEShgQLQxWxxWkYEjwMVcQWJ2JIMDEUmRjCqRgpno9rjdjiZAwJNobWiC1Ox5DgY2iN2OKEDAlGhtaILU7JkOBkaI3Y4qQMm1pKYmu0b7loQbzuOXgZirwMG0Oo14MGL0NrwJaNIVxr+ViGwutDcF6GTS3lcUVehnJeRopntpbolbucl2FTS1grYMvGEO64hn1riV6/y3kZNrWEtQK2lPMyUrw4roAt5byMFC9aAVs2hnDHNdwnXIJ9y8YQqjV4GbpEbLloQbyCO3gZirwM5byMFM/HFXkZynkZKV60IrY4L0ODl6HIy7AxhDuugS3kZdgYwrUGtpCXoZyXkeLFcUVscV6GBi9DkZdhYwh3BAJbyMuwMYRqDV6GIi/DxhCuNexbyMtQFy0IYoOXofByE+T1JsIFJxRecYK85ES45oS2aN8irzoRvAyF150gLzwRvAxFXoaS154IXoYiL8PGEO4cCPcJkZdhYwjXGthCXoaNIVRr8DK0RfsW52Vo8DIUeRk2hnBrDfcJkZdhYwjXWnoZirwMG0O41nCfsENscV6GTS3lvTfkZSjnZaR43mGQl2FjCHcEwmMZyMuwMYRrDWwhL8PGEK41sIW8DOW8jBQvjitii/MyNHgZirwM5byMFM9rRV6Gcl5Gihet6D6hixbEHhu8DEVehnJeRooXa0XzFudlaPAyFHkZynkZKV6sFd0n5LwMm1rKny7Iy7AxhLu1AlvIy1DOy0jx4gggtjgvQ4OXoT14nNDGEOoIBC9DkZehnJeR4vkIIC9DOS8jxYtWtG+5aMFciSzsW8jLsDGEaw3zFvIylPMyUrw4AogtzsuwqaVka0CPZXBehgYvQwf0WAbnZdjUEtaK2OK8DJtaQitii/MybGopWu3ivuDqtNxlLmxqCa1o3+K8DJtaQitii/MybGoJrWjfctGCIDZ4GTXyMpTzMlL8xFaNvAwbQ7i1lmzZ1aHROcDtW8HLqCvEFudlaPAy6gqxxXkZNrWEcwCxxXkZGryMGnkZynkZKZ7PAeRlKOdlpHjRitjivAwNXkYtiC3Oy9DgZdSCHoN30YIgNngZNfIylPMyUrw4rogtzsuwqaU8XwWxxXkZGryMWhBbnJehwcuokZdhYwh1awUvo0ZehnJeRornWwt5GTaGcGsN+xbyMuw69Fxr2LcUscV5GRq8jFoRW5yXocHLqBXdJ3TRgiA2eBk18jJsDOFaA1uK2OK8DJtaSmKRl6Gcl5HixfmK9i3Oy9DgZdTIy1DOy0jxvFbkZSjnZaR40Yr2Lc7L0OBl1MjLUM7LSPFirYgtzsuwqaU8s5CXoZyXkeLFWtG+xXkZGryMGnkZ/h4jxM+BFC/W+v/71vNbizy/fcfT6n7842p/v3k8zLbjnf2r6tyO3/75vUTSx8fpKf2t/Xj9OB3tDUFe/vRg76oz2vt5VOc23dxN0/HlD/72Jaf36bn+LwAAAP//AwBQSwMEFAAGAAgAAAAhALs3TSWhHAAAorYAABgAAAB4bC93b3Jrc2hlZXRzL3NoZWV0My54bWy03Vtz20aWwPH3rdrvoNJ7LJESKYlle0oXAiJuIm4zu/umyLStimV6JSXO7NZ+920A5xDd+HeplDGcmknsH/sC4jRuh93k27/9+fBl74/N49P99uu7/cmbw/29zde77Yf7r5/e7ddV8Mvp/t7T8+3XD7dftl837/b/uXna/9v7f/+3t9+3j789fd5snvdMC1+f3u1/fn7+tjg4eLr7vHm4fXqz/bb5al75uH18uH02f338dPD07XFz+6Gt9PDlYHp4OD94uL3/ut+1sHh8TRvbjx/v7zZX27vfHzZfn7tGHjdfbp/N9j99vv/2pK093L2muYfbx99+//bL3fbhm2ni1/sv98//bBvd33u4W6w+fd0+3v76xbzvPyfHt3d7fz6a/03N/4+0m9bR08P93eP2afvx+Y1p+aDbZr79s4Ozg9u7XUt8/69qZnJ88Lj5474JYN/U9F/bpMls19a0b+zoX2xsvmus2V2Pi9/vP7zb/9/lxcl8dno4/2V5dn75y/Hx5OqXi/nR/Jfjk2B+dHR5Pj08v/y//fdv23Gyftwzg3GT3T6YGJTNcDvaP3j/9sO9iX7zjvceNx/f7Z9PFufnp/PmlbbW3+8335+sP+893/5abr5s7p43Zgsm+3vP22/J5uPz5ebLF1P71LzX/9luH8q72ybUJ+YY2P01a8avKdRgM+R/3W5/a5pfmYYOm61sm2225Pbu+f6PTddkNj8zh81/txtn/ryoT6bNxh3sts7+s25p0B4q5h1/2Hy8/f3Lc7H9fr25//T52WzyzOzAZgQuPvzzavN0Z4a+6f7NdNa0erf9Ypow/957uG+OYTN0b/9s//v9/sPz53f7x29OJodnRyemkbvfn563D//ofCK1u3pmJ3T15uaddgXOXlVxcqI1zblCas7eTE9nk9ncbOFLfU7Mbuo6NX/4S51OTTzams0fpOb0de9Td9DU/EFqnrxU86DbwW3srm6fb9+/fdx+3zNHq6n/9O22OfdNFk1j5g/9++2i2RbSCHY7fxdTf0jNmLprGr/oGpx1IW7kEnIFWe5kz4yQJ9PYH+8P3x78YcbbnTQcoFIIuYasIBEkhiSQFJJBbiBrSC5iBuvujU7cN1rsijQHSbMHS0gFqW05MJHehdscID8x3E3r7/btcEOuIEtIAAkh15AVJILEkASSQjLIDWQNyUXMmWUX3aOjs0F8d4V28YVUkFrktD0h2/E1F6mfGN+mdTe+kCvIEhJAQsg1ZAWJIDEkgaSQDHIDWUNykZP2CtQcmgWkhFSQWoTBPB4EszkHz94cnjn/mJMHzsnPn+/vfrvYNhdd7wn6SE/Q500X5ozfXtm7U3Yn3WW1PWMP4WoIyyEEQwiHcD2E1RCiIcRDSIaQDiEbws0Q1kPIZYf0MR1COYRqCHUH7U2ec/I19xHOwdnE8+jszdns0PrHjIYfiGfThWnUvqJM3TPOhRTph+2lSDsA20Fw1cmx+U9/6nKbWUqR9oa1rRSI9A2HIn3D1yLmTmnX8LHb8KorMjvchSAS6cdkjM4TKWNOjLuGZ27DKRrOsMU32OK1p+G523COrSnQVSnSDop2d1WoVdudOwPHhPNnD5ymi2Y07vb6RSfHfRwuRfo4XHUys/f6yWCcSJG+mUCk7yqUhq1xImKPk9PBOEHDUSfzvqtYmukHaSJidvVunAwuySkqZdJVH7wbbPHasysmg7vWHC0XeBMl+qpQq7Z7dwaKudfwDZT+qcm6cX/9RaJp1TyBWRcJkX6HXIoc78bPVSfO2JgM7m6XUsYaHCJ9y6G0bA0OEXtwTAZnuRVajkT6LYylHWt0+Lb5aHAaQcsZtvkG27z2tTw48+XYngJ9lZ3MzXW9f2QYnOgqtFPb2+MMmCYzZj/+tZekH7oCNS2aO4o+gBed9Pv9cghXHfRPicshBB30kQqHcD2sshpCNKwSDyEZVkmHkA2r3AxhPaySD6EYQtlBf92shlB3wDvEJkMzDN/08M3EvqE4NGPlB+4omi7e7c+n7WP44Pi9kBfts+lkcHW8lDJ97K4gS0gACSHXImZw93cpg8vPCrUiSAxJICkkEzGh6Y/EQe83XZmjPrRrSA4pICWkgtQiHCQTk13CKJnMzH2nZ2D4M3SSezhvmzKjwTq6hUxy0NoPg2v1pRbq612RlqSAFCr1J5Tr12zDik1FpJiUkFJuQ6a7pr+tudFS/eVtTcpJBakkVaRaqb1Fc070kyYlNDxV/IuDoMsuOYOgI3OfoWmMy7ZDc9fQv/kr0pIUkEKhk/4ccq3U97hixYgUkxJSyh4zpf5G5IYV16ScVJBKUkWqlTwBbnJCIwW4Sy85Ae7ICXBHToBBywkoIIVCToDR44oVI1JMSkgpe8yU+hF7Q1qTclJBKkkVqVbyBLjJE40U4C7l5AS4IyfAHTkBBi0noIAUCjkBRo8rVoxIMSkhpewxE7LP0N029LJGmVzEnFr13FaQSlJFqpU80W2SRiNFt8s/OdHtyIluR050QcsJKCCFQk500eOKFSNSTEpIKXvMlMx90O6GbDp4Ar/RQv2NyJqUkwpSSapItdBpO3zcS3KT7xkp5F3qyAl5R07IJcFkX5JBy+Zz2uZx33ooJ4VCTsjR44oVI1JMSkgpe8yEzAfVemjeaCn7nku2q6ecpQpSSapItZLnmG7yNCMFWHJb9o13R06AJTFkBxi0nIACUijkBBg9rlgxIsWkhJSyx0zJvufC1q/ZVk4qSCWpItVKngA3eZWRAiyJHTvAHTkBlkSOHWDQspnMMDyCQaGUcgKMHldsKyLFpISUssdMyb7nko2wj2BQzooFqSRVpFrJE+Am8TJSgCVTZAdY8jz2U1NHzlUZtJyAAlIo5AQYPa5YMSLFpISUssdMyNzW9qfobiPMvZDSWkv1lCv1j/mFUr+/SpaqSLVS+/zoXoObrMlIAZaEkB3gjpwjuCMnwKBlM6doeASDQinlBBg9rthWRIpJCSllj5nQ1M7CTQfZ+Bst1J/H16ScVJBKUkWqldpbA3cKzHjpsGbeVpMctUKu1N+SXAqZ2Xc63q9IS1JACpX6jOI1e1yxYkSKSQkpZY+Z0JF10lbqPxtck3JSQSpJFalW6iYlOnOcfKmuV2XFX05+yvQ4J9pd3uvU/jhmOvjA6FLqmRmaffylXk9LlgpIoVI/4q6VzANX/9wz/FBdClljMCLFpISUkjIhZ0R0b/HIHhGgXCv2pQpSSapItZJnRPyV3NjrPxc102Zw+Hd06nwiMsyGS70TKxtOWpICUqhkZcOFXtyGFZuKSDEpIaXchkzoyPrIXKnf0jUpJxWkklSRaqX2fsG9BoyXSzOfdGIQMJcmpezLPmlJCkihkH3ZV7Ky4awYkWJSQkrZYyZk39cJWck0SK7VrGQaqVSyPhIl1Ur8xGs6XjKtbWpwiWcyTUo54WUyjaUCUijkhJfJNFaMSDEpIaXsMVNykmmDeQY3Wsi6kSflpIJUkipSLeRJpjWT1ke6kW+bGoScyTQp5YScyTSWCkihkBNyJtNYMSLFpISUssdM37aVTNNS1qM4KScVpJJUkWolPopPx0umtU0NAsxkmpRyAsxkGksFpFDICTCTaawYkWJSQkrZYyZ0aiXTtJR9BMt2WY/iLFWQSlJFqnUjmA5vniDHOoKZTGtbN0ugrFyLkBNgJtNYKiCF2rz1CTV7XLFiRIpJCSllj5mS/VzGZBpL5aSCVJIqUq3kOYLHS6ZNmUwTcgLMZJqUsmK+JAWkUJu3A8xkGitGpJiUkFL2mAnZZ+huG+x7rqHkUsv+AJNUkipSreSJ7niZtCkzaUJOdJlJk1JOdJlJY6lQm7ejy0waK0akmJSQUvaYCZ0691zDGWVazz5jy5baZ2xQwYolqSLVulk8Y5vH/bHO2G1T7iVZyA65kH3GJi1JASnU5q2Qs8cVK0akmJSQUvaY6du2jmgtZd1zkXJSQSpJFalW4jFtkjevD/Dr8yZts4NgcwKZlHKC3ZWyj2+WCkihkH3/pWQ9MrNiRIpJCSllj5nQmXV5ZsU1KScVpJJUkWolT7B/TpLsSJJkfdrnQqhb7N4uhrkUOrZSYqQlKSCF2ryVI2ePK1aMSDEpIaXsMWOpG9KalJMKUkmqSLVSu++dBJhJzPHI/rGlZ22TZua/HeguEWY+i9lNC5VSFl2RlqSAFJKuSStSRIpJCSklZaQb0pqU6/6yUmGkklSRaiXPUe1LhY28ROBIJpx1awSG6w311RcXCWgha5UAaUkKSCHpWunFpQKsF5FiUkJKSZnSiwsGpJC9YoCUkwpSSapItZJnLfl4KbUjzk8TenndgBayLwuSGOtpyVIBKVSyPil5zTas2FREikkJKeU2ZLpr7FtATFhba0VrDhupIJWkilQrec4f46XdmlXJg0/LhZx7fM5hk1LObR/nsLFUqM3b9/hMu7FiRIpJCSllj5mSlXZjxTUpJxWkklSRaiVPgMdLu5mzFgLMOWxSyrmvZ9qNpQJSKOTc13MOGytGpJiUkFL2mCnZ9/VMu7FUTipIJaki1UqeAI+Xdjti2k3IOYKZdpNSzhHMOWwsFWrz9hHMtBsrRqSYlJBS9pgJ2WdopN1QJhex026kklSRaiVPdMdLu5nF/zh8kQS7lFLO4csJbCwVkEIh5/Bl2o0VI1JMSkgpe8yUnLTbcN2AFrLSbqScVJBKUkWqhTwfdZpdP1barW3KzcQI2Qe0kB1y0pIUkEJt3jqg2eOKFSNSTEpIKXvM9G1bR7SWstJupJxUkEpSRaqVeEwfj7dWs21qEGCm2qSU2aLdDDXSkhSQQiH7mFayUm2sGJFiUkJK2WOmZN1zseKalJMKUkmqSLWSJ8DjrdU038oxPGkLOUcw12pKKSvmS1JACrV5+wjmWk1WjEgxKSGl7DFTsu65SGtSTipIJaki1UqeAI83v8ykRBFgzi+TUs4RzLWaLBWQQiHnCOZaTVaMSDEpIaXsMROy55cpWddgUq5krRtQstYNsFRFqpW4buB4vBlmbVODUzRnmEkpJ8CcYcZSASkUcgLMGWasGJFiUkJK2WMm1EwX6KctD9cNaCFr3QApJxWkklSRaiWuG2i+DW2k+SltU4OQS4bMSpVLKXvdAGlJCkihkvWZiG5E3+OKFSNSTEpIKXvMhMznjbu1QErWLHFSTipIJaki1UqcJX7sS3WNsG6gbXcQ7S4L9fK6AalnrxsgLUkBKVSy1g0ovbhuQArZ6wZIMSkhpaRMyBkR3a4xn1PuVodpqZ5yUkEqSRWpVvKMiL+SG/sLX7rJPNlxRy+vG5BC9roB0pIUkEIlKxv+mm1YsamIFJMSUsptyITM51n9KaLbNeZzin5AgHKtaF/2UapkqYpUK3HdwPF4ubS2qcFZAZmtSynlXPaROFuyVEAKhZzLPnNprBiRYlJCStljJuTc1yGZJoX6K1Ou1awPS0mlkrVugFQreb5wd7xk2jGTaULOcxnnsEkp57mMc9hYKtTm7ecyJtNYMSLFpISUssdMyUmmDdcNaCH7Rp5z2FiqIJWkilQLeZJp5ntmx7qra5tyj2ghO+RC9hFNWpICUqjNWyFnjytWjEgxKSGl7DHTt20l07SUlUwj5aSCVJIqUq3ER/HZeMm0tqlBgJlMk1JOgDlvjaUCUihkn7KVrGQaK0akmJSQUvaYCdnrBrSUdQSTclJBKkkVqdaN4CzU5htmR3oua5saBJhffCalnADzi89YKiCFQk6AmUxjxYgUkxJSyh4zJeu5jLQm5aSCVJIqUq3kOYLHS6bNmEwTck7RyJxdSSn7qkwKSKE2b5+imUxjxYgUkxJSyh4zIfsMjS8+Q5lcxP4Ak1SSKlKt5InueJm0mcxD689KF0JOdJE2u5JSTnRRKmCpUJu3o8tMGitGpJiUkFL2mAm9vG5A69lnbNlSa90ASxWkklSRat0szxl7vEya+a2ZYXZcyAk5FmZeSSkn5CgVsFSozdsh51pNVoxIMSkhpewx07dtH9GcNKYVrUljpIJUkipSreQ5pv/KpLHX501mnEAm5ASbE8iklBNsTiBjqVCbt4PNCWSsGJFiUkJK2WMmZK8bYMU1KScVpJJUkWolT7B/TpJsJkkyazq5kL1uQMh8ULP75Jq0JAWkUJu3cuTsccWKESkmJaSUPWYsdUNak3JSQSpJFalW4rqBmS8B9mPrBtom3XUDQva6AdIVaUkKSCHpmrQiRaSYlJBSUka6Ia1Jue4vKxVGKkkVqVbyHNW+VNjI6wZmkiPr1g0MvjbpQl99cd2AFrLWDZCWpIAUkq6VXlw3wHoRKSYlpJSUKb24bkAKmVmguyQ5KScVpJJUkWolZlHND/CM9UDeNuU+kAu9vG5AC1mXBdKSFJBCJeuTktdsw4pNRaSYlJBSbkOmu8a6BdRSVtqNlJMKUkmqSLUSzx/NL9SMlJVpmxoMAqbdpJSdlSEtSQEpFLKzMkpW2o0VI1JMSkgpe8yUrDlsrLgm5aSCVJIqUq3kCfB4aTfz4zLDhzgh+75eyAkw024sFZBCbd66r2ePK1aMSDEpIaXsMVOy0m6kNSknFaSSVJFqJU+Ax0u7ma9MRIA5h01KOQHmHDaWMj+53TZvVQyFnCOYaTdWjEgxKSGl7DETss/QSLuhTC5ip91IJaki1Uqe6I6Xdpsz7SbkHL5Mu0kp+7GcFJBCbd4+fJl2Y8WIFJMSUsoeM6UX1w1oISvtRspJBakkVaRayPNR53y8tFvb1OCSjCTYpZRyDmjk2JYsFZBCIeeAZtqNFSNSTEpIKXvM9G3bRzTTblrRSruRClJJqki1kueYHm+t5pypNiHnmGaqTUo5xzRTbSwVavP2Mc1UGytGpJiUkFL2mCnZ91zY+jXbykkFqSRVpFrJE+Dx1mrOOQdNyAkwFmZeSSknwCgVsFSozdsB5lpNVoxIMSkhpewxU7LvubhWk6VyUkEqSRWpVvIEeLz5ZXOu1RRyAsy1mlLKCTDXarJUqM3bAeb8MlaMSDEpIaXsMROy55cp2dfgbrusnyDItZQ1gVDJWjfAUhWpVuK6AfNTNaM9FnOGWdu6+yWHQs41mMs1WSoghdq8HWDOMGPFiBSTElLKHjOhl9cNaCFr3QApJxWkklSRaiWuGzBzp8cKeduUe9ulZK0bELLXDZCWpIAUKlmfibDHFStGpJiUkFL2mAnZs8SVrFnipJxUkEpSRaqVOEv85Cf93kDb7iDar/m9AalnrxsgLUkBKVSy1g0ovbhuQAqZnab56IgUkxJSSsqEnBHB3xvQUta6AVJBKkkVqVbyjIif81VqTc578AUqQi+vG5BC9roB0pIUkEIlKxv+mm1YsamIFJMSUsptyITsdQNK1roBUk4qSCWpItVKXDdgdv1o1wDm0trW3cu+kH3ZJy1JASnU5q3LPntcsWJEikkJKWWPmZB9XyfUX/nWkFyrWR+Wkkola90AqVbiJ14n4yXT2qYGJ32uBpVSTni5GpSlAlIoZGdWlKwPO1gxIsWkhJSyx0zpxXUDWsi6kSflpIJUkipSLeRJppmHqNGOaM5ha1sfHNGcwyal7Cc1UkAKtXn7iGYyjRUjUkxKSCl7zITMjfNuqZeWsj7AJOWkglSSKlKtxEdxczyMFmAm09rWBwFmMk1KOQFmMo2lQm3eDjCTaawYkWJSQkrZYyZkrxvQUvYRLNtlzUJlqYJUkipSrRvBWagn4yXT2qYGJ22kti6llHPSRuZsyVIBKRQy704Pnmsl+6SN5iO2FZMSUsoeMyUrmUZak3JSQSpJFalW8hzB4yXTTphMEzInkN132wo5AeZiTZYKSKE2bweYyTRWjEgxKSGl7DETss/QWKuJMrmI/QEmqSRVpFrJE93xMmknzKQJOdHlWk0p5ZyfuVaTpUJt3o4uM2msGJFiUkJK2WMm9PK6Aa1nn7G5VpOlClJJqki1bhbP2OY2bKxLctuUe8YWskMuZB/QpCUpIIXavBVy9rhixYgUkxJSyh4zfdvWEa2lrHsuUk4qSCWpItVKPKab78PoJ42Zi9dn89Uwk9mbM3Nzevf70/P24Xpz/6lF860xn+/vfrvYNiX29zy/3Hm0//5t++Xy522zg2BzApmUcoLNdZssFZBCIfvyrGRdnlkxIsWkhJSyx0zIXjfAimtSTipIJaki1UqeYDtJsvGCzd8bMJnQJm9mrxsQstcNkJakgBRq81aOnD2uWDEixaSElLLHjKVuSGtSTipIJaki1UpcN3DqJMAk2D+2bqBt0l03IGSvGyBdkZakgBSSrkkrUkSKSQkpJWWkG9KalOv+slJhpJJUkWolz1HtpMK654/mR/2603ewfXy47c7YzVn89PDN4Zn9j6n8I+f5LgV2Zs7a/VeKDX589+JUCvVp4Usl6wealZx81Nnbgz/evz24k+vKUgpNDvtdGuys/9Ah1Mbs7ToafMPsNbdrtWvLnL92b+ho4m5FtCvV36LFO+vfZeLdisEXrqXcikypuZ7+8f54PpkOvxyXldbezgYrPHLu96LvTJ+qSrZekWql/vL6dyHzi4Pa1j9I/0H6T9J/kc7PHXN+XaX5OMZzH/Njg7tL0J2ZdEk/FgY/833Rdvxu33yj1e6hlHRFWpICUki6Jq1IESkmJaSUlJFuSGtSTipIJaki1Ur8LMDcZYwf+y5HduYsQxp8q85F2/Eg9pIs7KdBXEmpM3s10dF8eGKTev0wCth6SLomrUgRKSYlpJSUkW5Ia1JOKkglqSLVStZoOHj6vNk8X90+375/+7B5/LS53Hz58rR3t/39qzl8zQONxXuPm4/v9s9PZ4tmXJkjd1dBX5kvLk0ffMWE09TxvWLOTqZOe/oftHZlXmkuU2zNXKgWzXWIr5hL0aK5qPAVc1lZNFcNvmIuHIsb/ysnx4sm8+6pc2Lej8no+V45M++0vdAO3o95zjGvtNfuwSvmKrRoLjJszVxnTB3vVs9Ma+a7SDxbMDN1zJp33ytmq82aWU/kZmeLc++2meWNi2bFIeuYRYeLZh2grx+zD8ySK75yMT9cNHN1PK2ZV5o5NZ7W5uadmuUfvlfMOzXTjH2vmHdqpi96tuBkYragvekZROHavNLM4fDVmS6apyfPVp9MzajybtuR2Wrzu2aebZuYvTP1vXJ+dLw498bnwrzS/F6UZwvMK82vOHn6Ma80P5jjeT9Hs0XzTau+1maL5htRfa2ZPdot2xyO6yPzfrpsxPCVY7MPzJdPe7bgeG62wHc+uDavNN/A6atzYkaiL6bXxydmJHqPxmOzbeY7t9jauRnw3vBcmPHuHe5mtHsHu+mhSyEP3v3FxIz1qXesm1dS7yvZxOyxbu7AcF9OTDQn3shMTGQm3rE+NWN96h3r5pXU+8rF1Ix17465Nq80vyDoGRtTs23m4cn3itk2MwXS94rZb928qOHZcL4ofPEvm0Pa01Jigu8/Ey9ufL4+WRQ+L0/MKd3TfmCGqnekzhcrn0fzRezdTjOwfZ7NFzfe8T5ZXLWPSoP9s5ws1j7Pm6Hj2f58uqh85eupCYAvMscmmuarRX1nVRMzM3nIczTNFpe+GlezxdI3LgJzAvKef2aLlc+j2SL2+dXJYumNmTlVeM8UJ4uVz6OTRezzxJwO/VfSxY3P17NF4fPS3DJ5t3++WLaxP+jvvd6//Xb7aZPePn66//q092Xz0dyHHb4xt9SPXUq5/fPz9lur5mHr1+2zSTjr3z5vbj9sHpu/mazVx+32Wf9igvbh8fb7/ddPe4+L+w/v9h9XH7q0yPft42/tXeD7/wcAAP//AwBQSwMEFAAGAAgAAAAhAOmmJbhmBgAAUxsAABMAAAB4bC90aGVtZS90aGVtZTEueG1s7FnNbhs3EL4X6DsQe08s2ZJiGZEDS5biNnFi2EqKHKldapcRd7kgKTu6FcmxQIGiadFLgd56KNoGSIBe0qdxm6JNgbxCh+RKWlpUbCcG+hcdbC334/zPcIa6eu1BytAhEZLyrBVUL1cCRLKQRzSLW8Gdfu/SeoCkwlmEGc9IK5gQGVzbfP+9q3hDJSQlCPZncgO3gkSpfGNlRYawjOVlnpMM3g25SLGCRxGvRAIfAd2UraxWKo2VFNMsQBlOgezt4ZCGBPU1yWBzSrzL4DFTUi+ETBxo0sTZYbDRqKoRciI7TKBDzFoB8In4UZ88UAFiWCp40Qoq5hOsbF5dwRvFJqaW7C3t65lPsa/YEI1WDU8RD2ZMq71a88r2jL4BMLWI63a7nW51Rs8AcBiCplaWMs1ab73antIsgezXRdqdSr1Sc/El+msLMjfb7Xa9WchiiRqQ/VpbwK9XGrWtVQdvQBZfX8DX2ludTsPBG5DFNxbwvSvNRs3FG1DCaDZaQGuH9noF9RlkyNmOF74O8PVKAZ+jIBpm0aVZDHmmlsVaiu9z0QOABjKsaIbUJCdDHEIUd3A6EBRrBniD4NIbuxTKhSXNC8lQ0Fy1gg9zDBkxp/fq+fevnj9Fr54/OX747PjhT8ePHh0//NHScjbu4Cwub3z57Wd/fv0x+uPpNy8ff+HHyzL+1x8++eXnz/1AyKC5RC++fPLbsycvvvr09+8ee+BbAg/K8D5NiUS3yBHa5ynoZgzjSk4G4nw7+gmmzg6cAG0P6a5KHOCtCWY+XJu4xrsroHj4gNfH9x1ZDxIxVtTD+UaSOsBdzlmbC68BbmheJQv3x1nsZy7GZdw+xoc+3h2cOa7tjnOomtOgdGzfSYgj5h7DmcIxyYhC+h0fEeLR7h6ljl13aSi45EOF7lHUxtRrkj4dOIE037RDU/DLxKczuNqxze5d1ObMp/U2OXSRkBCYeYTvE+aY8ToeK5z6SPZxysoGv4lV4hPyYCLCMq4rFXg6JoyjbkSk9O25LUDfktNvYKhXXrfvsknqIoWiIx/Nm5jzMnKbjzoJTnOvzDRLytgP5AhCFKM9rnzwXe5miH4GP+BsqbvvUuK4+/RCcIfGjkjzANFvxqKo2k79TWn2umLMKFTjd8V4ejptwdHkS4mdEyV4Ge5fWHi38TjbIxDriwfPu7r7ru4G//m6uyyXz1pt5wUWmuR5X2y65HRpkzykjB2oCSM3pemTJRwWUQ8WTQNvprjZ0JQn8LUo7g4uFtjsQYKrj6hKDhKcQ49dNSNfLAvSsUQ5lzDbmWUzfJITtM04SaHNNpNhXc8Mth5IrHZ5ZJfXyrPhjIyZFGMzf04ZrWkCZ2W2duXtmFWtVEvN5qpWNaKZUueoNlMZfLioGizOrAldCILeBazcgBFdyw6zCWYk0na3c/PULZr1hbpIJjgihY+03os+qhonTWNlGkYeH+k57xQflbg1Ndm34HYWJ5XZ1Zawm3rvbbw0HW7nXtJ5eyIdWVZOTpaho1bQrK/WAxTivBUMYayFr2kOXpe68cMshruhUAkb9qcmswnXuTeb/rCswk2FtfuCwk4dyIVU21gmNjTMqyIEWGaGcCP/ah3MelEK2Eh/AynW1iEY/jYpwI6ua8lwSEJVdnZpxdxRGEBRSvlYEXGQREdowMZiH4P7daiCPhGVcDthKoJ+gKs0bW3zyi3ORdKVL7AMzq5jlie4KLc6RaeZbOEmj2cymCcrrREPdPPKbpQ7vyom5S9IlXIY/89U0ecJXBesRdoDIdzkCox0vrYCLlTCoQrlCQ17Ai65TO2AaIHrWHgNQQX3yea/IIf6v805S8OkNUx9ap/GSFA4j1QiCNmDsmSi7xRi1eLssiRZQchEVElcmVuxB+SQsL6ugQ19tgcogVA31aQoAwZ3Mv7c5yKDBrFucv6pnY9N5vO2B7o7sC2W3X/GXqRWKvqlo6DpPftMTzUrB6852M951NqKtaDxav3MR20Olz5I/4Hzj4qQ2R8n9IHa5/tQWxH81mDbKwRRfck2HkgXSFseB9A42UUbTJqUbViK7vbC2yi4kS463RlfyNI36XTPaexZc+ayc3Lx9d3n+YxdWNixdbnT9ZgakvZkiur2aDrIGMeYX7XKPzzxwX1w9DZc8Y+ZkvZq/wFc8cGUYX8kgOS3zjVbN/8CAAD//wMAUEsDBBQABgAIAAAAIQAWDE8uFAgAAHVfAAANAAAAeGwvc3R5bGVzLnhtbNRcW4/aOBR+X2n/Q5R3mguEHRBQdWbKqlK3Wqmz0j7sSwiGsZoLSsIUutr/vscOhKTBxHGckMxDCyE5/nzuPsfx7P3Bc5U3FEY48Oeq8U5XFeQ7wRr727n618ty8KAqUWz7a9sNfDRXjyhS3y9+/WUWxUcXfX1FKFaAhB/N1dc43k01LXJekWdH74Id8uGXTRB6dgxfw60W7UJkryPykOdqpq6PNc/GvppQmHoODxHPDr/tdwMn8HZ2jFfYxfGR0lIVz5l+2vpBaK9cgHowRrajHIxxaCqH8DwIvVoYx8NOGETBJn4HdLVgs8EOKsKdaBPNdi6UgLIYJcPSdDM390MoSGmkhegNE/Gpi5m/95ZeHClOsPfjuWqml5Tkl09rkPF4pCqJVJ6CNfBJhz9VOz+cu9PK3+l5/wyO8Efu1k5jLWabwL8MafwG7CGMn37zg+/+kvwGYwIQcttiFv1Q3mwXrhiEiBO4QajEoDCAg17xbQ8ldzzZLl6FmNy2sT3sHpPLJrlAdex0n4dB4hRRMkLVcT6E2HavjqJlCa7IsCfw5kMZeAGihtUE0VHDRMPtaq4u4S/VolSAnDxINYLO/yJ/ocdvKZSITIZNsI+qsASkKeeaIVhZFDhjIkPqUmSLQ0xFsrgMWbxiEa1vEFlPY1AFrCAJ6rMicJnYddM4MCTuFy4sZhAyYxT6S/iinD6/HHfgfH2I7okTpfeV3L0N7aNhUmloya0lD3hojffe7/BYZhD6KGBdBeEaMpBz1BpOAG5ybTFz0SYGtQrx9pX8Hwc7+HcVxDGE6cVsje1t4NsuiUfnJ9hPQiIDOctcTcDAIIl6Yn+NDggiIwRGQgjGEBiCkzidCJ1HI2g4iRKu1mZq5RlXkVz1ibQgAM4hEvUs105RYVXXnIaAx6+QvLPMKKNj/RPN2eWckd+caR2bvrNkEpfaGYUq8cxXVKovgsnHsLsz/Gf9LmF8BzWcX1V+ShskWxwjOWloFPGY2+i0xfX5pv1WDaS88aiu17sL6BaD6F3nV9Gj37w9k8pzRXJehRN2n2kK3AicvF7zcqbCOkpYByVi4eKcjOTzLppVdXatibwPZinGvGZsUlRNG+ZzSdrUuAPkYktTbrumflRwlIycU0pgrbk2bSAsiDmhRiMypxmlti8QBKsOIXW+klxdQ+5AlDWS4HCmKeLlhCpq0x4YzpFu+fhTFR2K8g5y3a+kXv73Jq3MkzbuYZPpEMNuAdIjJW1l8hFaB6ePSRE++QLM0rLUEtpZsvpYiLBy2KQjsGCZTFjp04q927lH0qkmPejk2wcXb30PJZcWM2hFJ1+V1yDEP+BW0sN24HcEGwtg+0SMncwVMuPDhs0r0oC/zqv6oGiphhfTaS9AIkHS7L+GCtCeGZ0w58veW6FwSTd2XFiWZ+Aj7elUZCgxSAF2whaVxtjJlLHyPbR3L+hAlaZM4JPOIzTY9tugoVRiImwZ6bqcmRALNiTBWuSoJjTEU67Cx4vjNoHdebvPYyY95It9c9t7ZWcJG4rO1n0XfJV0lAkWoLfBzGoGxRR9ibq2JXqmboKz4ucmM5ZXVsUhQxU7BqcrlstiV0e0i2WrDUlTih9piXXCbgR4l03975hjZ6TbHihRGecQdjhXYCRgYOi33PElN4A8mbVQYyXx3LTbWLixlKpTICunyTekYoybz7nLFk+ZMJKzE7O7hsKEXJJTC62a5awDmJBLMtcuQoblNp8/kuEzKvl8FpdLco565SgpCHtobOC7eqYGsGu+b5DN/kGGQmdH9YJVhCuEDa6UikWt4NFrUSvkPrWoZbTpRlrCmlnBSdXCAi9f/KQo9chJRldQYy50rAjUXWfJNIoKVjEiFfhTp4qV0pI3F08NpYb7HHIyNiZfxOybpRiwMOuZtywg5jIMllr0L5kkvZyOioy1gC7Y3n37IwZ9RTppi7IgF/SisbaoHHfRyU74XdZDogUHGYs08fY8N6t49JBrQVmlHZ+1mF4HWJ6J9M30uQXSma40y+m2G9rKHAUrAxsWwlm9FZZYQsNyGKSm2156IMrDIspayxt6Lkkuu29S10UjjNG/ZNssaFM9XS8E2XqlDjFy2Y0HuZ6D0epqQNR2ihsA6slEcvmpwMN66ORWaYZiVRpWwOphedRstZtZqTnBrGjKLXnAySY9iJBDuXXc4qRrBdxiXLhnwGU6dLPVyoNwQilW62Xmp2Ixm+njupu1MLvavdhWAkfCdaqQl10bk+bI1V33Bd3qYl2MayYFte7tTHpRsOCSCdlh3/0uDd9UCvG74xVkOqtrbwh1t6HO2lXYqQSPuPmrrpRkLdIrNYyqbnke3snK6NWMigdp+45EHGrrjkIYavuBpvQ1vG4qgxxzKzqIxkNIOb/H9NTo283TYo2h/dh3VnL6cjK8jpx55zn3xnP67rJCzmGdq1/Ii69uxjOv9tiNsX/lbWeguT5c3p+mZ+LG5HRy+mZ1Ogqs09ZoY+/d+CX9ca5ePv9BT08F2zrd9Sd+C2JKYq5ePn8mZ6RC2RzmBa+lfo7ghE/4X9mHeK7++/Hxt8nzx6U5eNAfHwajIbIGE+vxeWCNnh6fn5cT3dSf/oM5kaPcp3CAd40T0umR7rDR3xhNIxfOUQ9Pkz2B/3q5NlczXxL49ARWgJ3FPjHH+gfL0AfLoW4MRmP7YfAwHlqDpWWYz+PR40draWWwW4InqeuaYSRnshPw1jTGHnKxf5bVWULZqyAk+HpjEtpZEtrlvPzF/wAAAP//AwBQSwMEFAAGAAgAAAAhABaYqtUUBwAA+SUAABQAAAB4bC9zaGFyZWRTdHJpbmdzLnhtbIRaTW/bOBC9L7D/gfA9MT9EigwSF2nSAEVbN02ye8pFtZlYWH14JTlb//sdSW6BtedpDwVaU6SGM2/evBn18t2PshBvsWnzurqaqXM5E7Fa1eu8er2a/fF0d+Znou2yap0VdRWvZvvYzt4tfv/tsm07QXur9mq26brtxXzerjaxzNrzehsrWnmpmzLr6J/N67zdNjFbt5sYu7KYayndvMzyaiZW9a7qrmaJCTOxq/K/d/Fm/MUkcra4bPPFZbe4jS9x1YnrKiv2bd6Kh7itm+5y3i0u5/0T41Pjr2JZXxyv3NTltqDXdeI2dlletMcPPOWrv2Inlrvye2xOFpfvxWOXdeQgbun+483Jz3WXFeIp78/8tsuqLu/2x8885l0Uf+ZtTkZlXTw1ueqajC69rM+P18gd5+K+qdc7Wv9Sr2PBPEEOy9+iuKkP53zr9qcHfWWP/8patIz//N9LC3pjsxef6xXrrU+fekfWzcll73ZFIT6Tn2JTiftN3dUnsa3rTtxkuzb+QsHxI8v5NRd3Am11ApXxpPqlv9D3IpbHGz9/uO0duytOTL2Nh0tyQfvYtrso5oSyHq+j+8sta8G3BwrN+uT4a9p2CrOfriGEF9kqlsyN+uj8fOwxNjnhb3nsxT5dL9otnXA1o3xsY/MWZ4v7Jm6zJq7F97049sN1mTfiel1lx8jnj7rZRMojeNKuENdlk5+AlYV/uc2qfZ935fYk0nUXL8TThpiAbCcmEPS3WJDLm7rKV1lR7EWbv1ZkCfGWWEfy2Fp0tdhklBHdJoo2K6N4y4p8TZk5PBRfBorJWpGJMqt2wyHxR1ztOtrbn5Z1uyaepNBdXkQEeN5J88fl3dcH8bjpnT6/u3j+UL3mVaSYVa+HX5+flu/7P6LHoUvVmZbKPx9Y8KGH1gtds0fKs5baPn/J9qL/i3g+EJkwKjVKnIlPPftk1XFgect+4ebiBAiJDcTaWvrb64fjcMhUpsqfOZcGsCS1d2BJ6eDxrgTtcjKFuwIyQwUNd6lUo3cpa9FS4pCFUqUKWUjVDy15j3cF6A2VQPcm2AyVQuMTHC8XcCg98qGU3kBvsPFyhCiCG2dhv2R8woVyWAqOM6NfkkpLxozhXdqgJeMlF5TRwsBhY7RQcVce36XRkvGGQ+/BG/DK3nJBGc0wyEKtNTbecAfa3vNpwgVlWEoSzvP9ktQK7Uqk5Dw/7pJol5YGvct4x7m3P1CnbC4PS5aFzXAvr7mgDLuU4nw4LCUe3cvwpDe8iy7GQHRYUpajyuFdRqN40XkwXlohH2pr0ZUpHdCVCb3QGy7hkmj0vERmKM0y23hluMs4lmBH93p8L4x5x2bKIR0g2LzE8VJcLo9gS6B7lYJZqRQEmw3QvSGFB3q2Vg6elwaFkmgZoldDM4zD70pS5ENjPVzyHuLQYkcph3xI+gs6SrK1cgylg1npLEwHyWqb8UADEeXhkvEpApu2rGYb6QsytlEK8uFEdaByBJmNrebDlUl5gl3asIpo2GVZKTKyKKuIhiWHd4UAMa/Y4jscGHA6BAvB5hyMsmSVw8iicJdJMIsmU2yDeV5jDcDqqBFRsLRpBzPFaAvf5TClaBgv7RNIlQ5j3mtYKyW+MilOhF4JDzQelhttIRFpi9kmgWYYx7YwQ7z4FuaQX5DnPawpNPaDiNJQOlI1h5liJazmDsuDgA9UFlZzm+KagilFK8jzHqqUvtNGsFFQEpsUU6WFHEWlDTKbw2zjDQQAPwcYYGPY9nbkXqhgQf81QNRh7vVsyzmWAOwohTnKG6jZNO5TFISoxoVDmymlNyGxcKbAkk1VFHc3bDd6kAdQERF8UaH3OL+cg5mioFgiQEEGsOykYmQ2zDYJFEtGQ2GmnYVXTrFK4ec2g4VhQsHitoJmhYg3HG7bLWzNSGND96bs3GYwfqJWmgk9j5kt4B5WYbVssA9NCjVAitFrNBRmCW5GFEwiYyBEVUhxmk9oAFjaTAp9SLMZyAB4UqEN7ugV9obDna/DYKPJAqKUABWsTnCUJwqixbmsoGinSEIfelhGSenhuQ3Wohp6Q2vYcZCOhrkcsHutx1UPvwurSiMx22hIKSZlh9WjqIAdvZ6gL53AXMY6ykg4JKRRFSwBCZws6QSnA6562k5RJR7OYEfRfBYWDgyAqRYGx8tjtZzgCYxix+mjFJlIPTwlxr2D9nhS4bF0lBNTYkxEkv3+NdzL4OGMxj4MWCw5aCGNdCB6LdQbxuG5jZ2YzwdIKQE7iuQXVJV4BGfwqIr/DDdQip9oiuHog6gXJRGlF/4kgYuvxDO94GA3OvGZwE9kysT4CDc+Ad7LhIm5DftBbfQ8LL7UE0EABKzZJJ6K6DBxIJxVyom5KPz8Qd9nIObVhJC2cHI7oYjcaSX6z2f7Of3XrcW/AAAA//8DAFBLAwQUAAYACAAAACEA6v6qDJcIAAAQIAAAGAAAAHhsL2RyYXdpbmdzL2RyYXdpbmcxLnhtbORZW4+jRhZ+X2n/A+JxJcZ1owBr3JHtNlGkZDOayWp3H2mM2ygYLKAvM1H++351AUO3PbhXijLanYfpMnDqVJ3Ld7469f6750PhPGZ1k1flwqXviOtkZVpt8/J+4f7jl9gLXadpk3KbFFWZLdzPWeN+d/PXv7x/3tbzp+a2djBB2czxc+Hu2/Y4n82adJ8dkuZddcxKvN1V9SFp8bO+n23r5AlTH4oZI0TOmmOdJdtmn2XtrXnj2vmS/2K2Q5KX7o1eGda6zopiWab7qjaPdnV1MKO0Km7o+5nagRpqAQx+3u1uKGHEF/079Ui/rqunG2Ieq2H3TL3ngtCwf6Ul9NQnfdlz66TPsK6IIs5h0PTzwvVFxMPQnZm5jnlqBuXjhzz9YJec/v3xQ+3k24XLXKdMDrA/3rYPdeZQbDSZY+Yfm9aOnIc6X7i/xTFb+ZtYeDFGniAr4a02IvJixsMNC+I14/J3JU3lPIX5W3j+h21ndipfGf6Qp3XVVLv2XVodZtVul6dZ50i4kYqZNrxe52/c5xFfxbfehmygfUkCL9wsqUfoWkZhtKR+xH9Xu57p1Xd/9S6MR9SerU8wNMZI5jDQj1X6a+OU1XqflPfZsjlmaQuj6sm0L/vP9UQjS94V+THOCzg7maux3e5VMWt2fFulD4esbE3g1lmhDdfs82PjOvU8O9xl8FP9w5bCv8iYFs461nnZXnQUC5eERGzlrX2yhqOCjbeMROAFZBMIIkK6pmvjKDF/aDJsPyluj3nvKfFmTxHjqQvWN5ZRFmraOmvTvRruYLSPMLSR6V9oC5+MqszfHJXfkvnzrkaeJXOYzUHQCyp5ELkOQl5ninKXdv3XUgLB0U1zrJv2+6w6OGoAA2Mt2qLJI2LGrKr7BMF0Woge9mmVFjl8d5u0iRJRX51BiGnQCM6jBmOSdRhggaRHjZCfh42AUCavgw0qOCEqrGDDUEjpi2thg7+EDfaNwgaCf7kkFOXmliw9sb4NvWiziT2+CiMSx9GKhOxPhA2nrtp/5u3+0z45Iq8N+P7BOALI73CkAU5nqrQl86Q47pOfqm2cPzvJAQnhE/zrk+pVNfiGQaZOLbD0gNPl5lXAQokkyAaUU6QFDbgMfCZfoMv5zPlz0KV9qr5OSU60Y0xJGJWRRYprwYUKSrgFK/CVc6SkrU5k6IJmEUoanCBvRIYuwVpIRcDO0aFOX3N0DgkIxcJ1nRb8pcjLXzHu2M8nXUW6cm7Ij+hQLEYaKDY5d3QiOvxbRTM/jtex8L3lehN6YhXFXuT7wqNss1ktpQjW6/ANaKZs0qVG+Xiy0IWKy1BvKUO90HkheCh4MM4LLgXohSkoEpkTGRJ1Krrpgym6GnFsoU3m91sLL/dby0bTqiybvM3+hWPD7lCAtP9t5kRMUuk8OVaLrvevJP49lDBrcPbOaTFK20slWHGvJBSUskklQwkahhGXk0oAu2/cyVCChjycVIGq3KtgINAhn9zIUEQIcEUyqQVZc9ISICDIpJahiPSlDNmkFn+kJQqomNQyFEFw4sAwqUUOtHAeMeFPahmK+JJTzie1BEMtIgj8aYuNRAJJuD+pBQVr4P1QhsHkXoYigvvwzKQWMO5eC4im5NN+GYqIyCdRMJ2RwywO/IBOZ+RQApgo2XS+0GEaM3aVyUYyggVE0tFuwAF6OEv25iiRzNPn0kIcRg6OmQuX6BpzrBp1mBniHdC1+wkwMycRSCnInBA20NwLd+B7nTDQZqiZvUkzYGQozN8kDHQYCutDCMx43bKR9ENh/02akctDYc3yrtaMFB0K61J4tbBhl72rdL/mamFz6u2Fo7cFCRJluG76xhh7GWSjKDNbsKGujtROsXALELKF26KbgWO269wt3DtDG45JqzJERbYaOk84ylsGsV+4tmart4fqMfsFvBLftSpdNBPQ2xixjNNn6cNdnq6yL0OhMAoUN1EtMlW0LKU3E+qqb3oJPvp0F96ZSt9Ze6TirEJFJQxbYtIfK+To1xk3ejSIiA7b17tTVf9qdZSqw4rWFzLAlOZHZnuMRr40mSIEOmV60k6dYQpazjCAqzWCEIbShJMPeh5YNmh1ampgnKRL/mhB516q4Jk0KlMtLLMVbCnCMW24TwG7GhMIGUjTBu33GSgWYZys2cH1+4wwl4E4Lrg6Gg50co4mjTUC2ilkZATDKUzQaa5wtU4uGAnttJhUjqcNQ6n6vYhlwTijGm+7fXLNMMw+NXO4XidFCexaaswno2mZBHGxMUSFzzTqnGJIFU+zIM0jro/aCG1ukyXCx2ZG+6SScmIWJAIpumO4CTDDPoxOzSquT8wAAWmRABp0wel24nEC25qX8DrRdbB7qZiIeaUJxlmFRTkEHcMqrKcUWbAy5itEvAI+3azuwVCnwemYVFaqv6wDrigVRLLQD0yZOzUl9aj9XGQKI4vyY7ZD1xxAadpy+t4kWxe185gAj5M0Rc/S9JqafbLNzGOqOj52eb2EXpqe8NS0tXPbCdSdzOu5DW0xTV5cdaAxu9uhHPQLM+znwsJsH7mT0Jqr8iR8yMuq1vTpxQQFdmU1m++7zq0xjbJS+7yqtp91ux5/cQuCO6r2Z/y3KyoYF/3co+vgbufLy2d1W6wrWA85mejbH1XSjFua9pNSYEBBVynMmxT3uPgqzCdZuf2Q1MlHPC80+8tK7/sVbsC+oMVEe7Mf7Xq7Rdq2s22ovGo1n+n8jB452TZvl83CtQ1o2wwZX1jZnsrLCyugCoDu/IXVpR4NR3pEX209D5tD1N57vdSsEJ6+sS8VSJzdznW9O4VT92AAtvE9mPhGW0DyNkYbhkTeeukHuNzhobeiS+6tmVwR3I8RESz/t+/BLjfn/j8uvnzVto1MiQLj45xSW4v7+2AmgkjY+2B84bPucPJHX369QCSNX+oy/+Y/AAAA//8DAFBLAwQKAAAAAAAAACEA+Vr0IdkLAADZCwAAFAAAAHhsL21lZGlhL2ltYWdlMS5qcGVn/9j/4AAQSkZJRgABAQEA3ADcAAD/2wBDAAIBAQEBAQIBAQECAgICAgQDAgICAgUEBAMEBgUGBgYFBgYGBwkIBgcJBwYGCAsICQoKCgoKBggLDAsKDAkKCgr/2wBDAQICAgICAgUDAwUKBwYHCgoKCgoKCgoKCgoKCgoKCgoKCgoKCgoKCgoKCgoKCgoKCgoKCgoKCgoKCgoKCgoKCgr/wAARCAAkAFgDASIAAhEBAxEB/8QAHwAAAQUBAQEBAQEAAAAAAAAAAAECAwQFBgcICQoL/8QAtRAAAgEDAwIEAwUFBAQAAAF9AQIDAAQRBRIhMUEGE1FhByJxFDKBkaEII0KxwRVS0fAkM2JyggkKFhcYGRolJicoKSo0NTY3ODk6Q0RFRkdISUpTVFVWV1hZWmNkZWZnaGlqc3R1dnd4eXqDhIWGh4iJipKTlJWWl5iZmqKjpKWmp6ipqrKztLW2t7i5usLDxMXGx8jJytLT1NXW19jZ2uHi4+Tl5ufo6erx8vP09fb3+Pn6/8QAHwEAAwEBAQEBAQEBAQAAAAAAAAECAwQFBgcICQoL/8QAtREAAgECBAQDBAcFBAQAAQJ3AAECAxEEBSExBhJBUQdhcRMiMoEIFEKRobHBCSMzUvAVYnLRChYkNOEl8RcYGRomJygpKjU2Nzg5OkNERUZHSElKU1RVVldYWVpjZGVmZ2hpanN0dXZ3eHl6goOEhYaHiImKkpOUlZaXmJmaoqOkpaanqKmqsrO0tba3uLm6wsPExcbHyMnK0tPU1dbX2Nna4uPk5ebn6Onq8vP09fb3+Pn6/9oADAMBAAIRAxEAPwD9/M980Ejpmvl/9of9vS68LeI7rwH8HdNtrq5srhre81W8QvGJlOGSNARuIb5SxOMggA9ax7r46f8ABQv4YJH4w8d/A+XXNB4e7/s2G3mmgj/vGK3lMw4/uxyY7gV8X/rtlNTMKmEw8KlV0tJuEHKMfV3u/kmfTf6p5lHB08TXlCmqmsVOSjKXov8ANo+uKbLKkUbSysqqoyzMcAD1rC8HeOE8Q+BNN8a69psmhi+tUme01J1R4NwyFbPQ/XB9QDxV3WNR8L3GlY1vULM2V0u0/aJlEcqntycEH0719VDEUJwU+ayaT100fk7NfM+dlRqRm4tbO2muq9Nzg/EHxd+Jni/UH0H9nz4e2+pRrHmbxd4kumtdJjYnGyEIGnu2HX5FWLHHnA8DhdV/Y4+PPxWvZNT+O/7dHjlYZOf+Ed+GdvbeHtPiH90SBJrx/qbgfQV7p4j8T6J4T8JX3jLUblF07TNPlvLiaPBCwxoXYjH+yDXyTY/taftD/Hn4oWvg/wCFV3a6Db30xFrE1ukjRxgEl5XYHoB0UD05614ufcXZfw5KjTqJynWdoRik5SendpdV1PeyXIcwziNWdDlhCkrzlLRL8G76PZG74u/4IzfsNfEiHZ8UtA8beKpNuGuPEnxP1u7b/wAfusD8BXmniT/g3Z/YlgZtU+Bfj74nfDTVBzBfeE/GkvyN2JE4cn8GH1r7i8IWevaL4Us7Lxjr6ajqENsov9QEKxLLJj5mCjgD09utWrHW9F1Z2Gmarb3BU4cQTK2364NfSU8wr8sbycW18Lav6btaeR5f1jF05NQm2l1V7fj380flz8UP+Cf/APwW7/ZAsZvG/wCxz+31q/xQsrE+Z/wivit1e9uEH8KreGWGU46jzIif4ecCqP7Mv/ByBq3hL4gx/Ar/AIKSfAa68D61a3C2up+INJsZkjs5P79zZS7pY17lozIOchcV+q1/q+maTGsuq6jDbIzBVaaUIGb0Ge9fLP8AwVF/4Jp/AX/goL8L3tNbfS9F+IVnZsvgzxczBZEcfMLeXHMtuxOCvJXcWXB69McdhakuTE29dE1fubUa9PEWjXhdPqlZ/hufS/gH4geCfil4O0/4gfDrxVY61oeqW4n03VNNuFlguIz/ABKynB5yPYgg8iivwM/4I8ft3fGH/gnN+2I37Hnxvkmt/B+ueKG0TxFol7MSuh6s0ohW8hOdqgyBVkx8rowfqoNFLEYKtRqWjqnszHFYGrh6nLHVPVM+uvG3hLXfhf8AEjXPhT4yhuIdU03UJ2hkvEKnULVpC0V3Gx4lV0ZdzLkK+5ThlIr1z4N/tvfFr4ZyW+leIrr/AISDR4wEa3vW/fxoP7kvXj0bOfbrXtPiP4ufsN/tk6O3hPx7qGnzLZ3ch0vUr5vss0EisV8+0u1IMZOP4XBIOGGCRXyV8T/AOm/DL4n6x4I8LeOo/E2i27xSaPqysjSlHTLQSmMBHdGyN6gBgyggMGJ/lniXLK3CuOqZzkWOTjKesVJcybbunH7SvfdXX4n7dkWPocSYWGV5xhGpRhpJxdmklZp/ZdrbOzPfP2+viLoXxS+Fnw58ZeFruR7HUNWum8t/lKsLZso65+8pyO/PSvn/AFfRNcn+HOk+MJPEkstlF4im0gaTcAskQNqbgSxkt8hyjArjB3A8Ec+5eIf2U/jV4t/Zp8CaT4V8NJcaha6xealfWN5eJbvDFPEQg/eEc9Mr1Gao3v7Gf7R0nwNtPDUXgq0bVYfGx1B7H+2YP+Pb7DJDv3ltud7D5c5xzXrZ3kfEXEWZRxtXCyanhle10vacraXre2j6nm5Rm2S5JgHhKeIiuXEPe1+TmSb9LdTS+BviHWdS/YW+L2iX99JNb6d4d1JbJJGJ8pWsXJUegzzjtk1wP7KNn8Rr74yafD8Mb3T7fURGxmm1LOxbfK+bgBTltvQcZPcda9r+DX7Nnxm8M/svfEz4e+I/DFva634l0m8t9Isf7SikEjvaNGmXUlVBc45PHU8Vnfsd/s0/HP4afGKPxX8QvBEem6fHps0fnf2rbzFnbaAoWJ2PrzjHFdEeG8+ryyFVqM/3N/aPW8Fzq13utEZ/29ktGnnPs6sP3tuRaWl7rvZddTmf25/j94x1v4pap8EdJ1WS00nRFt4r+GFipu55YI58uRyUCyoAvTOc54xynxl8IX37LfxQ0W8+GGv3lnNNoNreGUTZPmncsgP95GK52nI5Ptj1T9tD9i/4l+OPiXJ8avgpDa6lNqdrHF4g8PXFytvLJLEoSK5t5XITd5YEbxuVBCIysMFWwLD9mb9pn9obxzpur/G/wkvhiw0+xgsppXvbeSWSGLP3Fhlk+diT8xIAzkZxijibh7iPEZxjJRoznUqSpuhUW0Ip6q9/dt177k5DnWRUMrwqdSMIQjNVoPecmlZ2t719bdtjD/by8YT+P5vhH4ruIzH/AGn4R1S5mgydokL6d29snHsa6z9lX9lbwH8cPh9pPxK8f67rlxdaXrEiW9ouoZt3t4yMQMjq3yE5J2lTzwQOu1+2t+y78UPiFq3gMfBnwfDfaf4c0TULC4hbUYYGi8x7MxY8xhuGIHzjocetdJ8L/Emj/sM/sg6n8QP2ntTtfDdjoJutQ1NmukmKIT8ka7CRJI+AFRSSSwA54r6PDcN5hivEB18dRdSj7GKcmrxc0kvvvc8SpnWFp8Gxw+Dq8tX2smop+8otv8LWPxb/AOC9Pwz0/Sf+CtWsaP8AC+3/AOJp4mtNCuWtbRfmXUpY0hUAD+JvLif3L570V9df8E5f2GPih+3j+3Trf/BW79qvwbe6F4bvNc/tL4c+GdUTbcXvlKsVnPIh5SCGKNSuQPMkAcfIMuV+8SzGOEjGkleySf8AkfNSzP6rGNJLmaSv6n6FfFH9hH9mr4p+IJfGOoeCpdH1e4ObvUfDd9JYtdH+9MkREczf7Toze9Hwy/Yi+AXwz1uHxPYaHe6pfW7BrabWr9rhYW/vLHxHu9GKkjsRRRXw1bKcqqZhGtPDwc9+Zwi399rnDDMsxhgXSjWmo9lJ2+69j2FVXb92l2iiivaPJG7QTzShQKKKAA8t+FIAN3SiigBrnZE0gHIUmvn3w/8AAX4e/tW+KbH4wftCWlx4ofwzqkx8N+GtSmDaNp86uVFyLQAJNOAOJJ/NKZOzZk0UVtR0TaNqfuxbR9BQRRRRiGKNVVQAqquAB6UUUVkYn//ZUEsDBAoAAAAAAAAAIQDBXHP96hMAAOoTAAAUAAAAeGwvbWVkaWEvaW1hZ2UyLmpwZWf/2P/gABBKRklGAAEBAQCWAJYAAP/bAEMACAYGBwYFCAcHBwkJCAoMFA0MCwsMGRITDxQdGh8eHRocHCAkLicgIiwjHBwoNyksMDE0NDQfJzk9ODI8LjM0Mv/bAEMBCQkJDAsMGA0NGDIhHCEyMjIyMjIyMjIyMjIyMjIyMjIyMjIyMjIyMjIyMjIyMjIyMjIyMjIyMjIyMjIyMjIyMv/AABEIAIwAlgMBIgACEQEDEQH/xAAfAAABBQEBAQEBAQAAAAAAAAAAAQIDBAUGBwgJCgv/xAC1EAACAQMDAgQDBQUEBAAAAX0BAgMABBEFEiExQQYTUWEHInEUMoGRoQgjQrHBFVLR8CQzYnKCCQoWFxgZGiUmJygpKjQ1Njc4OTpDREVGR0hJSlNUVVZXWFlaY2RlZmdoaWpzdHV2d3h5eoOEhYaHiImKkpOUlZaXmJmaoqOkpaanqKmqsrO0tba3uLm6wsPExcbHyMnK0tPU1dbX2Nna4eLj5OXm5+jp6vHy8/T19vf4+fr/xAAfAQADAQEBAQEBAQEBAAAAAAAAAQIDBAUGBwgJCgv/xAC1EQACAQIEBAMEBwUEBAABAncAAQIDEQQFITEGEkFRB2FxEyIygQgUQpGhscEJIzNS8BVictEKFiQ04SXxFxgZGiYnKCkqNTY3ODk6Q0RFRkdISUpTVFVWV1hZWmNkZWZnaGlqc3R1dnd4eXqCg4SFhoeIiYqSk5SVlpeYmZqio6Slpqeoqaqys7S1tre4ubrCw8TFxsfIycrS09TV1tfY2dri4+Tl5ufo6ery8/T19vf4+fr/2gAMAwEAAhEDEQA/APf6KKKACkJpCcVn6lq9ppduZrqUIOwzyaaTewN2L5bFZmo+INO0xf8ASrlVP93OTXP3GralrcLHT8Wlrt5nkOM/SuL1DT7GdfLt4bzUrwHDSk4jz61006CfxEqV9js18byahK0WladLN2Er8J+dNuDr88ZluNVtrRMZKoeRWTZeENSOlqb/AFU2VuBnyoh0HvzW1pPg/Ro2k3XL327HyyPkD6VclCD90RgvM6wlpPFsvJwNu4j+dUYzBHO7N4r1AjGSqhsZ/OvTodD02GMRpZQBV7bBUx02yAP+iQ8/7ArP20ew+VnBWf2m52TWfipjk8RzEg/jk1pLr+u6dKUurOO8iAz5kDc/lV/WbXw7axk38ESFv7owf0rn7Hwzp19NJd6DrU0RPVM5GffpVJxkrtaE9Tc0/wAd6fdymKdJbVwcHzV4/Oulhuop0DxSK6noQa8/ng13S1ddVsYdUtWP3o+GX9KZYjEzz6DK0Mq/fspjj8qHRi1dDcmmek7qUGud0zxNDdS/ZrqNra5HVJO/0rfVsjI6VzSi47lJpklFIDxS1IwooooAQGkY0E4qlqWpQadatPcOFUDj3NNJt2QmytrWt22k2pklYbzwiDqTXGzQYuhrOuyCRm/497Nfmye3FSCARyTeIdV3SebxbWz84J6cetaumWEaquq6yVErEGNXPEeen411RSgjKTbehQTRbrV0N5rUhs7NRlbaNsAL74rT0/xJoSlbW1Vo41O0P5RCn8aXxcjS2dsgJ+zmUeZt4+Wtf7BZTWQiWGPyyuBgClKV1djV07IwvGIh/s+G5lkc228B1VuCuawLa4sI/E1jHoT7EIHnZJAI/wAa2PsTX+manpUh81YciLnOK5m3s9Ii0uKC8ujY6lGxAYA7jWsLctmPfU9bVqUniuFsfFcOmWSwyNdX7j/losdWx47sZoWD2l5Hkc5j6VyujK5XMiPxFa3Fvqqaqlsl3bom14zzj3qx4fbR7lZdUsohAwBWVBwAfcVy9tq91LdXNrpuoKUbol1wRmpNQx4Y0NbX7SGuL18u6njHeunkfKosi2uhpt4k1ZFnv4rU3NikhBUcNj1FW30jTvFFkmqWLNb3RGVkTggjsac+r6VpPhkRpcRyEx4Cqclial8JxjTvD5nuCI1djIdx6CplorxVh9Tn9RW6hMdrrsce88RXsY6fWtPSNcvNKvU07WJUkRlzFcKeGHarsWraX4ouJ9OaEuqjIYjr9K5LXbafw9BLaXjNcWMgIt3PLRnsM1orS92SsxJa3R6okgdQykFT0I71JmvNfA/iiePyNM1LI3L+4dj94V6QpzXHUpuDsaIkooFFZjI2OFJzXD6hH/wkevbpiRpthy+T8rtXTa7etZ6VO6f6wqQnue1cndW1xBptjpFvxPeHdcN7d66KMbamcnrYuaVpx1nV21S7VmtoDi1jJ+X/AHsUmqzRz+IEi1JmjsYQGWNh8rsO5rduIbuw0hIdMjjaWMABWOAQOtYtxq1ler9j1yya3fpvblc+xFVG8ncTsi5pc9xrb3XnRKNP+5FkYLe9ZupW+oaGpKawIbHoEcZYfQ1NYzXehu4eRJtL25jkBGV9qZp+lXHiG6/tDVhm2zmGHPGPUiqXuyu9geqsjL0K01m+E6W8rQWsjljORh3rp7Hwhplq3mTQrdTHkvMNxz+NbsMCQxhI1CqBwBUuKyqV3Ju2hUY2IY7WCJQscKKvoFAFPMKEYKKfwqWisbsqxkXnh3Sr4s01jCZCMeYEAb8xXFa38P5jNFLbO1zDCci3kc4A9q9LNNZeKuFWURnmemL4Yj1ZY59Pkivl+7BJllyPTNaV7Ff603mzRta6Xb5ZojwZAO2BW3rnh2DUlE0SLHdpykgHP0rAjv8AULywu9DeVYtQ+6hboV//AFV0xfP7yJehzvhzXbPTr+8dYpWvHbZDGPu47V6A2lHXNCWDWYUMrDOB/Cap+HvBtppaie5RZrsnJcjpXV7flxUV6sXL3QieXvp0XltpNxGI76yBNpN0JQcgZrsfC2tLq2mgbwZ4jskHvVTxhpxNmNRgjDXFsQ2fVe4ritM8Qiz1+K9tbSW3tLkhZEK4Ut0Jq+V1YaBsexA0VFG29A3YjIoriKOa1+Z59YsLBBu+bzH56AUaJIupaze3bZPkny4yenvVSSUS+JdTuFGTbW+0H0OKf4aeWy8KzXipulYs+09667WhZGNru7JNbi1yK8M1u8j23dIsZA+lZS6ZZa4ypd6jOJM/6mYbWz7VsRa1d2egG/1DY8kjfu0QY69BWJPqCljd6xpywPtLwSL69s4rSnz2sgfLsWTpIe8i0i0mZ7eM7pcnIHtXbRRrFGqKMBRgCuX8DpJJYTXs3LzvnNauuyajHArWEsEWD87zHgCsKrcpchcVbU2RS1zHhrX7jUILr7b5X+jtgyrwrCtXT9bs9TaYWsu8RHDHtWMqcolXRp0VBHcwykiOWNyOoVgak3c9akY+kPSjNLQAwjrXFeL9OmtZ49asI8zwkF8HGRXbVWvLdbmzmibo6Ff0rSlPlkJrQq6PqB1LTYbkpsLjlc9Kuy3UMIJkkRQO5YCuC0aa8GgalZQu6T2zEKR1xWTq9jv0FL99SlecY8yNpPzGK3VBSlvYnmaPQpNW0u+kbT1u4ZJZVPyK2SRXCpYl7fVdLdSWtX86Aeg+tUdNbT/+Eq0l9HDn5AJgcnBxzXQatJ/Z/jaKMIdmoQ7OPXGKuMVTdl1E7vY6jQLz7do1vNk7toBz60VheArh20+6tZG5glIHNFc9SNpMpPQjnxZXPiGYEbim79K1PC8clx4aijnRV3qcBfSsu4j8/wAQ61aSfcktgRx14ra8JyK2gQKpz5Y21pN+4ZxV2VB4cuppIo7u5D2kLbkjAwfxNU/E2mXl+0imNUs7eIsrZ6nFdptqrqMW/TrhB3jI/SohWlzJl8iMzwpH5Xh+245K54rL8Ra1EftGnX2l3LpkCN0zhvQ5Favhi4SfSI1UY2EqRW00av8AeAP1qZStNtjWx5r/AMI7rUmgx5kSK1QFzb5Kkjr8xHWsW0vrq10yS0topI3u34K55HoDXsckKyxtGwyrDBA9KwNW8LpdWtslk/2ea1OYWIyPoa6KeJW0kLlOR0DSjNd+TBJNZanb/M/mfMGH0rWt9Y1fT/E8em3d4l6jgbtqAbfyFNvdI1zSvOv7eUXV7P8ALIyR42D2p3hn+z7ITXl9FPFfIMyyXI459KcnF+9uBv2Piiy1DVjp8AkMigkkrgcda29wAyTXA+H5Fl1PV9WtIftMinEUYbbu+nFGs+L7o6VLBLplxZzyfKGPK+/PFYug3LliVc74EEZBBFIeQa82tPGBtdAgsrXzFvgQoZ1BBGeteh2zM1nGZD85UFuMc1E6Uqe4XucvpW1fFGpo2AGUHFY1wfB8mpSG4mlXLZZWJCk1o6VOj+LdVuTxFCgDt1pniC98PX2lT+RJaNPtyvHOa6E/eJsamnaj4W85TZS2azYwMAA1l+KireLdDH+3kEd6i8M6H4cl021dzBJeNyxD87vpUmvxiTx1oyR5xEuSMdBUpL2gpaIpeAWuU13XYZhhFlyo/Gir3gtornWtcu4mJV5sYPaipqP3hot604s/FdnI3Edynkk+/ap/B58hb2yPBilJAPXBp/jCykuNMW5hGZLVxKAB1wc1jQavHD4jsL2Nh5OoR7G/2WFNe9TFbU9AHSmSLuRlPQjFCnPINOPSuXY0OU0P/iWa3daexIRzvTNdUDXNeJtNnfZqNmxWeDnAH3h6Vp6PqkWp2aTIw34+de4NbVFzJTREXZ2ZqYoxSA0tYliFailtYZ0KSxq6nqGGc1NijFF7AULPR7DT5HltLaOJ3GGKjGaztf0KXWJ7QeYi28b7pFI5NdDTapTkncDjtX8Oy3fiDTzDbKlnCMvIuB07Yre1a8XTtKuJ8gFIztz3OK0HZUBJOAK4e7lfxT4jjtIT/wAS+1YGVh0YjtWik525tkFjDtdU1Dw7pEcn2JZ77UZifLIzuT8KsWup+HdQuWg1vSo9Lmz9+b5AT9T0ra8RLJputafqAiZ7OFSrBFzt9/pT9X1/wzd6Y7XMsM4AyI/4s1re+qRN7E2n+E9BWWK9sDuCncjRzblNZM94W8Y6leOP3Nhb4BPTOM0nhqEeHvC19qxV083dJFC3AUfwjFY+24h8LTTXJH2/WpxhfRScVUU+Z3Je2h0nw6sPI0SW8IIa7kMnI7UV1GmWiWOnQWyLhY0AxRXPOV5NlJFuaISROhGQwwa84GhJHdXWkTschvPs3PGD3FelnpXP+JNEbVLZZYJPKuoDujcD9KKU7OzG+4zwtrBvrU2twwF3bfLKvf610deU3Opvp93Hr1sFFxAPLv4M4JHTIr0PRtWt9Y06K8t33K4z16GnVpte8hmiygggiuR1PSZ9Iuf7T0pWZi2ZIR0IPeuw60hUEVnCbiJq5kaXr9pqACbvLnA+aNuCDWwGB71hap4atb9jNGWguR92WPrWRH/wk+jHaRHfwjvyrYrTkjLWLJu1udpmlz71x3/CaTRfLcaNdIe5Xn+lLH45t5Dt/s+9B7Zj61PspD5kdeTx1qrd3tvZQtLcTpEi85Y4rlpPEWu3jFNP0cop/wCWkx/pUB8F3OsTpc65qEr7TuWCLhQar2VtZMOa5T1DWtS8XXB0/RYilln95dHgEegrs9G0mHSbBLaJRkDLNj7x9as2tnDZwrDDGqIvQAYq0BxUzqXXKthjXjV0IZQQR0NYw8MaKtz9pGnw+aDndjvW2TxWF4h1saVaEQxma6fiOJeSTSg5N2QM53xBqL6l4hg8PW4QWseJLp+ygc4p1jEniDxV5qDdY6cAkZHRmrPWCW0shpMYMmsakd9zMP8AlmDyc9e1dzoukw6PpsdrFztHLdyfWuiUlCJFrs0wox0opw6UVyGgYprDin0YoA4Dxz4NTVbSa7tWaOcL8yqOHHWuZ8Pte+GtOiv7QtNEr7bq3OcqM4yBXsTAY6Vzuq+HlnkN5Zt5V0o6fwv7EV006qtyyJdzS0vVrbVbNLi2lV1YdAelaGa8sNrI16RA/wDZGpKciNTiOWuotvFYtCltrETW0uMeYfuMfrUzodYgpI6p8bTziuM1zxwujXv2c2hYAgM5bHFdXBd291EJIZUkUjqpzXN6v4LttVvTO0vloR90DPzetKkoRl74O7Wh0VjcwahZxXURDRyDcKsiJOu0flUNrAttbRQr91FCjjGcVZBrJvXQpBtHpRtFLmkJ4pAGBSE4FZOreIdN0eEvd3SIeyhuT+FcPqfjTWb+EtZW66fYk4N3cNg49QK0hSlIDpvE/jC20GLy0X7RePwkKHLfiK5iO4n06+WSQm71u9AIiAyIAfX0rP0q3utQeSPQ8vcSn99qVwu7aP8AZrvPDfhW30BZJPMe4u5jmaaT7zH/AArZ2pqwNFnRNJFjEZp8SXcvzSSY7+lbQFAA9KcK5pSbeoAKKWikAUUUUAJgUjD2pTUeT5gHagCjqOk2epwmO6hV/Q45H0NcdqHhjW7XIt54b+yB/wCPW5G7I9ic4NehUhAxWkKriJpHkVzLZ2xBktL7RrkdGTJjB/CtjTNU8TPbKYL2wviD0Y7WxXeywRTqRLGrj0YZrIvPDOlXIMjWwRyPvRnaa2VVS3ROxWbXdXtkdptK8zau791JnP0qpD4x1O4kCr4dugCOCzAVC/h+G3DiK9v1B7Cf/wCtTofD0LfK19qDBhk5uDT5IiTJrjxB4gIzFpUUK5+9NKBiuV8Qa74hitmM2tWUQPSO2OX/AErsYPCGlE5lWafH/PWUtWjbeH9JtyDFYQAg9dgNHNGPQpM8N0vQPEmu3RlhgdmzuWafIB/E16bo/wAPXYxXPiK8kv5k5WF2LRp9AeK7uNFUBVUKPQCpAKzniG9tCrkFtaQ2sYjgiSNB0VFAA/Kp8ClpKw33AdiiiigAooooA//ZUEsDBAoAAAAAAAAAIQDF5APHTwwAAE8MAAATAAAAeGwvbWVkaWEvaW1hZ2UzLlBOR4lQTkcNChoKAAAADUlIRFIAAACCAAAATQgGAAAAar28nAAAAAFzUkdCAK7OHOkAAAAEZ0FNQQAAsY8L/GEFAAAACXBIWXMAAA7DAAAOwwHHb6hkAAAL5ElEQVR4Xu2cP6wNXRTFXyGiIJEQREgUEiISJAoRERKFQhBRKEgkFAqFQqGQKBQKhUJCoVAoFAqFQqFQKBQKhUKhUCgUCoVCobif383syXrz5t/5s899z3dXcuKZuXdmzjn77L32OnvuwmSOOf5ibghzTDE3hDmmmBvCHFMUN4TPnz9Pbt26NXn79m11ZPnhx48fk3fv3k3+/PlTHfn3UdQQGNidO3dOFhYWpu3BgwfVmeUDDGD9+vXT5zty5Mjk9+/f1RlfPHv2bDo2e/funS6W0ihqCC9fvqyNwNrz58+rs8sDu3fvXvR8Dx8+rM74gQVixkc7duxYdaYcihrCtWvXFg0ybe3atZPv379Xn5gtnjx5suT58Are+Pjx45L74plKopghYPVbtmypO3ry5Mn67xs3blSfmh0IARs3bqyfSdvPnz+rT/ng6dOnS+55//796mwZFDOE9+/f1508dOjQ5Nu3b5NVq1ZN/49X8B7sIcBX7PmI1bhn+793zD579mx9L2vXr1+vzpZBMUO4d+9e3UmyBnD58uUlx2YBjFC9FStUn807w9m2bVt9L2snTpyozpZBMUPQUIB3AF++fKm9wpo1a2bGFW7fvl0/G96AMKbHILmesPtogzyWRBFDYGCZ6LYO6sorwdCbwBsoY7dJ51nsGN7MC1+/fq3vQ7NxohE+S6GIISgrxjMoPnz4UJ8r7Q7BnTt36vvv37+/OjqZhgM7fuHChepofmhKjTfiGez/3p5IUcQQdHUx8E1YjGQ1/Pr1qzrqDzIF5QaqaaAu2nE1kNy4efNmfZ+LFy8u8pBtY+WFIoagrLgtP7569Wp9vuQqePz4cX3fHTt2TEOYwtJJDNQLZFD2DCwYTSVLCktFDMFiMGlic7DBixcv6s57uuEmhuRuQpWd90ghlTvRCJPKGThXSuJ2N4RPnz7VHcMztIHOmrGUCg8am7l3m46hSqiHp1J+RL9tkeCd7Pjr16+nx7zhbgjqfvtinoYHpF5v6GrvUjZVZPLIHPT6Bw8erI4uzqTgECXgbgi4eutUn36uyqN3eNAshpXYpV+8evXK9ZlUW9FFwk6kHS/FE9wNwdxcFz9QWHiAyXtCySurrwsar3NnDowFY2LXV/VSwyljUgKuhqAp2BiNQCfIS99XNZOGd+iDTRaeIydUp2gukiaJLCEsuRqCErIxsU5jJtzCAxp/u8irQtM7PEQuqITdFNmACkuEKG+4GoJ2lhRxCBq7x0xSKHTHkzbkDYBmDjknRMlq25azkucSwpKrIWhnccljYDzBIzbevXu3fp4xoQowCfadnNmM9ZPWRqKVMJ46dao66gdXQ7D4GjKp58+frwdgzIoNgebnY0vktGop18pkUdg18VBtolFpwuhmCNqREItW3aFN7YuFkjOk47GKnfIcwkQOqIys+kETmlWM9aixcDMEXUm45LFQA8I75ILqGSFFMLjt3M+jhLWvTI96Sfucd5GvmyHQQevEmzdvqqPjYCuBvYAcQDBSkhiyulRLyCXuaEVSn3StY+hd1+lmCKqahdYj6kpAi0iFlsm1pWp9IITYd+EYqUAfsevR+sYGL2CfixG0uBeGRlYCv8Ezd5XduRmC7fPHqIS6ElJTNsQZXYFj0tgmlOGnbogpBxqaXNJd+ywerXlvzhO6yDDYwmbc4GO8m6EesNngKE24GAJWbjeNSX00dUKLSAFhya6FQQzJ3G3QNJgdwxQoPxiqVGbit27dWn/+9OnT02dholV5DG1tY+piCEqwYnbPNL0am+93gaofu1ZspbRWEaUqnqoYWo0m4Y/tZrIkDIXQqJVTsY3siPHDUxAaMDzCZNticDEElYpZ3TGw6iCIYyxYUZqCxeoSGqtT3jdgwlevXl1f68CBA50v1YxtfJ8UFCUWDsCzspMbyq1cDEFXIelgDAgpdo3YCdR8PWX3UFNa9h76QNyGkJE+44FIOZko5Rkhbd26dfXfxP1Hjx5Nwx1EMGf1kosh8EYvD56ymlUOjnXHGttTXyEzz2J9MnfOdTF8DK2PoA01QgHGTxjCgHVVK9n1qljKbgi4Y3volBdIleSxARMKzf+ZoNitXOIpHmH79u319TZt2lT/Hdt27do15QjwqaH0Wje+vPSE7IaglUYpZVZqUHiYUOhm0VjtgHsyMUwQxscqj2HnCGHckwnEY5DLE96OHz9efyZEPtdKKTIGD2Q3BM2TU2VRCzE0JikEQ9oBaiMDDIsmjmtF89hm7hwuwD3wHH3p6dCOYxfgAmqQOesiDNkNQcWgWKJo0Jw7RKYmjtr3YNW4Xgae1cmkq5GMaYQWDQ3nzp0LZuWaEtNCDVvJc2oK24bshmDSMhYcI94ocNHW+ZAq4jNnztTf27x5cxCJY9Wyp4BBQ9pw6fRDtZEYkUzf3YgJdeppQ2XyMchuCJYXp6RrBibBOt9VscQkofYRc/nMhg0b6u8MNcIB34FPEMf7XC4ewL4XE6fVU5JlhELlZhZZqEcZQlZD0MHCraeCSba0DQMDDAB5OpNHeqiCUV/j+0w6noUwEzOQGuNDvZ3WPsbWWShn6tu1jEFWQ1B2m5q3G44ePVpfMzRXP3z48JSw5iJXOpkh/KdJ9mL3K7QGNMar9CGrIeh2b4rwgWqGIRELx0w8K9VcvEq4udm1kteQFakpdQp3Up6CJ1y2yiKM3B40hFXjpiFT5O5aV9jViNF8FhlX33/QqiiPX0PDOO36IeRV915Sn0szntT0XJHVECwXt3jeB1YJMjIDM9bd79u3r1eFU0k5Ng73AS9g1w9xzbpAUkQ2gGZh1+oi0DHIZggINPaAbekV50nHqB20zKKrIdTghvk83zOShjvscquwajMo/g3N88dAq4tCsiIVq1JJnmZS9JPxyYFshqBbtaxGJgx2jwUzaHaurTHBGA/fU1dvMG2C1kW01P2m1jB0gT4Z6Rsb67VIhxa756FgN9OuFxKi+pDNEC5dulQ/3J49exax5LaGcWAkGMsQ6VES2uXydXDgCl7Q+4zZHtfNszEhcwxUXMpRRwlGGwLpEiycBnuF3PFAkDbImz1YV+OB1d2HQFl3m6rWfGHEIywY6K/da0zRjRLMXIog5Fr1kxyv4o0yBM1fxzbiInwAY0l9OQMXbB1v4wn6fF5hwaAhaMyWcE6iqNBUNsdvN4wyBN3SbWtK/q5cuTK4vx4D5Ql4CIWmnHgcT2guP8bo1Fu27YLGQkMOYThVch5lCKxA2C6DbI3YzoTghnVFctwDXTyB+9lxvEXqgAxByR/ZTB9UcqelekYFc6KaQmxtqCELWdTV6jUR+sNT6grVRfJ3CWg62Kdeall+LlKn0MqlVE0hiyFYnp9jx7ELZBamE6AzsCKapAm3XQIa9/t0AS3izfUCrYG+N8NDSkhONgQVWWDUntCfzifeqqTsVcLVBs0E2l4WMSh3ShWSFBYOubfyo5CXjZtINgR1f575O9DiDgQo3KH9P5ewMgbKS7pejNWUF0+WM2Sy8o2E6tZ4TMGLIdkQtOBijMCSAsKBdRxXqGEhJxEbgm4rt6WzQDMtj5/IQ4tp6jc8UyySDcH26BmQEtDfObAGeSsNDVPNdBZ4yMBN4GV0PIZevulDkiGwEmxleGz7tkHrGK0hVZeG7gI2i3C0rIyWWsQ7BEImr+Kl1F8kGYLuhOVUzfpAbq5xEUPMtQMXAq2Ubu62Kokl118JSDIEXZ05VbMh2EBDwniGWUCFpeZmEoZh57wzqVxIMgQVc3Jsr4aA+3luLo2BFpNa/5v1iSUXSAqSDMEGwkM1Wwlo24nUFLeE5J0L0YagrjF3Re1KAXsuNgYmb6u2sZLGJdoQlBB57/gtV0BSbQzwioQqk8FpIa/pzRrRhmB59Epyfx5QvYAUzv4mW2gTmpYrogyBvNg6/H8NC4auop1S6XQuRBmCWr7XL3isFOi+g7VZaRspiDIE2/FK2eT4l9B8zX4WSmcqgg0BGdM6zMbKHIs33tgISqkLmBWCDUHz5LbNlv8jEJPYa6FSa6WFBEOwIWDtEMRZSbtz+CCKI8zx72FuCHNMMTeEOf5iMvkP0hJkdJcckAQAAAAASUVORK5CYIJQSwMEFAAGAAgAAAAhAP47pTqZCAAADyAAABgAAAB4bC9kcmF3aW5ncy9kcmF3aW5nMi54bWzkWW2Po0YS/n7S/QfEx5NY93uDtZ7I9oyjSMlltZvT3X1kMB6jYLCAedmN8t/v6W7AMONZPCdFWd3th502UF1dVV1PPV39/runQ+49pFWdlcXCp++I76VFUm6z4m7h/+OXTRD6Xt3ExTbOyyJd+J/T2v/u6q9/ef+0reaP9XXlYYKinuPnwt83zXE+m9XJPj3E9bvymBZ4uyurQ9zgZ3U321bxI6Y+5DNGiJrVxyqNt/U+TZtr98Zv54v/i9kOcVb4V3ZlWOs6zfNlkezLyj3aVeXBjZIyv6LvZ8YCM7QCGPy8211RwogU/TvzyL6uyscr4h6bYffMvOeC0LB/ZSXs1Cd96VPjJU/wrogizuHQ5PPClyLiYejP3FzHLHGD4uFDlnxol5z8/eFD5WXbhc98r4gP8D/eNvdV6lEYGs8x84910468+ypb+L9tNmwlbzYi2GAUCLISwepGRMGG8fCG6c2acfW7kaZqnsD9DSL/w7ZzO1UvHH/Ikqqsy13zLikPs3K3y5K0CyTCSMXMOt6u87dNKKGWLIMbtV4FQoarYKX1TaBkGK2U4NdyyX83Vs/s6ru/1goXEWNzGxMMnTPiORz0Y5n8WntFud7HxV26rI9p0sCpdjIby/5zO9HIk7d5dtxkOYIdz824NfeiPessvi6T+0NaNG7jVmluHVfvs2Pte9U8PdymiFP1w5YivsiYBsE6VlnRvBooFi4JidgqWEuyRqDgpWUkdKDJjRZEhHRN1y5QYn5fpzA/zq+PWR8p8eZIERepV7zvPGM8VDdV2iR7M9zBaR/haCfTv7AePjnVuL8+mrjF86ddhTyL53Cbh00vqOI68j1seZspJlw29F9LCWyObppjVTffp+XBMwM4GGuxHo0fsGfcqrpPsJlOC7HDPq2SPEPsruMmNiLmqzMIMQ0a+jxqMKZYhwEtkPSoEfLzsKEJZeoy2KCCE2K2FXwYCqWkuBQ2+HPYYN8obOglXyoV8mC51GEg9DoKVkDiQN9wdqMpJciJPxE2vKps/pk1+0/7+Ii8duD7B+MIIL/DkRo4nZrSFs/j/LiPfyq3m+zJiw9ICEnwr0+qF9XgGwaZKmmBpQecLjcvAhZKFEE2oJwiLagmkqMWPUOX85nz56BL81h+nZK0QPGckshQUnICnhEleQ1cKOUhOwHPOVLSlCcyJM/DmhSCtczmYlQLtaQtfwJROinu1NVH7xCDTyx832tAX/Ks+BXjjvx8skWkq+aO+4gOxDbIAkMm557NQ49/qxzomkbLcE0DQhjAjNysgxXfkCCKVmzDJVlGMnoDmBmfdJlRPJw89ErBZYwILnWbFlyCWPBxWmgtFcEHpp4ozVXkONSp5ib3ruZawGnrbDy/27bocrdtyWhSFkWdNem/cGrYHXJw9r/NvIgpqrxHjysBDuM0v5D491DCrcHbe6fFGG3PlaAC9kpCQSmbVDKUoGEYcTWpBKj7RkuGEjTk4aQKFOVeBYIFmUlDhiJCIKJkUguy5qRFU0bJpJahiJIox2xSixxpiTQVk1qGIjwUkkeTWtRAC+cRE3JSy1BEKk45n9SCjOg9xgWSZNpjIxGtCJeTWlCvBtEPVagnbRmKILURmUktINy9FvBMxafjMhQRkSSRns7IYRZrqel0Rg4lpBSKTecLHaYxYxe5bCQjmCaKjqwBBejhLN67k0Q8T56KFuIw8nDKXPjE1phjWZuzzBDvgJ3dT4CZO4hAykDmhDDMGQp34HuZMNBmKMzepBkwMhS2ZQGeuEwz0GEobM8gFwsj6YfC8k3LRi4PhdWbhF0h7EOl3yTsyGUvbNs1F9vsDr29cPS2TYJEGe2SN+6x55tstMucCe1WNydqL1/4OQjZwm/QzMAp2/duF/6tK97HuDEZYna2GXqPOMm72u7tewJh3h7Kh/QX0Ep815h0sUzAmjFiGafPkvvbLFmlX4ZCYaTBGozt0hQtG+xuQlv1XStBok33yjtX6Ttvj1ScVWiohGNLTMmxQo52nQtjQHVE7LZ9aZ2p+hero5SCcll9IQNM2XaI8xejkVQuU4SgMrKTduocU7ByjgFcrJFxHSq3ncDNtbYp0E9rqYELki35owWde2k2z6RTmelgOVNgUoRT2tBOAb86FwilleuCnhZkWIQLsmUHl9sZYS4HcVxwczIc6OQcPZrWCeimgP6OXhpO4Tad5QoX6+SCkbCdFpOq8bRhqEy7F3tZMM5oS8NdsB3DcHZa5nC5TooS2HXUmCSjaZkCcWn3EBWSWdQ57SFTPN2CLI+4fNdG6HK7LBESxozspIpy4hYktBJslJeOfTidllVcnpg4prhdIo0GW3A6SwJO4FsXMUSd2DrYvTRMxL2yBOOswrwYgo5jFW2kDFloZdxX2PEG+GyvugdDmwanY1JRmvay3VN5YSCShVK7MnfqSdpR8zlPDUbmxcd0h6Y5gNJ15ey1SbrOK+8hBh7HSYKWpWs11ft4m7rH1DR82uX1EnZpdsJTz7adu53AXMm8nNvRFtfjxU0H+rK7HcpBvzDHfl5ZWNtG7iSs5rI4CR+yoqwsfXo2QQ6rWs3u+65x61xjvNQ8rcrtZ9utx19cguCKqvkZ/+3yEs5FO/foe7ja+fL8WdXk6xLeQ07G9vLHlDQXlrr5ZBS4vLdVCvPG+R3uvXL3SVpsP8RV/BHPc8v+0iL4foULsC/oMKEH2a762K63W2TbdW77KS86zWcaP6NHXrrNmmW98Nv+c9sMGd9Xnfowo/sqoAqA7vx91WvNITQHaPTVzvOwN0Tba6/nbSmD8LRtYF/eHYqw3jOqO4VT12AAtvE1mPhWW0C4bVttKAmWao0WkJQsWEmxCdhyfR0ystGE/Y9fg73enPv/uPeSjOC6syWWmmmQoLYW99fBTIYhyprtw2nJcehuAeaPvvt6hkgWv8xd/tV/AAAA//8DAFBLAwQUAAYACAAAACEAOTG1kdsAAADQAQAAIwAAAHhsL3dvcmtzaGVldHMvX3JlbHMvc2hlZXQxLnhtbC5yZWxzrJHNasMwDIDvg76D0b120sMYo04vY9Dr2j2AZyuJWSIbS1vXt593KCylsMtu+kGfPqHt7mue1CcWjokstLoBheRTiDRYeD0+rx9AsTgKbkqEFs7IsOtWd9sXnJzUIR5jZlUpxBZGkfxoDPsRZ8c6ZaTa6VOZndS0DCY7/+4GNJumuTflNwO6BVPtg4WyDxtQx3Oum/9mp76PHp+S/5iR5MYKE4o71csq0pUBxYLWlxpfglZXZTC3bdr/tMklkmA5oEiV4oXVVc9c5a1+i/QjaRZ/6L4BAAD//wMAUEsDBBQABgAIAAAAIQA9BrRkvQAAACsBAAAjAAAAeGwvd29ya3NoZWV0cy9fcmVscy9zaGVldDMueG1sLnJlbHOEj80KwjAQhO+C7xD2blJ7EJGmvYjgVeoDLMn2B9skZOPf25uLoCB429llv5mpmsc8iRtFHr3TsJYFCHLG29H1Gs7tYbUFwQmdxck70vAkhqZeLqoTTZjyEw9jYJEpjjUMKYWdUmwGmpGlD+TypfNxxpRl7FVAc8GeVFkUGxU/GVB/McXRaohHuwbRPkN2/s/2XTca2ntzncmlHxbKRrznZhmJsaekQcr3jt9DKXNkUHWlvirWLwAAAP//AwBQSwMEFAAGAAgAAAAhADl6AmLUAAAANAIAACMAAAB4bC9kcmF3aW5ncy9fcmVscy9kcmF3aW5nMS54bWwucmVsc7yRTUsDMRCG74L/IczdzO4WRKTZXsTSi4jUHzAks9no5oMkFfvvDQhioeKtx5nhfd4HZr359Iv44FxcDAp62YHgoKNxwSp43T/e3IEolYKhJQZWcOQCm/H6av3CC9UWKrNLRTRKKArmWtM9YtEzeyoyJg7tMsXsqbYxW0yk38kyDl13i/k3A8YTptgZBXlnViD2x9Sa/2fHaXKaH6I+eA71TAU637obkLLlqkBK9Gwcfe9X8vlpC3heY7iYxiDfEtu/PPqLefQ/Hnjy6/ELAAD//wMAUEsDBBQABgAIAAAAIQA5egJi1AAAADQCAAAjAAAAeGwvZHJhd2luZ3MvX3JlbHMvZHJhd2luZzIueG1sLnJlbHO8kU1LAzEQhu+C/yHM3czuFkSk2V7E0ouI1B8wJLPZ6OaDJBX77w0IYqHirceZ4X3eB2a9+fSL+OBcXAwKetmB4KCjccEqeN0/3tyBKJWCoSUGVnDkApvx+mr9wgvVFiqzS0U0SigK5lrTPWLRM3sqMiYO7TLF7Km2MVtMpN/JMg5dd4v5NwPGE6bYGQV5Z1Yg9sfUmv9nx2lymh+iPngO9UwFOt+6G5Cy5apASvRsHH3vV/L5aQt4XmO4mMYg3xLbvzz6i3n0Px548uvxCwAA//8DAFBLAwQUAAYACAAAACEAnifDJrsBAAA0FQAAJwAAAHhsL3ByaW50ZXJTZXR0aW5ncy9wcmludGVyU2V0dGluZ3MxLmJpbuyUz0rjUBTGvzajFjcqCG5ciEtpsaXxz1JpolYSU5JUuhKKjRDQpKQRmREXMut5Ax+mj9AHcOXChYgP4Gbmu7GiDEUquBHODeeeP/frSfIjPTZCHCNBjB7tBCmW0GAeIsrilFVVMbCDUSv3Q5u8RWteK+WgrofpuNChn0Ern6dv5TXuFgJ2S7knI7t8rpgbypXP05T/y7Vb9/T3nYz6QXMZAxS14tzd0R/vo7tMZIerWa8veERp8Q0JvH5X4zz6gCLP9veVdhZ9XKKMTej8l5RR4b6NEkyso8paiWZgg1eJmirrJqMyc515hb7GrIq1LLtiR9f0DMtCMwqToKeiRrsbJF74K4Bl+r7pwknCIErbaRhHaDiu727XfbhBLz49z2oMna6KKqjFp3Fix53gJRr9dsU54FA37FcGN9Pd5UVKH2ka7TnnFPT7C/v309TeQn/tWr2/NTxD4a2n0qp8ZehVvkU7VPksyCHmvDnHGWeBmjBNzh01FRpoM+rhgucJOhT/r3R4Fo2prbHHT3TZ3+Mv1P3UREtZkyUEhIAQEAJCQAgIASEgBISAEBACQkAICIFxCPwDAAD//wMAUEsDBBQABgAIAAAAIQCtMdRdpAAAANoAAAAVAAAAeGwvcGVyc29ucy9wZXJzb24ueG1sZM29DsIwDATgHYl3qLyTtAyoqvqzMTHCA0Sp20Rq7Cq2UHl7ihi7nu6+a4ctLcUbs0SmDipTQoHkeYw0d/B63i81FKKORrcwYQcfFBj686ld9w3TI4oWO0HSQVBdG2vFB0xOTIo+s/CkxnOyPE3Ro5U1oxslIGpa7LWsaqvhF+G4txKSCvy9ZjuIvCLtXxPn5FQM5/nglTebXCSw/RcAAP//AwBQSwMEFAAGAAgAAAAhALhpvS5pAQAApQIAABEACAFkb2NQcm9wcy9jb3JlLnhtbCCiBAEooAABAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAIySy27CMBRE95X6D5b3wU6CoLJCkPpgVSokUrXqzrIvIWrsWLZb4O/rJBDoY9FlcmeOZkbO5ntVo0+wrmr0DMcjihFo0chKlzP8XCyiG4yc51ryutEwwwdweJ5fX2XCMNFYWNnGgPUVOBRI2jFhZnjrvWGEOLEFxd0oKHQ4bhqruA+ftiSGi3deAkkonRAFnkvuOWmBkRmI+IiUYkCaD1t3ACkI1KBAe0fiUUzOWg9WuT8N3eVCqSp/MKHTMe4lW4r+OKj3rhqEu91utEu7GCF/TF6Xj+uualTpdisBOM+kYMIC943N2/7msK8zcvGzHbDmzi/D1psK5O0hf9BlpQFs2B4VwBVKUITWTxn5LT25V0HsQeYJTSYRHUd0UtCExSkbJ2+D7yQKmboJ+mAgUSjF+glOl5f07r5Y4DMvLmjK6JTRNPB++NuSPVAdS/yDGE8LStk4ZcnkgngC5F3o7w8r/wIAAP//AwBQSwMEFAAGAAgAAAAhAPgOWk3IAQAABwQAABAACAFkb2NQcm9wcy9hcHAueG1sIKIEASigAAEAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAnFPLbtswELwX6D+wvKQ92JTs1AkMikGQtMihD6N20mPBUiuLiEQK5Eaw+/WlxEZREqdAe9sXZ2dnl/xsV1ekBee1NRlNpwklYJTNtdlm9HrzcXJKiUdpcllZAxndg6dn4vUrvnK2AYcaPAkQxme0RGyWjHlVQi39NKRNyBTW1RKD67bMFoVWcGnVXQ0G2SxJFgx2CCaHfNIMgDQiLlv8X9Dcqo6fv9nsm0BY8POmqbSSGKYUn7Vy1tsCyYedgoqzcZIHdmtQd07jXiScjV2+VrKCiwAsCll54OwhwK9AdqKtpHZe8BaXLSi0jnj9K8h2TMlP6aGjk9FWOi0NBlpdWXR6u2o8OvHdultfAqDnLBTEYG+Oa8e2PhbzviAYfy2MWF9kDTn5Js0W/qXF7HCLjmOcNfR+rMJGYwX+a7GSDg+I8n4sSk8tShJZLk7SySxJT8nbeXoyT9+NxRhkWXc6pS+nIuc/Eg6vjp5iH71ZOW3wx7kDeRDsxRdxxGeL6ncf9HiiwCdtbv11s7GXEuH+iB4H+bqUDvJwd8ORDQF+Fe7HVR3IRdmtL7+veZ7oTv4m/muRLqbJPAnXPIpx9vCDxW8AAAD//wMAUEsBAi0AFAAGAAgAAAAhAHKx13C8AQAAyQcAABMAAAAAAAAAAAAAAAAAAAAAAFtDb250ZW50X1R5cGVzXS54bWxQSwECLQAUAAYACAAAACEAtVUwI/QAAABMAgAACwAAAAAAAAAAAAAAAAD1AwAAX3JlbHMvLnJlbHNQSwECLQAUAAYACAAAACEAeww40TwDAAB6BwAADwAAAAAAAAAAAAAAAAAaBwAAeGwvd29ya2Jvb2sueG1sUEsBAi0AFAAGAAgAAAAhANUTPQknAQAAUQQAABoAAAAAAAAAAAAAAAAAgwoAAHhsL19yZWxzL3dvcmtib29rLnhtbC5yZWxzUEsBAi0AFAAGAAgAAAAhAB8xll73HAAAjLcAABgAAAAAAAAAAAAAAAAA6gwAAHhsL3dvcmtzaGVldHMvc2hlZXQxLnhtbFBLAQItABQABgAIAAAAIQCCYWbM2wwAAOxnAAAYAAAAAAAAAAAAAAAAABcqAAB4bC93b3Jrc2hlZXRzL3NoZWV0Mi54bWxQSwECLQAUAAYACAAAACEAuzdNJaEcAACitgAAGAAAAAAAAAAAAAAAAAAoNwAAeGwvd29ya3NoZWV0cy9zaGVldDMueG1sUEsBAi0AFAAGAAgAAAAhAOmmJbhmBgAAUxsAABMAAAAAAAAAAAAAAAAA/1MAAHhsL3RoZW1lL3RoZW1lMS54bWxQSwECLQAUAAYACAAAACEAFgxPLhQIAAB1XwAADQAAAAAAAAAAAAAAAACWWgAAeGwvc3R5bGVzLnhtbFBLAQItABQABgAIAAAAIQAWmKrVFAcAAPklAAAUAAAAAAAAAAAAAAAAANViAAB4bC9zaGFyZWRTdHJpbmdzLnhtbFBLAQItABQABgAIAAAAIQDq/qoMlwgAABAgAAAYAAAAAAAAAAAAAAAAABtqAAB4bC9kcmF3aW5ncy9kcmF3aW5nMS54bWxQSwECLQAKAAAAAAAAACEA+Vr0IdkLAADZCwAAFAAAAAAAAAAAAAAAAADocgAAeGwvbWVkaWEvaW1hZ2UxLmpwZWdQSwECLQAKAAAAAAAAACEAwVxz/eoTAADqEwAAFAAAAAAAAAAAAAAAAADzfgAAeGwvbWVkaWEvaW1hZ2UyLmpwZWdQSwECLQAKAAAAAAAAACEAxeQDx08MAABPDAAAEwAAAAAAAAAAAAAAAAAPkwAAeGwvbWVkaWEvaW1hZ2UzLlBOR1BLAQItABQABgAIAAAAIQD+O6U6mQgAAA8gAAAYAAAAAAAAAAAAAAAAAI+fAAB4bC9kcmF3aW5ncy9kcmF3aW5nMi54bWxQSwECLQAUAAYACAAAACEAOTG1kdsAAADQAQAAIwAAAAAAAAAAAAAAAABeqAAAeGwvd29ya3NoZWV0cy9fcmVscy9zaGVldDEueG1sLnJlbHNQSwECLQAUAAYACAAAACEAPQa0ZL0AAAArAQAAIwAAAAAAAAAAAAAAAAB6qQAAeGwvd29ya3NoZWV0cy9fcmVscy9zaGVldDMueG1sLnJlbHNQSwECLQAUAAYACAAAACEAOXoCYtQAAAA0AgAAIwAAAAAAAAAAAAAAAAB4qgAAeGwvZHJhd2luZ3MvX3JlbHMvZHJhd2luZzEueG1sLnJlbHNQSwECLQAUAAYACAAAACEAOXoCYtQAAAA0AgAAIwAAAAAAAAAAAAAAAACNqwAAeGwvZHJhd2luZ3MvX3JlbHMvZHJhd2luZzIueG1sLnJlbHNQSwECLQAUAAYACAAAACEAnifDJrsBAAA0FQAAJwAAAAAAAAAAAAAAAACirAAAeGwvcHJpbnRlclNldHRpbmdzL3ByaW50ZXJTZXR0aW5nczEuYmluUEsBAi0AFAAGAAgAAAAhAK0x1F2kAAAA2gAAABUAAAAAAAAAAAAAAAAAoq4AAHhsL3BlcnNvbnMvcGVyc29uLnhtbFBLAQItABQABgAIAAAAIQC4ab0uaQEAAKUCAAARAAAAAAAAAAAAAAAAAHmvAABkb2NQcm9wcy9jb3JlLnhtbFBLAQItABQABgAIAAAAIQD4DlpNyAEAAAcEAAAQAAAAAAAAAAAAAAAAABmyAABkb2NQcm9wcy9hcHAueG1sUEsFBgAAAAAXABcAOQYAABe1AAAAAA=="

def load_template():
    # 1. Try cached template (uploaded by user/admin)
    if os.path.exists(TEMPLATE_CACHE):
        with open(TEMPLATE_CACHE,'rb') as f: return f.read()
    # 2. Fall back to bundled default template
    try:
        return base64.b64decode(_BUNDLED_TPL_B64)
    except: return None

STATE['template'] = load_template()

# ─── HTML Mobile UI ────────────────────────────────────────────────────────────
HTML = r'''<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width,initial-scale=1,maximum-scale=1,user-scalable=no">
<meta name="apple-mobile-web-app-capable" content="yes">
<meta name="apple-mobile-web-app-status-bar-style" content="black-translucent">
<meta name="mobile-web-app-capable" content="yes">
<meta name="theme-color" content="#0F2557">
<title>DAR Portal — SNFOR</title>
<style>
:root{
  --navy:#0F2557;--navy2:#1A3A7C;--accent:#2563EB;--accent2:#1D4ED8;
  --green:#059669;--green-bg:#D1FAE5;--green-text:#065F46;
  --yellow:#D97706;--yellow-bg:#FEF3C7;--yellow-text:#92400E;
  --red:#DC2626;--red-bg:#FEE2E2;--red-text:#991B1B;
  --blue-bg:#EFF6FF;--blue-text:#1E40AF;
  --bg:#F1F5F9;--card:#FFFFFF;--border:#E2E8F0;
  --text:#0F172A;--muted:#64748B;--light:#F8FAFC;
}
*{box-sizing:border-box;margin:0;padding:0;-webkit-tap-highlight-color:transparent;}
body{background:var(--bg);color:var(--text);font-family:-apple-system,BlinkMacSystemFont,'SF Pro Text',Segoe UI,sans-serif;min-height:100vh;}

/* ── SPLASH / HOME ── */
#home-screen{
  min-height:100vh;
  background:linear-gradient(160deg,#0F2557 0%,#1A3A7C 50%,#2563EB 100%);
  display:flex;flex-direction:column;align-items:center;justify-content:center;
  padding:32px 24px;padding-top:calc(32px + env(safe-area-inset-top));
}
.home-logo{
  width:80px;height:80px;background:rgba(255,255,255,0.12);border-radius:20px;
  display:flex;align-items:center;justify-content:center;margin-bottom:16px;
  border:1.5px solid rgba(255,255,255,0.2);backdrop-filter:blur(10px);
}
.home-logo span{font-size:28px;font-weight:800;color:#fff;letter-spacing:-1px;}
.home-company{font-size:11px;font-weight:600;color:rgba(255,255,255,0.6);letter-spacing:.15em;text-transform:uppercase;margin-bottom:6px;}
.home-title{font-size:26px;font-weight:800;color:#fff;text-align:center;margin-bottom:6px;letter-spacing:-.5px;}
.home-subtitle{font-size:13px;color:rgba(255,255,255,0.65);text-align:center;margin-bottom:40px;}
.home-divider{width:40px;height:3px;background:rgba(255,255,255,0.25);border-radius:2px;margin:0 auto 32px;}

.home-cards{width:100%;max-width:380px;display:flex;flex-direction:column;gap:14px;}
.home-card{
  background:rgba(255,255,255,0.08);border:1.5px solid rgba(255,255,255,0.15);
  border-radius:18px;padding:20px;cursor:pointer;
  display:flex;align-items:center;gap:16px;
  transition:all .2s;backdrop-filter:blur(10px);
}
.home-card:active{transform:scale(0.97);background:rgba(255,255,255,0.15);}
.home-card-icon{
  width:52px;height:52px;border-radius:14px;display:flex;align-items:center;
  justify-content:center;font-size:24px;flex-shrink:0;
}
.home-card-icon.history{background:rgba(250,204,21,0.15);}
.home-card-icon.generate{background:rgba(52,211,153,0.15);}
.home-card-body{flex:1;}
.home-card-title{font-size:16px;font-weight:700;color:#fff;margin-bottom:3px;}
.home-card-desc{font-size:12px;color:rgba(255,255,255,0.6);line-height:1.4;}
.home-card-arrow{font-size:18px;color:rgba(255,255,255,0.4);}
.home-badge{
  display:inline-block;background:rgba(250,204,21,0.2);border:1px solid rgba(250,204,21,0.3);
  color:#FCD34D;font-size:10px;font-weight:700;padding:2px 8px;border-radius:20px;margin-top:6px;
}

.home-footer{
  margin-top:40px;font-size:10px;color:rgba(255,255,255,0.35);text-align:center;letter-spacing:.05em;
}

/* ── SCREENS ── */
.screen{display:none;min-height:100vh;}
.screen.active{display:block;}

/* ── HEADER ── */
.hdr{
  background:var(--navy);
  padding:14px 16px 12px;
  padding-top:calc(14px + env(safe-area-inset-top));
  position:sticky;top:0;z-index:100;
  box-shadow:0 2px 12px rgba(15,37,87,0.3);
}
.hdr-row{display:flex;align-items:center;gap:10px;}
.hdr-back{
  width:34px;height:34px;border-radius:10px;
  background:rgba(255,255,255,0.1);border:none;
  color:#fff;font-size:18px;cursor:pointer;
  display:flex;align-items:center;justify-content:center;flex-shrink:0;
}
.hdr-title{color:#fff;font-size:16px;font-weight:700;flex:1;}
.hdr-sub{font-size:10px;color:rgba(255,255,255,0.55);margin-top:1px;}
.hdr-ai{font-size:11px;color:rgba(255,255,255,0.65);display:flex;align-items:center;gap:5px;}
.ai-dot{width:7px;height:7px;border-radius:50%;background:#34D399;}
.ai-dot.off{background:#94A3B8;}

/* ── HISTORY SCREEN ── */
.content{padding:14px;padding-bottom:80px;}

/* Stats bar */
.stats-row{display:grid;grid-template-columns:repeat(4,1fr);gap:8px;margin-bottom:14px;}
.stat-card{background:var(--card);border-radius:12px;padding:10px 8px;text-align:center;box-shadow:0 1px 3px rgba(0,0,0,.06);}
.stat-num{font-size:20px;font-weight:800;color:var(--navy);}
.stat-lbl{font-size:9px;color:var(--muted);font-weight:600;text-transform:uppercase;letter-spacing:.04em;margin-top:2px;}

/* Search & Filter */
.search-bar{
  background:var(--card);border-radius:12px;padding:10px 14px;
  display:flex;align-items:center;gap:10px;margin-bottom:10px;
  box-shadow:0 1px 3px rgba(0,0,0,.06);border:1.5px solid var(--border);
}
.search-bar input{
  flex:1;border:none;outline:none;font-size:14px;color:var(--text);background:transparent;
  font-family:inherit;
}
.search-bar input::placeholder{color:var(--muted);}
.search-icon{font-size:16px;color:var(--muted);}

.filter-row{display:flex;gap:6px;margin-bottom:14px;overflow-x:auto;padding-bottom:2px;}
.filter-row::-webkit-scrollbar{display:none;}
.filter-btn{
  flex-shrink:0;padding:6px 14px;border-radius:20px;border:1.5px solid var(--border);
  background:var(--card);color:var(--muted);font-size:12px;font-weight:600;cursor:pointer;
  font-family:inherit;white-space:nowrap;
}
.filter-btn.active{background:var(--navy);border-color:var(--navy);color:#fff;}

/* Section label */
.sec-lbl{font-size:11px;font-weight:700;color:var(--muted);text-transform:uppercase;letter-spacing:.1em;margin-bottom:8px;display:flex;align-items:center;justify-content:space-between;}

/* Ticket cards */
.ticket-card{
  background:var(--card);border-radius:14px;margin-bottom:10px;
  box-shadow:0 1px 4px rgba(0,0,0,.07);border:1px solid var(--border);
  overflow:hidden;cursor:pointer;transition:all .15s;
}
.ticket-card:active{transform:scale(0.985);}
.tc-header{padding:12px 14px;display:flex;align-items:flex-start;justify-content:space-between;gap:8px;}
.tc-left{flex:1;}
.tc-ticket{font-size:15px;font-weight:800;color:var(--navy);font-family:monospace;}
.tc-location{font-size:13px;color:var(--text);font-weight:600;margin-top:1px;}
.tc-contract{font-size:11px;color:var(--muted);margin-top:2px;}
.tc-right{display:flex;flex-direction:column;align-items:flex-end;gap:4px;}
.status-badge{
  display:inline-block;padding:3px 9px;border-radius:20px;font-size:10px;font-weight:700;
  white-space:nowrap;
}
.s-closed{background:var(--green-bg);color:var(--green-text);}
.s-progress{background:var(--yellow-bg);color:var(--yellow-text);}
.s-docs{background:var(--blue-bg);color:var(--blue-text);}
.tc-date{font-size:10px;color:var(--muted);}

.tc-body{padding:0 14px 12px;border-top:1px solid var(--border);padding-top:10px;}
.tc-detail-row{display:flex;gap:6px;flex-wrap:wrap;margin-top:6px;}
.defect-tag{
  display:inline-block;padding:2px 8px;border-radius:6px;font-size:10px;font-weight:600;
  background:var(--red-bg);color:var(--red-text);
}
.tc-meta{display:grid;grid-template-columns:1fr 1fr;gap:6px;margin-top:8px;}
.tc-meta-item{font-size:11px;}
.tc-meta-lbl{color:var(--muted);font-weight:600;}
.tc-meta-val{color:var(--text);font-weight:500;font-family:monospace;font-size:10px;}

/* Detail modal */
.modal-overlay{
  position:fixed;inset:0;background:rgba(0,0,0,0.5);z-index:200;
  display:flex;align-items:flex-end;justify-content:center;
  opacity:0;pointer-events:none;transition:opacity .25s;
}
.modal-overlay.open{opacity:1;pointer-events:all;}
.modal{
  background:var(--card);border-radius:24px 24px 0 0;width:100%;max-width:500px;
  max-height:88vh;overflow-y:auto;padding:0 0 env(safe-area-inset-bottom,20px);
  transform:translateY(100%);transition:transform .3s cubic-bezier(.34,1.2,.64,1);
}
.modal-overlay.open .modal{transform:translateY(0);}
.modal-handle{width:36px;height:4px;background:var(--border);border-radius:2px;margin:12px auto 0;}
.modal-header{padding:16px 20px;border-bottom:1px solid var(--border);}
.modal-ticket{font-size:22px;font-weight:800;color:var(--navy);font-family:monospace;}
.modal-location{font-size:15px;color:var(--text);font-weight:600;margin-top:2px;}
.modal-body{padding:16px 20px;}
.detail-section{margin-bottom:18px;}
.detail-section-title{
  font-size:10px;font-weight:800;color:var(--muted);text-transform:uppercase;
  letter-spacing:.12em;margin-bottom:10px;display:flex;align-items:center;gap:6px;
}
.detail-section-title::after{content:'';flex:1;height:1px;background:var(--border);}
.detail-grid{display:grid;grid-template-columns:1fr 1fr;gap:10px;}
.detail-item{background:var(--light);border-radius:10px;padding:10px 12px;}
.detail-item.full{grid-column:1/-1;}
.detail-lbl{font-size:10px;font-weight:700;color:var(--muted);text-transform:uppercase;letter-spacing:.06em;margin-bottom:3px;}
.detail-val{font-size:13px;font-weight:600;color:var(--text);}
.detail-val.mono{font-family:monospace;font-size:12px;}
.defect-list{display:flex;gap:6px;flex-wrap:wrap;margin-top:4px;}

/* Upload zone */
.upload-section{background:var(--light);border:2px dashed var(--border);border-radius:14px;padding:20px;text-align:center;cursor:pointer;position:relative;margin-bottom:12px;}
.upload-section input{position:absolute;inset:0;opacity:0;cursor:pointer;}
.upload-icon{font-size:32px;margin-bottom:8px;}
.upload-title{font-size:14px;font-weight:700;color:var(--text);margin-bottom:4px;}
.upload-desc{font-size:12px;color:var(--muted);}
.upload-done{font-size:12px;color:var(--green);font-weight:600;margin-top:8px;font-family:monospace;}

/* Generate screen tabs */
.tab-bar{display:flex;background:rgba(255,255,255,0.1);border-radius:10px;padding:3px;margin-top:10px;gap:3px;}
.tab-btn{
  flex:1;padding:7px;border:none;border-radius:8px;
  font-size:12px;font-weight:600;cursor:pointer;font-family:inherit;
  background:transparent;color:rgba(255,255,255,0.6);transition:all .2s;
}
.tab-btn.active{background:rgba(255,255,255,0.2);color:#fff;}

/* Cards & Fields */
.card{background:var(--card);border-radius:14px;padding:14px;margin-bottom:12px;box-shadow:0 1px 3px rgba(0,0,0,.07);}
.card-title{font-size:13px;font-weight:700;margin-bottom:12px;display:flex;align-items:center;gap:8px;color:var(--navy);}
.field{margin-bottom:10px;}
.lbl{font-size:10px;font-weight:700;color:var(--muted);text-transform:uppercase;letter-spacing:.07em;display:block;margin-bottom:4px;}
input[type=text],input[type=date]{
  width:100%;background:var(--light);border:1.5px solid var(--border);
  border-radius:10px;color:var(--text);font-family:inherit;font-size:15px;
  padding:10px 12px;outline:none;-webkit-appearance:none;
}
input:focus{border-color:var(--accent);background:#fff;}
.auto-filled{border-color:var(--green)!important;background:#F0FDF4!important;}
.g2{display:grid;grid-template-columns:1fr 1fr;gap:10px;}
.ai-badge{display:inline-block;padding:1px 5px;background:var(--green-bg);color:var(--green-text);border-radius:3px;font-size:9px;font-weight:700;margin-left:4px;vertical-align:middle;}
select{
  width:100%;background:var(--light);border:1.5px solid var(--border);
  border-radius:10px;color:var(--text);font-size:15px;padding:10px 12px;
  outline:none;-webkit-appearance:none;font-family:inherit;
  background-image:url("data:image/svg+xml,%3Csvg xmlns='http://www.w3.org/2000/svg' width='12' height='12' viewBox='0 0 24 24' fill='none' stroke='%2364748B' stroke-width='2'%3E%3Cpath d='M6 9l6 6 6-6'/%3E%3C/svg%3E");
  background-repeat:no-repeat;background-position:right 12px center;
}

/* Camera grid */
.cam-grid{display:grid;grid-template-columns:1fr 1fr 1fr;gap:8px;margin-bottom:10px;}
.cam-thumb{position:absolute;inset:0;object-fit:cover;border-radius:10px;}
.cam-check{position:absolute;top:6px;right:6px;background:var(--green);border-radius:50%;width:20px;height:20px;display:flex;align-items:center;justify-content:center;color:#fff;font-size:11px;}

/* Buttons */
.btn-primary{
  width:100%;background:linear-gradient(135deg,var(--accent),var(--accent2));
  color:#fff;border:none;border-radius:12px;padding:16px;
  font-size:16px;font-weight:700;cursor:pointer;
  display:flex;align-items:center;justify-content:center;gap:8px;
  font-family:inherit;box-shadow:0 4px 12px rgba(37,99,235,0.3);
}
.btn-primary:disabled{opacity:.5;box-shadow:none;}
.btn-secondary{
  width:100%;background:var(--blue-bg);color:var(--accent);
  border:1.5px solid var(--accent);border-radius:12px;padding:12px;
  font-size:14px;font-weight:700;cursor:pointer;
  display:flex;align-items:center;justify-content:center;gap:8px;
  font-family:inherit;margin-bottom:8px;
}
.add-unit-btn{
  width:100%;padding:14px;border-radius:12px;
  border:2px dashed var(--accent);background:var(--blue-bg);
  color:var(--accent);font-size:14px;font-weight:700;cursor:pointer;
  display:flex;align-items:center;justify-content:center;gap:8px;margin-bottom:12px;
  font-family:inherit;
}
.del-btn{background:var(--red-bg);color:var(--red-text);border:none;border-radius:6px;padding:4px 8px;font-size:11px;cursor:pointer;}

/* Status boxes */
.sbox{padding:10px 14px;border-radius:10px;font-size:12px;font-family:monospace;line-height:1.5;}
.s-ok{background:var(--green-bg);color:var(--green-text);}
.s-warn{background:var(--yellow-bg);color:var(--yellow-text);}
.s-err{background:var(--red-bg);color:var(--red-text);}
.s-info{background:var(--blue-bg);color:var(--blue-text);}

/* Progress */
.prog{height:5px;background:var(--border);border-radius:3px;overflow:hidden;margin:8px 0;display:none;}
.prog-fill{height:100%;background:var(--accent);transition:width .3s;width:0;}

/* Unit card */
.unit-card{background:var(--light);border-radius:10px;border:1px solid var(--border);padding:10px 12px;margin-bottom:8px;}
.unit-header{display:flex;align-items:center;justify-content:space-between;margin-bottom:8px;}
.unit-sn{font-size:12px;font-weight:700;font-family:monospace;color:var(--accent);}
.tag{display:inline-block;padding:2px 7px;border-radius:5px;font-size:10px;}
.tag-g{background:var(--green-bg);color:var(--green-text);}
.tag-y{background:var(--yellow-bg);color:var(--yellow-text);}
.tag-b{background:var(--blue-bg);color:var(--blue-text);}

/* Spin */
.spin{width:16px;height:16px;border:2px solid rgba(255,255,255,.3);border-top-color:#fff;border-radius:50%;animation:spin .7s linear infinite;display:inline-block;}
@keyframes spin{to{transform:rotate(360deg);}}
.proc-bar{background:var(--blue-bg);border-radius:8px;padding:8px 12px;font-size:11px;color:var(--accent);font-family:monospace;margin-top:6px;}

.tpl-warn{background:var(--red-bg);border:1px solid #FCA5A5;border-radius:10px;padding:10px 14px;font-size:12px;color:var(--red-text);margin-bottom:12px;display:none;}

.safe-bottom{height:calc(env(safe-area-inset-bottom,0px) + 20px);}

/* History record card */
.hist-card{background:var(--card);border-radius:14px;padding:14px;margin-bottom:10px;box-shadow:0 1px 3px rgba(0,0,0,.07);}
</style>
</head>
<body>

<!-- ══════════════ HOME SCREEN ══════════════ -->
<div id="home-screen">
  <div class="home-logo"><span>DAR</span></div>
  <div class="home-company">SNFOR SDN BHD</div>
  <div class="home-title">Defect Analysis Report</div>
  <div class="home-subtitle">SMB Lanterns — Athena Series</div>
  <div class="home-divider"></div>

  <div class="home-cards">
    <div class="home-card" onclick="showScreen('history-screen')">
      <div class="home-card-icon history">📁</div>
      <div class="home-card-body">
        <div class="home-card-title">Ticket History</div>
        <div class="home-card-desc">View all previous DAR records, defect reports &amp; closed tickets</div>
        <div class="home-badge" id="total-badge">Loading...</div>
      </div>
      <div class="home-card-arrow">›</div>
    </div>

    <div class="home-card" onclick="showScreen('generate-screen')">
      <div class="home-card-icon generate">📋</div>
      <div class="home-card-body">
        <div class="home-card-title">Generate DAR Report</div>
        <div class="home-card-desc">Capture site photos, scan QR codes &amp; generate Excel report</div>
        <div class="home-badge" style="background:rgba(52,211,153,0.15);border-color:rgba(52,211,153,0.3);color:#6EE7B7;">New Report</div>
      </div>
      <div class="home-card-arrow">›</div>
    </div>
  </div>

  <div class="home-footer">DAR Portal v2.0 &nbsp;·&nbsp; Internal Use Only</div>
</div>

<!-- ══════════════ HISTORY SCREEN ══════════════ -->
<div id="history-screen" class="screen">
  <div class="hdr">
    <div class="hdr-row">
      <button class="hdr-back" onclick="showHome()">‹</button>
      <div style="flex:1">
        <div class="hdr-title">Ticket History</div>
        <div class="hdr-sub">SMB DAR List — Athena Series</div>
      </div>
      <div style="font-size:11px;color:rgba(255,255,255,0.6)" id="rec-count"></div>
    </div>
  </div>

  <div class="content">
    <!-- Stats -->
    <div class="stats-row">
      <div class="stat-card">
        <div class="stat-num" id="stat-total">—</div>
        <div class="stat-lbl">Total</div>
      </div>
      <div class="stat-card">
        <div class="stat-num" style="color:var(--green)" id="stat-closed">—</div>
        <div class="stat-lbl">Closed</div>
      </div>
      <div class="stat-card">
        <div class="stat-num" style="color:var(--yellow)" id="stat-progress">—</div>
        <div class="stat-lbl">In Progress</div>
      </div>
      <div class="stat-card">
        <div class="stat-num" style="color:var(--blue-text)" id="stat-docs">—</div>
        <div class="stat-lbl">Docs Sub.</div>
      </div>
    </div>

    <!-- Upload new data -->
    <div class="upload-section" id="upload-zone" style="margin-bottom:12px;">
      <input type="file" accept=".xlsx,.xls" onchange="uploadTicketData(this)">
      <div class="upload-icon">📊</div>
      <div class="upload-title">Update Ticket Data</div>
      <div class="upload-desc">Upload new tracking Excel to refresh records</div>
      <div class="upload-done" id="upload-done" style="display:none"></div>
    </div>

    <!-- Search -->
    <div class="search-bar">
      <span class="search-icon">🔍</span>
      <input type="text" id="search-input" placeholder="Search ticket no., location, contract..." oninput="filterTickets()">
    </div>

    <!-- Filter pills -->
    <div class="filter-row">
      <button class="filter-btn active" onclick="setFilter('all',this)">All</button>
      <button class="filter-btn" onclick="setFilter('closed',this)">✅ Closed</button>
      <button class="filter-btn" onclick="setFilter('progress',this)">🔄 In Progress</button>
      <button class="filter-btn" onclick="setFilter('docs',this)">📄 Docs Submitted</button>
    </div>

    <div class="sec-lbl">
      <span id="list-label">All Records</span>
      <span id="list-count" style="font-size:11px;color:var(--muted);font-weight:600;text-transform:none;letter-spacing:0;"></span>
    </div>

    <div id="ticket-list"></div>
    <div class="safe-bottom"></div>
  </div>
</div>

<!-- ══════════════ GENERATE SCREEN ══════════════ -->
<div id="generate-screen" class="screen">
  <div class="hdr">
    <div class="hdr-row">
      <button class="hdr-back" onclick="showHome()">‹</button>
      <div style="flex:1">
        <div class="hdr-title">Generate DAR Report</div>
        <div class="hdr-sub" id="ai-lbl-sub">AI: Checking...</div>
      </div>
      <div class="hdr-ai"><div class="ai-dot off" id="ai-dot"></div></div>
    </div>
    <div class="tab-bar">
      <button class="tab-btn active" id="tab-gen-btn" onclick="switchGenTab('generate')">📋 Generate</button>
      <button class="tab-btn" id="tab-hist-btn" onclick="switchGenTab('genhistory')">📁 My Reports</button>
    </div>
  </div>

  <!-- Generate Tab -->
  <div id="gen-tab" class="content">
    <div id="status-box" class="sbox s-warn" style="margin-bottom:12px;">⚠ Add lanterns and capture photos to begin</div>
    <div class="tpl-warn" id="tpl-warn">⚠ DAR Template not ready — contact supervisor to setup.</div>

    <div style="font-size:11px;font-weight:700;color:var(--muted);text-transform:uppercase;letter-spacing:.1em;margin:4px 0 8px;display:flex;align-items:center;gap:6px;">
      Lantern Photos<span style="flex:1;height:1px;background:var(--border);display:block;"></span>
    </div>

    <div id="units-list"></div>
    <button class="add-unit-btn" onclick="addUnit()"><span style="font-size:20px;">+</span> Add Lantern</button>

    <div style="font-size:11px;font-weight:700;color:var(--muted);text-transform:uppercase;letter-spacing:.1em;margin:4px 0 8px;display:flex;align-items:center;gap:6px;">
      DAR Information<span style="flex:1;height:1px;background:var(--border);display:block;"></span>
    </div>

    <div class="card">
      <div class="g2">
        <div class="field"><label class="lbl">Ticket No</label><input type="text" id="f-ticket" placeholder="e.g. 31731"></div>
        <div class="field"><label class="lbl">Contract No <span class="ai-badge" id="ai-contract" style="display:none">AI</span></label><input type="text" id="f-contract" placeholder="TNB 671/2018"></div>
      </div>
      <div class="g2">
        <div class="field"><label class="lbl">Station</label><input type="text" id="f-station" placeholder="Kuantan"></div>
        <div class="field"><label class="lbl">PIC</label><input type="text" id="f-pic" placeholder="PIC Name"></div>
      </div>
      <div class="g2">
        <div class="field"><label class="lbl">Site Visit Date</label><input type="date" id="f-date"></div>
        <div class="field"><label class="lbl">DO No</label><input type="text" id="f-dono" placeholder="03/06-0003"></div>
      </div>
      <div class="g2">
        <div class="field"><label class="lbl">Total Qty</label><input type="text" id="f-totalqty" placeholder="159 Nos"></div>
        <div class="field"><label class="lbl">Defective Qty</label><input type="text" id="f-defqty" placeholder="4 Nos"></div>
      </div>
      <div class="field"><label class="lbl">Def. Product Model <span class="ai-badge" id="ai-defmodel" style="display:none">AI</span></label><input type="text" id="f-defmodel" placeholder="RL151028B"></div>
      <div class="field"><label class="lbl">New Product Model</label><input type="text" id="f-newmodel" placeholder="Athena RL151026"></div>
      <div class="g2">
        <div class="field"><label class="lbl">Delivery Date <span class="ai-badge" id="ai-deliverydate" style="display:none">AI</span></label><input type="text" id="f-deliverydate" placeholder="08-2023"></div>
        <div class="field"><label class="lbl">Delivery Location</label><input type="text" id="f-delivery" placeholder="KKB Store"></div>
      </div>
      <div class="field">
        <label class="lbl" style="color:var(--navy);font-weight:800;">👷 SNFOR Staff Name</label>
        <input type="text" id="f-staffname" placeholder="Staff name on site" style="border-color:var(--accent);font-weight:600;">
      </div>
    </div>

    <div class="prog" id="prog"><div class="prog-fill" id="pfill"></div></div>
    <button class="btn-primary" id="gen-btn" onclick="generate()">⬇ Generate DAR Excel</button>
    <div id="gen-msg" style="margin-top:8px;"></div>
    <div class="safe-bottom"></div>
  </div>

  <!-- My Reports Tab -->
  <div id="genhistory-tab" class="content" style="display:none;">
    <div style="display:flex;align-items:center;justify-content:space-between;margin-bottom:12px;">
      <div style="font-size:15px;font-weight:800;color:var(--navy);">📁 My Generated Reports</div>
      <div id="hist-count" style="font-size:11px;color:var(--muted);"></div>
    </div>
    <div id="history-list"><div class="sbox s-warn">No reports yet. Generate a DAR first!</div></div>
    <div class="safe-bottom"></div>
  </div>
</div>

<!-- ══════════════ TICKET DETAIL MODAL ══════════════ -->
<div class="modal-overlay" id="modal-overlay" onclick="closeModal(event)">
  <div class="modal" id="modal">
    <div class="modal-handle"></div>
    <div class="modal-header">
      <div style="display:flex;justify-content:space-between;align-items:flex-start;">
        <div>
          <div class="modal-ticket" id="m-ticket"></div>
          <div class="modal-location" id="m-location"></div>
        </div>
        <div id="m-status-badge"></div>
      </div>
    </div>
    <div class="modal-body" id="modal-body"></div>
  </div>
</div>

<script>
// ── TICKET DATA ──
const STATIC_TICKETS = [{"ticket": "36353", "location": "Pasir Mas", "contract": "TNB 2047/2023", "model": "RL151028B", "reported": "22/06/2024", "site_visit": "01/10/2024", "serial": "A-210823-02719", "defects": ["LED Driver"], "reason": "Experience Surge At Site", "status": "Docs Submitted In SMB", "closed_date": "", "do_no": "2215", "new_serial": "A-150925-00008"}, {"ticket": "36646", "location": "Bayan Baru", "contract": "TNB 1211/2023", "model": "RL151028B", "reported": "16/07/2024", "site_visit": "13/08/2025", "serial": "A-150423-03220", "defects": ["LED Driver"], "reason": "Experience Surge At Site", "status": "Closed - Admin Approved", "closed_date": "28/01/2026", "do_no": "2210", "new_serial": "A-150925-00001"}, {"ticket": "36963", "location": "Pasir Mas", "contract": "TNB 2047/2023", "model": "RL151028B", "reported": "16/08/2024", "site_visit": "29/08/2025", "serial": "A-210823-00764", "defects": [], "reason": "N/A", "status": "Closed - Admin Approved", "closed_date": "28/01/2026", "do_no": "N/A", "new_serial": "Non Faulty  (no replacement)"}, {"ticket": "37636", "location": "Pasir Mas", "contract": "TNB 2047/2023", "model": "RL151028B", "reported": "09/10/2024", "site_visit": "29/08/2025", "serial": "A-210823-00692", "defects": [], "reason": "N/A", "status": "Closed - Admin Approved", "closed_date": "26/02/2026", "do_no": "N/A", "new_serial": "Non Faulty  (no replacement)"}, {"ticket": "38599", "location": "Kemaman", "contract": "TNB 2047/2023", "model": "RL151028B", "reported": "17/12/2024", "site_visit": "26/08/2025", "serial": "A-210823-03495", "defects": ["LED Driver"], "reason": "Experience Surge At Site", "status": "Closed - Admin Approved", "closed_date": "26/02/2026", "do_no": "2216", "new_serial": "A-150925-00008"}, {"ticket": "38738", "location": "Pasir Mas", "contract": "TNB 1211/2023", "model": "RL151028B", "reported": "26/12/2024", "site_visit": "29/08/2025", "serial": "A-150423-02881", "defects": [], "reason": "N/A", "status": "Closed - Admin Approved", "closed_date": "26/02/2026", "do_no": "N/A", "new_serial": "Non Faulty  (no replacement)"}, {"ticket": "38811", "location": "Seberang Jaya", "contract": "TNB 2047/2023", "model": "RL151028B", "reported": "01/01/2025", "site_visit": "13/08/2025", "serial": "A-210823-00126", "defects": ["LED Driver"], "reason": "Experience Surge At Site", "status": "Closed - Admin Approved", "closed_date": "26/02/2026", "do_no": "2211", "new_serial": "A-150925-00002"}, {"ticket": "38886", "location": "Bahau", "contract": "TNB 2047/2023", "model": "RL151028B", "reported": "08/01/2025", "site_visit": "07/02/2025", "serial": "A-210823-04078", "defects": [], "reason": "N/A", "status": "Closed - Admin Approved", "closed_date": "10/06/2025", "do_no": "N/A", "new_serial": "Non Faulty  (no replacement)"}, {"ticket": "38923", "location": "Rompin", "contract": "TNB 671/2018", "model": "RL151026", "reported": "11/01/2025", "site_visit": "25/08/2025", "serial": "A-151223-00293", "defects": ["LED Driver"], "reason": "Experience Surge At Site", "status": "Closed - Admin Approved", "closed_date": "26/02/2026", "do_no": "2197", "new_serial": "I-260525-01106"}, {"ticket": "38969", "location": "Kuala Lumpur", "contract": "TNB 671/2018", "model": "RL151026", "reported": "13/01/2025", "site_visit": "13/02/2025", "serial": "C-150524-01173", "defects": ["LED Driver"], "reason": "Experience Surge At Site", "status": "Closed - Admin Approved", "closed_date": "29/08/2025", "do_no": "1878", "new_serial": "F-231224-00049"}, {"ticket": "38969", "location": "Kuala Lumpur", "contract": "TNB 671/2018", "model": "RL151026", "reported": "13/01/2025", "site_visit": "13/02/2025", "serial": "C-150524-00450", "defects": ["SPD"], "reason": "Suspected Experience Surge At Site", "status": "Closed - Admin Approved", "closed_date": "29/08/2025", "do_no": "1878", "new_serial": "F-231224-00050"}, {"ticket": "38998", "location": "Melaka Barat", "contract": "TNB 671/2018", "model": "RL151026", "reported": "14/01/2025", "site_visit": "24/02/2025", "serial": "C-150524-01419", "defects": ["LED Driver"], "reason": "Experience Surge At Site", "status": "Closed - Admin Approved", "closed_date": "29/08/2025", "do_no": "1890", "new_serial": "F-231224-00649"}, {"ticket": "39272", "location": "Temerloh", "contract": "TNB 1211/2023", "model": "RL151028B", "reported": "27/01/2025", "site_visit": "1/27/25", "serial": "A-150423-00399", "defects": [], "reason": "N/A", "status": "Closed - Admin Approved", "closed_date": "22/08/2025", "do_no": "N/A", "new_serial": "Non Faulty  (no replacement)"}, {"ticket": "39280", "location": "Pulau Pinang", "contract": "TNB 671/2018", "model": "RL151026", "reported": "27/01/2025", "site_visit": "13/08/2025", "serial": "B-250324-01795", "defects": [], "reason": "N/A", "status": "Closed - Admin Approved", "closed_date": "29/08/2025", "do_no": "N/A", "new_serial": "Non Faulty  (no replacement)"}, {"ticket": "40373", "location": "Ipoh", "contract": "TNB 671/2018", "model": "RL151026", "reported": "16/04/2025", "site_visit": "16/12/2025", "serial": "C-150524-02683", "defects": [], "reason": "N/A", "status": "Closed - Admin Approved", "closed_date": "26/02/2026", "do_no": "N/A", "new_serial": "Non Faulty  (no replacement)"}, {"ticket": "40713", "location": "Besut", "contract": "TNB 1211/2023", "model": "RL151028B", "reported": "02/05/2025", "site_visit": "29/08/2025", "serial": "A-150925-01736", "defects": ["LED Driver"], "reason": "Experience Surge At Site", "status": "Closed - Admin Approved", "closed_date": "26/02/2026", "do_no": "2192", "new_serial": "A-150925-00021"}, {"ticket": "40838", "location": "Muar", "contract": "TNB 1211/2023", "model": "RL151028B", "reported": "07/05/2025", "site_visit": "16/07/2025", "serial": "A-150423-00274", "defects": [], "reason": "N/A", "status": "Closed - Admin Approved", "closed_date": "16/07/2025", "do_no": "N/A", "new_serial": "Non Faulty  (no replacement)"}, {"ticket": "40838", "location": "Muar", "contract": "TNB 1211/2023", "model": "RL151028B", "reported": "07/05/2025", "site_visit": "16/07/2025", "serial": "A-150423-01629", "defects": [], "reason": "N/A", "status": "Closed - Admin Approved", "closed_date": "16/07/2025", "do_no": "N/A", "new_serial": "Non Faulty  (no replacement)"}, {"ticket": "41400", "location": "Besut", "contract": "TNB 1211/2023", "model": "RL151028B", "reported": "17/06/2025", "site_visit": "29/08/2025", "serial": "A-150423-02780", "defects": [], "reason": "N/A", "status": "Closed - Admin Approved", "closed_date": "26/02/2026", "do_no": "N/A", "new_serial": "Non Faulty  (no replacement)"}, {"ticket": "41422", "location": "Kuala Berang", "contract": "TNB 1211/2023", "model": "RL151028B", "reported": "19/06/2025", "site_visit": "28/08/2025", "serial": "A-150423-01819", "defects": [], "reason": "N/A", "status": "Closed - Admin Approved", "closed_date": "26/02/2026", "do_no": "N/A", "new_serial": "Non Faulty  (no replacement)"}, {"ticket": "41608", "location": "Alor Setar", "contract": "TNB 1211/2023", "model": "RL151028B", "reported": "02/07/2025", "site_visit": "02/07/2025", "serial": "A-150423-03024", "defects": ["SPD"], "reason": "Experience Surge At Site", "status": "Closed - Admin Approved", "closed_date": "26/02/2026", "do_no": "2195", "new_serial": "A-150925-00005"}, {"ticket": "41726", "location": "Taiping", "contract": "TNB 1211/2023", "model": "RL151028B", "reported": "11/07/2025", "site_visit": "13/08/2025", "serial": "A-150423-02104", "defects": [], "reason": "N/A", "status": "Closed - Admin Approved", "closed_date": "26/08/2025", "do_no": "N/A", "new_serial": "Non Faulty  (no replacement)"}, {"ticket": "41733", "location": "Taiping", "contract": "TNB 1211/2023", "model": "RL151028B", "reported": "11/07/2025", "site_visit": "13/08/2025", "serial": "A-150423-00917", "defects": ["LED Driver"], "reason": "Experience Surge At Site", "status": "Closed - Admin Approved", "closed_date": "26/02/2026", "do_no": "2209", "new_serial": "A-150925-00003"}, {"ticket": "41736", "location": "Taiping", "contract": "TNB 1211/2023", "model": "RL151028B", "reported": "11/07/2025", "site_visit": "13/08/2025", "serial": "A-150423-01490", "defects": [], "reason": "N/A", "status": "Closed - Admin Approved", "closed_date": "25/08/2025", "do_no": "N/A", "new_serial": "Non Faulty  (no replacement)"}, {"ticket": "41736", "location": "Taiping", "contract": "TNB 1211/2023", "model": "RL151028B", "reported": "11/07/2025", "site_visit": "13/08/2025", "serial": "A-150423-01493", "defects": [], "reason": "N/A", "status": "Closed - Admin Approved", "closed_date": "25/08/2025", "do_no": "N/A", "new_serial": "Non Faulty  (no replacement)"}, {"ticket": "41736", "location": "Taiping", "contract": "TNB 1211/2023", "model": "RL151028B", "reported": "11/07/2025", "site_visit": "13/08/2025", "serial": "A-150423-02013", "defects": [], "reason": "N/A", "status": "Closed - Admin Approved", "closed_date": "25/08/2025", "do_no": "N/A", "new_serial": "Non Faulty  (no replacement)"}, {"ticket": "41981", "location": "Ipoh", "contract": "TNB 671/2018", "model": "RL151026", "reported": "28/07/2025", "site_visit": "18/08/2025", "serial": "C-150524-02425", "defects": [], "reason": "N/A", "status": "Closed - Admin Approved", "closed_date": "25/08/2025", "do_no": "N/A", "new_serial": "Non Faulty  (no replacement)"}, {"ticket": "42073", "location": "Gerik", "contract": "TNB 1211/2023", "model": "RL151028B", "reported": "29/07/2025", "site_visit": "04/11/2025", "serial": "A-150423-00251", "defects": [], "reason": "N/A", "status": "Closed - Admin Approved", "closed_date": "26/02/2026", "do_no": "N/A", "new_serial": "Non Faulty  (no replacement)"}, {"ticket": "42073", "location": "Gerik", "contract": "TNB 1211/2023", "model": "RL151028B", "reported": "29/07/2025", "site_visit": "04/11/2025", "serial": "A-150423-03084", "defects": [], "reason": "N/A", "status": "Closed - Admin Approved", "closed_date": "26/02/2026", "do_no": "N/A", "new_serial": "Non Faulty  (no replacement)"}, {"ticket": "42073", "location": "Gerik", "contract": "TNB 1211/2023", "model": "RL151028B", "reported": "29/07/2025", "site_visit": "04/11/2025", "serial": "A-150423-03597", "defects": [], "reason": "N/A", "status": "Closed - Admin Approved", "closed_date": "26/02/2026", "do_no": "N/A", "new_serial": "Non Faulty  (no replacement)"}, {"ticket": "42073", "location": "Gerik", "contract": "TNB 1211/2023", "model": "RL151028B", "reported": "29/07/2025", "site_visit": "04/11/2025", "serial": "A-150423-00242", "defects": ["LED Module"], "reason": "Miss-Handling At Site", "status": "Closed - Admin Approved", "closed_date": "26/02/2026", "do_no": "2197", "new_serial": "A-150925-00018"}, {"ticket": "42102", "location": "Kuala Berang", "contract": "TNB 1211/2023", "model": "RL151028B", "reported": "31/07/2025", "site_visit": "28/08/2025", "serial": "A-150423-01820", "defects": [], "reason": "N/A", "status": "Closed - Admin Approved", "closed_date": "26/02/2026", "do_no": "N/A", "new_serial": "Non Faulty  (no replacement)"}, {"ticket": "42390", "location": "Kuala Terengganu", "contract": "TNB 1211/2023", "model": "RL151028B", "reported": "24/08/2025", "site_visit": "28/08/2025", "serial": "A-150423-02548", "defects": ["LED Driver"], "reason": "Suspected Experience Surge At Site", "status": "Closed - Admin Approved", "closed_date": "26/02/2026", "do_no": "2194", "new_serial": "A-150925-00023"}, {"ticket": "42394", "location": "Kuala Terengganu", "contract": "TNB 671/2018", "model": "RL151026", "reported": "24/08/2025", "site_visit": "28/08/2025", "serial": "F-231224-00243", "defects": ["SPD"], "reason": "Suspected Experience Surge At Site", "status": "Closed - Admin Approved", "closed_date": "26/02/2026", "do_no": "2193", "new_serial": "I-260525-01596"}, {"ticket": "42399", "location": "Kuala Terengganu", "contract": "TNB 1211/2023", "model": "RL151028B", "reported": "24/08/2025", "site_visit": "28/08/2025", "serial": "A-150423-03287", "defects": [], "reason": "N/A", "status": "Closed - Admin Approved", "closed_date": "26/02/2026", "do_no": "N/A", "new_serial": "Non Faulty  (no replacement)"}, {"ticket": "42441", "location": "Kuala Terengganu", "contract": "TNB 1211/2023", "model": "RL151028B", "reported": "25/08/2025", "site_visit": "28/08/2025", "serial": "A-150423-03105", "defects": [], "reason": "N/A", "status": "Closed - Admin Approved", "closed_date": "26/02/2026", "do_no": "N/A", "new_serial": "Non Faulty  (no replacement)"}, {"ticket": "42441", "location": "Kuala Terengganu", "contract": "TNB 671/2018", "model": "RL151026", "reported": "25/08/2025", "site_visit": "28/08/2025", "serial": "F-231224-00125", "defects": ["LED Driver"], "reason": "Suspected Experience Surge At Site", "status": "Closed - Admin Approved", "closed_date": "26/02/2026", "do_no": "2206", "new_serial": "I-260525-01602"}, {"ticket": "42455", "location": "Kuala Terengganu", "contract": "TNB 1211/2023", "model": "RL151028B", "reported": "25/08/2025", "site_visit": "28/08/2025", "serial": "A-150423-03108", "defects": ["LED Driver"], "reason": "Suspected Experience Surge At Site", "status": "Closed - Admin Approved", "closed_date": "26/02/2026", "do_no": "2207", "new_serial": "A-150925-00022"}, {"ticket": "42639", "location": "AMBS", "contract": "TNB 1211/2023", "model": "RL151028B", "reported": "09/09/2025", "site_visit": "09/01/2026", "serial": "A-150423-01732", "defects": [], "reason": "N/A", "status": "Closed - Admin Approved", "closed_date": "26/02/2026", "do_no": "N/A", "new_serial": "Non Faulty  (no replacement)"}, {"ticket": "42695", "location": "Klang", "contract": "TNB 1211/2023", "model": "RL151028B", "reported": "15/09/2025", "site_visit": "12/12/2025", "serial": "A-150423-01979", "defects": [], "reason": "N/A", "status": "Closed - Admin Approved", "closed_date": "26/02/2026", "do_no": "N/A", "new_serial": "Non Faulty  (no replacement)"}, {"ticket": "42695", "location": "Klang", "contract": "TNB 1211/2023", "model": "RL151028B", "reported": "15/09/2025", "site_visit": "12/12/2025", "serial": "A-150423-03551", "defects": [], "reason": "N/A", "status": "Closed - Admin Approved", "closed_date": "26/02/2026", "do_no": "N/A", "new_serial": "Non Faulty  (no replacement)"}, {"ticket": "42935", "location": "Besut", "contract": "TNB 1211/2023", "model": "RL151028B", "reported": "02/10/2025", "site_visit": "09/01/2026", "serial": "A-150423-03719", "defects": ["LED Driver"], "reason": "Suspected Experience Surge At Site", "status": "Closed - Admin Approved", "closed_date": "27/02/2026", "do_no": "2220", "new_serial": "A-150925-00007"}, {"ticket": "42935", "location": "Besut", "contract": "TNB 1211/2023", "model": "RL151028B", "reported": "02/10/2025", "site_visit": "09/01/2026", "serial": "A-150423-03673", "defects": ["LED Driver"], "reason": "Suspected Experience Surge At Site", "status": "Closed - Admin Approved", "closed_date": "27/02/2026", "do_no": "2220", "new_serial": "A-150925-00009"}, {"ticket": "43037", "location": "Marang", "contract": "TNB 1211/2023", "model": "RL151028B", "reported": "09/10/2025", "site_visit": "09/01/2026", "serial": "A-150423-03287", "defects": [], "reason": "N/A", "status": "Closed - Admin Approved", "closed_date": "27/02/2026", "do_no": "N/A", "new_serial": "Non Faulty  (no replacement)"}, {"ticket": "43037", "location": "Marang", "contract": "TNB 1211/2023", "model": "RL151028B", "reported": "09/10/2025", "site_visit": "09/01/2026", "serial": "A-150423-03741", "defects": [], "reason": "N/A", "status": "Closed - Admin Approved", "closed_date": "27/02/2026", "do_no": "N/A", "new_serial": "Non Faulty  (no replacement)"}, {"ticket": "43039", "location": "Marang", "contract": "TNB 671/2018", "model": "RL151026", "reported": "09/10/2025", "site_visit": "09/01/2026", "serial": "A-151223-01325", "defects": [], "reason": "N/A", "status": "Closed - Admin Approved", "closed_date": "27/02/2026", "do_no": "N/A", "new_serial": "Non Faulty  (no replacement)"}, {"ticket": "43354", "location": "Pasir Mas", "contract": "TNB 2047/2023", "model": "RL151028B", "reported": "03/11/2025", "site_visit": "25/03/2026", "serial": "A-210823-03722", "defects": ["LED Driver"], "reason": "N/A", "status": "Ticket In Progress", "closed_date": "", "do_no": "0326-0006", "new_serial": "A-150925-00020"}, {"ticket": "43414", "location": "Muar", "contract": "TNB 1211/2023", "model": "RL151028B", "reported": "05/11/2025", "site_visit": "15/12/2025", "serial": "A-150423-00289", "defects": [], "reason": "N/A", "status": "Closed - Admin Approved", "closed_date": "27/02/2026", "do_no": "N/A", "new_serial": "Non Faulty  (no replacement)"}, {"ticket": "43414", "location": "Muar", "contract": "TNB 1211/2023", "model": "RL151028B", "reported": "05/11/2025", "site_visit": "15/12/2025", "serial": "A-150423-00324", "defects": [], "reason": "N/A", "status": "Closed - Admin Approved", "closed_date": "27/02/2026", "do_no": "N/A", "new_serial": "Non Faulty  (no replacement)"}, {"ticket": "43460", "location": "Kuala Lumpur", "contract": "TNB 671/2018", "model": "RL151026", "reported": "11/11/2025", "site_visit": "28/08/2025", "serial": "H-170325-03957", "defects": ["LED Driver"], "reason": "Suspected Experience Surge At Site", "status": "Closed - Admin Approved", "closed_date": "27/02/2026", "do_no": "2213", "new_serial": "I-260525-02112"}, {"ticket": "43460", "location": "Kuala Lumpur", "contract": "TNB 671/2018", "model": "RL151026", "reported": "11/11/2025", "site_visit": "28/08/2025", "serial": "H-170325-03788", "defects": [], "reason": "N/A", "status": "Closed - Admin Approved", "closed_date": "27/02/2026", "do_no": "N/A", "new_serial": "Non Faulty  (no replacement)"}, {"ticket": "43460", "location": "Kuala Lumpur", "contract": "TNB 671/2018", "model": "RL151026", "reported": "11/11/2025", "site_visit": "28/08/2025", "serial": "H-170325-00740", "defects": [], "reason": "N/A", "status": "Closed - Admin Approved", "closed_date": "27/02/2026", "do_no": "N/A", "new_serial": "Non Faulty  (no replacement)"}, {"ticket": "43559", "location": "Kota Tinggi", "contract": "TNB 1211/2023", "model": "RL151028B", "reported": "16/11/2025", "site_visit": "09/01/2026", "serial": "A-150423-01063", "defects": [], "reason": "N/A", "status": "Closed - Admin Approved", "closed_date": "27/02/2026", "do_no": "N/A", "new_serial": "Non Faulty  (no replacement)"}, {"ticket": "43559", "location": "Kota Tinggi", "contract": "TNB 1211/2023", "model": "RL151028B", "reported": "16/11/2025", "site_visit": "09/01/2026", "serial": "A-150423-01060", "defects": [], "reason": "N/A", "status": "Closed - Admin Approved", "closed_date": "27/02/2026", "do_no": "N/A", "new_serial": "Non Faulty  (no replacement)"}, {"ticket": "43636", "location": "Kuala Berang", "contract": "TNB 1211/2023", "model": "RL151028B", "reported": "21/11/2025", "site_visit": "09/01/2026", "serial": "A-150423-01817", "defects": [], "reason": "N/A", "status": "Closed - Admin Approved", "closed_date": "27/02/2026", "do_no": "N/A", "new_serial": "Non Faulty  (no replacement)"}, {"ticket": "43715", "location": "Pulau Pinang", "contract": "TNB 1211/2023", "model": "RL151028B", "reported": "26/11/2025", "site_visit": "29/01/2026", "serial": "A-150423-03465", "defects": [], "reason": "N/A", "status": "Docs Submitted In SMB", "closed_date": "", "do_no": "N/A", "new_serial": "Non Faulty  (no replacement)"}, {"ticket": "43777", "location": "Melaka Barat", "contract": "TNB 1211/2023", "model": "RL151028B", "reported": "03/12/2025", "site_visit": "15/12/2025", "serial": "A-150423-03350", "defects": ["LED Driver"], "reason": "Suspected Experience Surge At Site", "status": "Closed - Admin Approved", "closed_date": "16/03/2026", "do_no": "2208", "new_serial": "A-150925-00006"}, {"ticket": "43813", "location": "Pasir Mas", "contract": "TNB 1211/2023", "model": "RL151028B", "reported": "05/12/2025", "site_visit": "09/01/2026", "serial": "A-150423-02700", "defects": [], "reason": "N/A", "status": "Closed - Admin Approved", "closed_date": "27/02/2026", "do_no": "N/A", "new_serial": "Non Faulty  (no replacement)"}, {"ticket": "43825", "location": "Alor Setar", "contract": "TNB 1211/2023", "model": "RL151028B", "reported": "05/12/2025", "site_visit": "29/01/2026", "serial": "A-150423-02754", "defects": [], "reason": "N/A", "status": "Docs Submitted In SMB", "closed_date": "", "do_no": "N/A", "new_serial": "Non Faulty  (no replacement)"}, {"ticket": "43854", "location": "Besut", "contract": "TNB 1211/2023", "model": "RL151028B", "reported": "08/12/2025", "site_visit": "09/01/2026", "serial": "A-150423-01730", "defects": [], "reason": "N/A", "status": "Closed - Admin Approved", "closed_date": "27/02/2026", "do_no": "N/A", "new_serial": "Non Faulty  (no replacement)"}, {"ticket": "43854", "location": "Besut", "contract": "TNB 1211/2023", "model": "RL151028B", "reported": "08/12/2025", "site_visit": "09/01/2026", "serial": "A-150423-01736", "defects": [], "reason": "N/A", "status": "Closed - Admin Approved", "closed_date": "27/02/2026", "do_no": "N/A", "new_serial": "Non Faulty  (no replacement)"}, {"ticket": "43886", "location": "Kuala Terengganu", "contract": "TNB 671/2018", "model": "RL151026", "reported": "12/12/2025", "site_visit": "09/01/2026", "serial": "B-250324-03885", "defects": ["LED Driver"], "reason": "Suspected Experience Surge At Site", "status": "Closed - Admin Approved", "closed_date": "16/03/2026", "do_no": "2249", "new_serial": "J-151025-03347"}, {"ticket": "43886", "location": "Kuala Terengganu", "contract": "TNB 671/2018", "model": "RL151026", "reported": "12/12/2025", "site_visit": "09/01/2026", "serial": "B-250324-02036", "defects": ["LED Driver"], "reason": "Suspected Experience Surge At Site", "status": "Closed - Admin Approved", "closed_date": "16/03/2026", "do_no": "2249", "new_serial": "J-151025-02903"}, {"ticket": "43893", "location": "Kuala Terengganu", "contract": "TNB 671/2018", "model": "RL151026", "reported": "12/12/2025", "site_visit": "09/01/2026", "serial": "A-151223-01228", "defects": ["LED Driver"], "reason": "Suspected Experience Surge At Site", "status": "Closed - Admin Approved", "closed_date": "16/03/2026", "do_no": "2250", "new_serial": "J-151025-2106"}, {"ticket": "43902", "location": "Kuala Terengganu", "contract": "TNB 1211/2023", "model": "RL151028B", "reported": "12/12/2025", "site_visit": "09/01/2026", "serial": "A-151223-02488", "defects": [], "reason": "N/A", "status": "Closed - Admin Approved", "closed_date": "27/02/2026", "do_no": "N/A", "new_serial": "Non Faulty  (no replacement)"}, {"ticket": "44142", "location": "Kuala Berang", "contract": "TNB 1211/2023", "model": "RL151028B", "reported": "04/01/2026", "site_visit": "09/01/2026", "serial": "A-150423-01820", "defects": ["LED Driver"], "reason": "Suspected Experience Surge At Site", "status": "Closed - Admin Approved", "closed_date": "16/03/2026", "do_no": "2219", "new_serial": "A-150925-00013"}, {"ticket": "44142", "location": "Kuala Berang", "contract": "TNB 1211/2023", "model": "RL151028B", "reported": "04/01/2026", "site_visit": "09/01/2026", "serial": "A-150423-01819", "defects": ["LED Driver"], "reason": "Suspected Experience Surge At Site", "status": "Closed - Admin Approved", "closed_date": "16/03/2026", "do_no": "2219", "new_serial": "A-150925-00017"}, {"ticket": "44211", "location": "Besut", "contract": "TNB 1211/2023", "model": "RL151028B", "reported": "05/01/2026", "site_visit": "09/01/2026", "serial": "A-150423-01739", "defects": [], "reason": "N/A", "status": "Closed - Admin Approved", "closed_date": "26/02/2026", "do_no": "N/A", "new_serial": "Non Faulty  (no replacement)"}, {"ticket": "44211", "location": "Besut", "contract": "TNB 1211/2023", "model": "RL151028B", "reported": "05/01/2026", "site_visit": "09/01/2026", "serial": "A-150423-01745", "defects": [], "reason": "N/A", "status": "Closed - Admin Approved", "closed_date": "26/02/2026", "do_no": "N/A", "new_serial": "Non Faulty  (no replacement)"}, {"ticket": "44211", "location": "Besut", "contract": "TNB 1211/2023", "model": "RL151028B", "reported": "05/01/2026", "site_visit": "09/01/2026", "serial": "A-150423-01750", "defects": [], "reason": "N/A", "status": "Closed - Admin Approved", "closed_date": "26/02/2026", "do_no": "N/A", "new_serial": "Non Faulty  (no replacement)"}, {"ticket": "44211", "location": "Besut", "contract": "TNB 1211/2023", "model": "RL151028B", "reported": "05/01/2026", "site_visit": "09/01/2026", "serial": "A-150423-02484", "defects": [], "reason": "N/A", "status": "Closed - Admin Approved", "closed_date": "26/02/2026", "do_no": "N/A", "new_serial": "Non Faulty  (no replacement)"}, {"ticket": "44211", "location": "Besut", "contract": "TNB 1211/2023", "model": "RL151028B", "reported": "05/01/2026", "site_visit": "09/01/2026", "serial": "A-150423-02486", "defects": [], "reason": "N/A", "status": "Closed - Admin Approved", "closed_date": "26/02/2026", "do_no": "N/A", "new_serial": "Non Faulty  (no replacement)"}, {"ticket": "44211", "location": "Besut", "contract": "TNB 1211/2023", "model": "RL151028B", "reported": "05/01/2026", "site_visit": "09/01/2026", "serial": "A-150423-02780", "defects": [], "reason": "N/A", "status": "Closed - Admin Approved", "closed_date": "26/02/2026", "do_no": "N/A", "new_serial": "Non Faulty  (no replacement)"}, {"ticket": "44282", "location": "Marang", "contract": "TNB 1211/2023", "model": "RL151028B", "reported": "06/01/2026", "site_visit": "09/01/2026", "serial": "A-150423-02382", "defects": [], "reason": "N/A", "status": "Closed - Admin Approved", "closed_date": "12/03/2026", "do_no": "N/A", "new_serial": "Non Faulty  (no replacement)"}, {"ticket": "44282", "location": "Marang", "contract": "TNB 1211/2023", "model": "RL151028B", "reported": "06/01/2026", "site_visit": "09/01/2026", "serial": "A-150423-02467", "defects": [], "reason": "N/A", "status": "Closed - Admin Approved", "closed_date": "12/03/2026", "do_no": "N/A", "new_serial": "Non Faulty  (no replacement)"}, {"ticket": "44282", "location": "Marang", "contract": "TNB 1211/2023", "model": "RL151028B", "reported": "06/01/2026", "site_visit": "09/01/2026", "serial": "A-150423-03161", "defects": [], "reason": "N/A", "status": "Closed - Admin Approved", "closed_date": "12/03/2026", "do_no": "N/A", "new_serial": "Non Faulty  (no replacement)"}, {"ticket": "44282", "location": "Marang", "contract": "TNB 1211/2023", "model": "RL151028B", "reported": "06/01/2026", "site_visit": "09/01/2026", "serial": "A-150423-03324", "defects": [], "reason": "N/A", "status": "Closed - Admin Approved", "closed_date": "12/03/2026", "do_no": "N/A", "new_serial": "Non Faulty  (no replacement)"}, {"ticket": "44282", "location": "Marang", "contract": "TNB 671/2018", "model": "RL151026", "reported": "06/01/2026", "site_visit": "09/01/2026", "serial": "A-151223-01961", "defects": ["SPD"], "reason": "Suspected Experience Surge At Site", "status": "Closed - Admin Approved", "closed_date": "12/03/2026", "do_no": "2248", "new_serial": "J-151025-03372"}, {"ticket": "44282", "location": "Marang", "contract": "TNB 671/2018", "model": "RL151026", "reported": "06/01/2026", "site_visit": "09/01/2026", "serial": "F-231224-00008", "defects": ["LED Driver"], "reason": "Suspected Experience Surge At Site", "status": "Closed - Admin Approved", "closed_date": "12/03/2026", "do_no": "2248", "new_serial": "J-151025-02920"}, {"ticket": "44282", "location": "Marang", "contract": "TNB 671/2018", "model": "RL151026", "reported": "06/01/2026", "site_visit": "09/01/2026", "serial": "F-231224-00269", "defects": [], "reason": "N/A", "status": "Closed - Admin Approved", "closed_date": "12/03/2026", "do_no": "N/A", "new_serial": "Non Faulty  (no replacement)"}, {"ticket": "44282", "location": "Marang", "contract": "TNB 671/2018", "model": "RL151026", "reported": "06/01/2026", "site_visit": "09/01/2026", "serial": "F-231224-00501", "defects": [], "reason": "N/A", "status": "Closed - Admin Approved", "closed_date": "12/03/2026", "do_no": "N/A", "new_serial": "Non Faulty  (no replacement)"}, {"ticket": "44806", "location": "Kuantan", "contract": "TNB 671/2018", "model": "RL151026", "reported": "14/02/2026", "site_visit": "09/03/2026", "serial": "A-151223-01902", "defects": [], "reason": "Suspected Experience Surge At Site", "status": "Docs Submitted In SMB", "closed_date": "", "do_no": "03/26-0002", "new_serial": "J-151025-00159"}, {"ticket": "44984", "location": "Muar", "contract": "TNB 1211/2023", "model": "RL151028B", "reported": "27/02/2026", "site_visit": "30/03/2026", "serial": "", "defects": [], "reason": "N/A", "status": "In Progress", "closed_date": "", "do_no": "", "new_serial": ""}, {"ticket": "44984", "location": "Muar", "contract": "TNB 1211/2023", "model": "RL151028B", "reported": "27/02/2026", "site_visit": "30/03/2026", "serial": "", "defects": [], "reason": "N/A", "status": "In Progress", "closed_date": "", "do_no": "", "new_serial": ""}, {"ticket": "44984", "location": "Muar", "contract": "TNB 1211/2023", "model": "RL151028B", "reported": "27/02/2026", "site_visit": "30/03/2026", "serial": "", "defects": [], "reason": "N/A", "status": "In Progress", "closed_date": "", "do_no": "", "new_serial": ""}, {"ticket": "44984", "location": "Muar", "contract": "TNB 1211/2023", "model": "RL151028B", "reported": "27/02/2026", "site_visit": "30/03/2026", "serial": "", "defects": [], "reason": "N/A", "status": "In Progress", "closed_date": "", "do_no": "", "new_serial": ""}, {"ticket": "44984", "location": "Muar", "contract": "TNB 1211/2023", "model": "RL151028B", "reported": "27/02/2026", "site_visit": "30/03/2026", "serial": "", "defects": [], "reason": "N/A", "status": "In Progress", "closed_date": "", "do_no": "", "new_serial": ""}, {"ticket": "TNB Conract", "location": "Total Supply Qty", "contract": "Defective Component", "model": "", "reported": "", "site_visit": "", "serial": "Total Reported", "defects": [], "reason": "N/A", "status": "In Progress", "closed_date": "", "do_no": "", "new_serial": ""}, {"ticket": "TNB 671/2018", "location": "30170", "contract": "", "model": "", "reported": "9", "site_visit": "", "serial": "21", "defects": [], "reason": "N/A", "status": "In Progress", "closed_date": "", "do_no": "", "new_serial": ""}, {"ticket": "TNB 2047/2023", "location": "5000", "contract": "", "model": "", "reported": "4", "site_visit": "", "serial": "7", "defects": [], "reason": "N/A", "status": "In Progress", "closed_date": "", "do_no": "", "new_serial": ""}, {"ticket": "TNB 1211/2023", "location": "3775", "contract": "", "model": "1", "reported": "10", "site_visit": "", "serial": "59", "defects": [], "reason": "N/A", "status": "In Progress", "closed_date": "", "do_no": "", "new_serial": ""}, {"ticket": "TNB 638/2025", "location": "2880", "contract": "", "model": "", "reported": "", "site_visit": "", "serial": "", "defects": [], "reason": "N/A", "status": "In Progress", "closed_date": "", "do_no": "", "new_serial": ""}];

let allTickets = [...STATIC_TICKETS];
let currentFilter = 'all';

// ── AI STATUS ──
fetch('/ai_status').then(r=>r.json()).then(d=>{
  const dot=document.getElementById('ai-dot');
  const sub=document.getElementById('ai-lbl-sub');
  if(d.active){dot.classList.remove('off');sub.textContent='AI: '+d.name+' — Ready';}
  else sub.textContent='No AI Key configured';
}).catch(()=>{});

fetch('/template_status').then(r=>r.json()).then(d=>{
  if(!d.saved) document.getElementById('tpl-warn').style.display='block';
}).catch(()=>{});

// ── HOME ──
function initHome(){
  const total = allTickets.length;
  const badge = document.getElementById('total-badge');
  if(badge) badge.textContent = total+' Records';
}
initHome();

function showScreen(id){
  document.getElementById('home-screen').style.display='none';
  document.querySelectorAll('.screen').forEach(s=>s.classList.remove('active'));
  document.getElementById(id).classList.add('active');
  if(id==='history-screen') renderTickets();
}

function showHome(){
  document.querySelectorAll('.screen').forEach(s=>s.classList.remove('active'));
  document.getElementById('home-screen').style.display='flex';
}

// ── TICKET HISTORY ──
function statusClass(s){
  if(!s) return 's-progress';
  const sl = s.toLowerCase();
  if(sl.includes('closed')) return 's-closed';
  if(sl.includes('docs')) return 's-docs';
  return 's-progress';
}
function statusLabel(s){
  if(!s) return 'In Progress';
  if(s.toLowerCase().includes('closed')) return '✅ Closed';
  if(s.toLowerCase().includes('docs')) return '📄 Docs Submitted';
  return '🔄 In Progress';
}

function renderTickets(){
  const query = (document.getElementById('search-input')?.value||'').toLowerCase();
  let filtered = allTickets.filter(t=>{
    if(currentFilter==='closed' && !t.status.toLowerCase().includes('closed')) return false;
    if(currentFilter==='progress' && (t.status.toLowerCase().includes('closed')||t.status.toLowerCase().includes('docs'))) return false;
    if(currentFilter==='docs' && !t.status.toLowerCase().includes('docs')) return false;
    if(query){
      const hay = (t.ticket+t.location+t.contract+t.serial+t.model).toLowerCase();
      if(!hay.includes(query)) return false;
    }
    return true;
  });

  // Stats (always from full data)
  document.getElementById('stat-total').textContent = allTickets.length;
  document.getElementById('stat-closed').textContent = allTickets.filter(t=>t.status.toLowerCase().includes('closed')).length;
  document.getElementById('stat-progress').textContent = allTickets.filter(t=>!t.status.toLowerCase().includes('closed')&&!t.status.toLowerCase().includes('docs')).length;
  document.getElementById('stat-docs').textContent = allTickets.filter(t=>t.status.toLowerCase().includes('docs')).length;
  document.getElementById('rec-count').textContent = allTickets.length+' records';
  document.getElementById('list-count').textContent = filtered.length+' results';

  const list = document.getElementById('ticket-list');
  if(filtered.length===0){
    list.innerHTML='<div class="sbox s-warn">No tickets found matching your search.</div>'; return;
  }

  list.innerHTML = filtered.map((t,i)=>`
    <div class="ticket-card" onclick="openTicket(${allTickets.indexOf(t)})">
      <div class="tc-header">
        <div class="tc-left">
          <div class="tc-ticket">#${t.ticket}</div>
          <div class="tc-location">📍 ${t.location}</div>
          <div class="tc-contract">${t.contract} &nbsp;·&nbsp; ${t.model}</div>
        </div>
        <div class="tc-right">
          <span class="status-badge ${statusClass(t.status)}">${statusLabel(t.status)}</span>
          <span class="tc-date">Reported: ${t.reported}</span>
        </div>
      </div>
      ${t.defects.length>0||t.reason!=='N/A'?`
      <div class="tc-body">
        <div style="font-size:10px;font-weight:700;color:var(--muted);text-transform:uppercase;letter-spacing:.06em;margin-bottom:4px;">Defective Components</div>
        <div class="tc-detail-row">
          ${t.defects.length>0?t.defects.map(d=>`<span class="defect-tag">⚡ ${d}</span>`).join(''):'<span style="font-size:11px;color:var(--muted);">No defect detected</span>'}
        </div>
        ${t.reason&&t.reason!=='N/A'?`<div style="font-size:11px;color:var(--muted);margin-top:6px;">📝 ${t.reason}</div>`:''}
      </div>`:''}
    </div>
  `).join('');
}

function setFilter(f, btn){
  currentFilter=f;
  document.querySelectorAll('.filter-btn').forEach(b=>b.classList.remove('active'));
  btn.classList.add('active');
  const labels={all:'All Records',closed:'Closed Tickets',progress:'In Progress',docs:'Docs Submitted'};
  document.getElementById('list-label').textContent=labels[f];
  renderTickets();
}

function filterTickets(){ renderTickets(); }

function openTicket(idx){
  const t = allTickets[idx];
  document.getElementById('m-ticket').textContent = 'Ticket #'+t.ticket;
  document.getElementById('m-location').textContent = '📍 '+t.location;
  document.getElementById('m-status-badge').innerHTML = `<span class="status-badge ${statusClass(t.status)}">${statusLabel(t.status)}</span>`;

  document.getElementById('modal-body').innerHTML = `
    <div class="detail-section">
      <div class="detail-section-title">Contract Details</div>
      <div class="detail-grid">
        <div class="detail-item"><div class="detail-lbl">Contract No</div><div class="detail-val">${t.contract}</div></div>
        <div class="detail-item"><div class="detail-lbl">Lantern Model</div><div class="detail-val">${t.model}</div></div>
        <div class="detail-item"><div class="detail-lbl">Reported Date</div><div class="detail-val">${t.reported}</div></div>
        <div class="detail-item"><div class="detail-lbl">Site Visit Date</div><div class="detail-val">${t.site_visit||'—'}</div></div>
      </div>
    </div>

    <div class="detail-section">
      <div class="detail-section-title">Lantern Serial</div>
      <div class="detail-grid">
        <div class="detail-item full"><div class="detail-lbl">Defective Serial No</div><div class="detail-val mono">${t.serial||'—'}</div></div>
        <div class="detail-item full"><div class="detail-lbl">Replacement Serial No</div><div class="detail-val mono">${t.new_serial||'—'}</div></div>
      </div>
    </div>

    <div class="detail-section">
      <div class="detail-section-title">Defect Analysis</div>
      <div class="detail-item full" style="margin-bottom:8px;">
        <div class="detail-lbl">Defective Components</div>
        <div class="defect-list">
          ${t.defects.length>0?t.defects.map(d=>`<span class="defect-tag">⚡ ${d}</span>`).join(''):'<span style="font-size:12px;color:var(--muted);">No defect detected (Non-Faulty)</span>'}
        </div>
      </div>
      <div class="detail-item full">
        <div class="detail-lbl">Root Cause / Reason</div>
        <div class="detail-val" style="font-size:12px;line-height:1.5;">${t.reason||'N/A'}</div>
      </div>
    </div>

    <div class="detail-section">
      <div class="detail-section-title">Submission Details</div>
      <div class="detail-grid">
        <div class="detail-item"><div class="detail-lbl">DO No</div><div class="detail-val mono">${t.do_no||'—'}</div></div>
        <div class="detail-item"><div class="detail-lbl">Closed Date</div><div class="detail-val">${t.closed_date||'—'}</div></div>
      </div>
    </div>
  `;

  document.getElementById('modal-overlay').classList.add('open');
}

function closeModal(e){
  if(e.target===document.getElementById('modal-overlay'))
    document.getElementById('modal-overlay').classList.remove('open');
}

// Upload new Excel data
async function uploadTicketData(input){
  const file = input.files[0]; if(!file) return;
  const fd = new FormData();
  fd.append('file', file);
  const done = document.getElementById('upload-done');
  done.style.display='block';
  done.textContent='⏳ Processing...';
  try{
    const r = await fetch('/upload_tracking', {method:'POST', body:fd});
    const d = await r.json();
    if(d.ok){
      allTickets = d.tickets;
      done.textContent = `✓ Loaded ${d.tickets.length} records`;
      document.getElementById('total-badge').textContent = d.tickets.length+' Records';
      renderTickets();
    } else {
      done.textContent = '✗ Error: '+d.error;
    }
  } catch(e){
    done.textContent = '✗ '+e.message;
  }
}

// ── GENERATE TAB ──
function switchGenTab(tab){
  const isGen = tab==='generate';
  document.getElementById('gen-tab').style.display = isGen?'block':'none';
  document.getElementById('genhistory-tab').style.display = isGen?'none':'block';
  document.getElementById('tab-gen-btn').classList.toggle('active', isGen);
  document.getElementById('tab-hist-btn').classList.toggle('active', !isGen);
  if(!isGen) loadHistory();
}

const CAUSES=['LED Driver','LED Module','SPD','Natural Disaster','N/A'];
const ACTIONS=['Lantern Replacement','Return Back TNB','Void'];
const RCAS=['Good','Water Ingression','Natural Disaster','Broken','Lantern Damage','Body Crack','Component','Warranty Expired','N/A'];

let units=[];
let unitIdCounter=0;

function updateDefQty(){
  const el=document.getElementById('f-defqty');
  if(el) el.value = units.length > 0 ? units.length+' Nos' : '';
}

function fillField(id,val){
  if(!val) return false;
  const el=document.getElementById(id); if(!el||el.value) return false;
  el.value=val; el.classList.add('auto-filled');
  const badgeId='ai-'+id.replace(/^f-/,'');
  const b=document.getElementById(badgeId); if(b) b.style.display='inline';
  return true;
}

function addUnit(){
  const id=++unitIdCounter;
  units.push({id,files:{},extracted:{},cause:'',action:'',rca:'',lamputest:'',processing:false});
  renderUnits(); updateDefQty();
  setTimeout(()=>{
    const el=document.getElementById('unit-'+id);
    if(el) el.scrollIntoView({behavior:'smooth',block:'center'});
  },100);
}

function removeUnit(id){
  units=units.filter(u=>u.id!==id);
  renderUnits(); updateStatus(); updateDefQty();
}

function renderUnits(){
  const list=document.getElementById('units-list');
  list.innerHTML=units.map(u=>{
    const sn=u.extracted.serial||`Unit ${u.id}`;
    const contract=u.extracted.contract?`<span class="tag tag-g">📋 ${u.extracted.contract}</span>`:'';
    const date=u.extracted.delivery_date?`<span class="tag tag-g">📅 ${u.extracted.delivery_date}</span>`:'';
    const procMsg=u.processing?`<div class="proc-bar">🤖 AI scanning...</div>`:'';

    return `<div class="unit-card" id="unit-${u.id}">
      <div class="unit-header">
        <span class="unit-sn">#${u.id} — ${sn}</span>
        <button class="del-btn" onclick="removeUnit(${u.id})">✕ Remove</button>
      </div>
      <div style="display:flex;gap:4px;flex-wrap:wrap;margin-bottom:8px;">${contract}${date}${!contract&&!date?'<span class="tag tag-y">Pending AI scan</span>':''}</div>

      <div style="font-size:10px;font-weight:700;color:var(--muted);text-transform:uppercase;letter-spacing:.05em;margin-bottom:4px;">📸 Lantern Photos</div>
      <div class="cam-grid">
        ${['full','serial','issue'].map(type=>{
          const icons={'full':'🏮','serial':'🏷️','issue':'⚠️'};
          const labels={'full':'Full Lantern','serial':'Serial Label','issue':'Issue'};
          const captured=!!u.files[type];
          const thumb=u.thumbs&&u.thumbs[type]?`<img class="cam-thumb" src="${u.thumbs[type]}">`:'';
          const check=captured?`<div class="cam-check">✓</div>`:'';
          return `<div style="display:flex;flex-direction:column;gap:4px;">
            ${thumb?`<div style="position:relative;aspect-ratio:1;border-radius:12px;overflow:hidden;border:2px solid var(--green);">${thumb}${check}</div>`
            :`<div style="aspect-ratio:1;border-radius:12px;border:2px dashed var(--border);background:var(--light);display:flex;flex-direction:column;align-items:center;justify-content:center;gap:3px;">
                <div style="font-size:22px;">${icons[type]}</div>
                <div style="font-size:10px;color:var(--muted);font-weight:600;">${labels[type]}</div>
              </div>`}
            <div style="display:grid;grid-template-columns:1fr 1fr;gap:3px;">
              <label style="background:var(--accent);color:#fff;border-radius:7px;padding:5px 2px;font-size:10px;font-weight:700;text-align:center;cursor:pointer;display:block;">
                <input type="file" accept="image/*" capture="environment" style="display:none;" onchange="capturePhoto(${u.id},'${type}',this)">📷</label>
              <label style="background:var(--blue-bg);color:var(--accent);border:1px solid var(--accent);border-radius:7px;padding:5px 2px;font-size:10px;font-weight:700;text-align:center;cursor:pointer;display:block;">
                <input type="file" accept="image/*" style="display:none;" onchange="capturePhoto(${u.id},'${type}',this)">🖼️</label>
            </div>
          </div>`;
        }).join('')}
      </div>

      <div style="font-size:10px;font-weight:700;color:var(--muted);text-transform:uppercase;letter-spacing:.05em;margin:10px 0 4px;">🔧 Component Photos</div>
      <div class="cam-grid">
        ${['lampu_test','spd','driver'].map(type=>{
          const icons={'lampu_test':'💡','spd':'⚡','driver':'🔌'};
          const labels={'lampu_test':'Lamp Test','spd':'SPD','driver':'Driver'};
          const captured=!!u.files[type];
          const thumb=u.thumbs&&u.thumbs[type]?`<img class="cam-thumb" src="${u.thumbs[type]}">`:'';
          const check=captured?`<div class="cam-check">✓</div>`:'';
          return `<div style="display:flex;flex-direction:column;gap:4px;">
            ${thumb?`<div style="position:relative;aspect-ratio:1;border-radius:12px;overflow:hidden;border:2px solid var(--green);">${thumb}${check}</div>`
            :`<div style="aspect-ratio:1;border-radius:12px;border:2px dashed var(--border);background:var(--light);display:flex;flex-direction:column;align-items:center;justify-content:center;gap:3px;">
                <div style="font-size:22px;">${icons[type]}</div>
                <div style="font-size:10px;color:var(--muted);font-weight:600;">${labels[type]}</div>
              </div>`}
            <div style="display:grid;grid-template-columns:1fr 1fr;gap:3px;">
              <label style="background:var(--accent);color:#fff;border-radius:7px;padding:5px 2px;font-size:10px;font-weight:700;text-align:center;cursor:pointer;display:block;">
                <input type="file" accept="image/*" capture="environment" style="display:none;" onchange="capturePhoto(${u.id},'${type}',this)">📷</label>
              <label style="background:var(--blue-bg);color:var(--accent);border:1px solid var(--accent);border-radius:7px;padding:5px 2px;font-size:10px;font-weight:700;text-align:center;cursor:pointer;display:block;">
                <input type="file" accept="image/*" style="display:none;" onchange="capturePhoto(${u.id},'${type}',this)">🖼️</label>
            </div>
          </div>`;
        }).join('')}
      </div>

      <div class="field" style="margin:6px 0 10px;">
        <label class="lbl" style="font-size:9px;">💡 Lamp Test Result</label>
        <select onchange="units.find(u=>u.id==${u.id}).lamputest=this.value">
          <option value="">— Select result —</option>
          <option value="Lamp lights up after testing"${u.lamputest==='Lamp lights up after testing'?' selected':''}>✅ Lamp lights up after testing</option>
          <option value="Lamp does not light up"${u.lamputest==='Lamp does not light up'?' selected':''}>❌ Lamp does not light up</option>
        </select>
      </div>
      ${procMsg}
      <div style="display:grid;grid-template-columns:1fr 1fr 1fr;gap:6px;margin-top:4px;">
        <div class="field" style="margin:0"><label class="lbl" style="font-size:9px;">Cause</label>
          <select onchange="units.find(u=>u.id==${u.id}).cause=this.value">
            <option value="">—</option>${CAUSES.map(c=>`<option${u.cause===c?' selected':''}>${c}</option>`).join('')}
          </select></div>
        <div class="field" style="margin:0"><label class="lbl" style="font-size:9px;">Action</label>
          <select onchange="units.find(u=>u.id==${u.id}).action=this.value">
            <option value="">—</option>${ACTIONS.map(c=>`<option${u.action===c?' selected':''}>${c}</option>`).join('')}
          </select></div>
        <div class="field" style="margin:0"><label class="lbl" style="font-size:9px;">RCA</label>
          <select onchange="units.find(u=>u.id==${u.id}).rca=this.value">
            <option value="">—</option>${RCAS.map(c=>`<option${u.rca===c?' selected':''}>${c}</option>`).join('')}
          </select></div>
      </div>
    </div>`;
  }).join('');
  updateStatus();
}

async function capturePhoto(uid,type,input){
  const file=input.files[0]; if(!file) return;
  const unit=units.find(u=>u.id===uid); if(!unit) return;
  unit.files[type]=file;
  if(!unit.thumbs) unit.thumbs={};
  unit.thumbs[type]=URL.createObjectURL(file);
  renderUnits();
  if(type==='serial'){
    unit.processing=true; renderUnits();
    const fd=new FormData();
    fd.append(`img__${uid}__${type}`,file);
    for(const t of ['full','issue']) if(unit.files[t]) fd.append(`img__${uid}__${t}`,unit.files[t]);
    try{
      const r=await fetch('/scan_unit',{method:'POST',body:fd});
      const d=await r.json();
      if(d.extracted){
        unit.extracted=d.extracted;
        const ex=d.extracted;
        if(ex.contract) fillField('f-contract',ex.contract);
        if(ex.defmodel) fillField('f-defmodel',ex.defmodel);
        if(ex.delivery_date) fillField('f-deliverydate',ex.delivery_date);
      }
    }catch(e){}
    unit.processing=false; renderUnits();
  }
}

function updateStatus(){
  const box=document.getElementById('status-box'); if(!box) return;
  if(units.length===0){box.className='sbox s-warn';box.textContent='⚠ Add at least one lantern to begin';}
  else{const ready=units.filter(u=>u.files.serial||u.files.full).length;box.className='sbox s-ok';box.textContent=`✓ ${units.length} unit(s) — ${ready} with photos — Ready to generate`;}
}

async function generate(){
  if(units.length===0){alert('Add at least one unit!');return;}
  const btn=document.getElementById('gen-btn');
  const prog=document.getElementById('prog');
  const pfill=document.getElementById('pfill');
  btn.disabled=true;btn.innerHTML='<span class="spin"></span> Generating...';
  prog.style.display='block';pfill.style.width='20%';
  const fd=new FormData();
  const metadata=units.map((u,i)=>({id:u.id,idx:i,extracted:u.extracted,cause:u.cause,action:u.action,rca:u.rca,new_serial:''}));
  fd.append('metadata',JSON.stringify(metadata));
  fd.append('send_email','false');
  const info={
    ticket:document.getElementById('f-ticket').value,
    contract:document.getElementById('f-contract').value,
    station:document.getElementById('f-station').value,
    pic:document.getElementById('f-pic').value,
    sitedate:document.getElementById('f-date').value,
    dono:document.getElementById('f-dono').value,
    totalqty:document.getElementById('f-totalqty').value,
    defqty:document.getElementById('f-defqty').value,
    defmodel:document.getElementById('f-defmodel').value,
    newmodel:document.getElementById('f-newmodel').value,
    dodate:document.getElementById('f-deliverydate').value,
    delivery:document.getElementById('f-delivery').value,
    staffname:document.getElementById('f-staffname').value,
  };
  fd.append('info',JSON.stringify(info));
  for(const u of units) for(const[type,file] of Object.entries(u.files)) if(file) fd.append(`photo__${u.id}__${type}`,file);
  pfill.style.width='50%';
  try{
    const r=await fetch('/generate_mobile',{method:'POST',body:fd});
    pfill.style.width='90%';
    if(r.ok){
      const blob=await r.blob();
      const url=URL.createObjectURL(blob);
      const a=document.createElement('a');
      const contract=info.contract.replace(/\//g,'-').replace(/\s/g,'')||'DAR';
      const fname=`DAR_${contract}_${info.ticket||'output'}.xlsx`;
      a.href=url;a.download=fname;
      document.body.appendChild(a);a.click();document.body.removeChild(a);
      window._lastDarBlob=blob;window._lastDarFname=fname;
      pfill.style.width='100%';
      document.getElementById('gen-msg').innerHTML=`<div class="sbox s-ok" style="margin-top:8px">✓ DAR Excel generated &amp; downloaded successfully!<br>
        <button onclick="shareWhatsApp()" style="margin-top:8px;width:100%;background:#25D366;color:#fff;border:none;border-radius:10px;padding:12px;font-size:14px;font-weight:700;cursor:pointer;">
          Share via WhatsApp
        </button></div>`;
    }else{
      const err=await r.text();
      document.getElementById('gen-msg').innerHTML=`<div class="sbox s-err" style="margin-top:8px">✗ ${err}</div>`;
    }
  }catch(e){
    document.getElementById('gen-msg').innerHTML=`<div class="sbox s-err" style="margin-top:8px">✗ ${e.message}</div>`;
  }
  btn.disabled=false;btn.innerHTML='⬇ Generate DAR Excel';
  setTimeout(()=>{prog.style.display='none';pfill.style.width='0';},2000);
}

async function shareWhatsApp(){
  const darBlob=window._lastDarBlob;
  if(!darBlob){alert('Generate report first!');return;}
  const info={ticket:document.getElementById('f-ticket').value,station:document.getElementById('f-station').value,contract:document.getElementById('f-contract').value,staff:document.getElementById('f-staffname').value};
  const text=`DAR Report — SNFOR SDN BHD\nTicket: ${info.ticket}\nStation: ${info.station}\nContract: ${info.contract}${info.staff?'\nStaff: '+info.staff:''}`;
  const files=[new File([darBlob],window._lastDarFname,{type:'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'})];
  if(navigator.canShare&&navigator.canShare({files})){
    try{await navigator.share({files,title:window._lastDarFname,text});return;}catch(e){if(e.name==='AbortError')return;}
  }
  window.open(`https://wa.me/?text=${encodeURIComponent(text+'\n\nPlease check downloaded file.')}`, '_blank');
}

async function loadHistory(){
  try{
    const r=await fetch('/history');
    const data=await r.json();
    const list=document.getElementById('history-list');
    document.getElementById('hist-count').textContent=data.length+' reports';
    if(data.length===0){list.innerHTML='<div class="sbox s-warn">No reports yet.</div>';return;}
    list.innerHTML=data.map(rec=>`
      <div class="hist-card">
        <div style="display:flex;justify-content:space-between;align-items:flex-start;margin-bottom:10px;">
          <div><div style="font-size:16px;font-weight:800;color:var(--navy);">🎫 Ticket #${rec.ticket}</div>
          <div style="font-size:11px;color:var(--muted);margin-top:2px;">📅 ${rec.date}</div></div>
          <button onclick="deleteRecord('${rec.id}')" style="background:var(--red-bg);color:var(--red-text);border:none;border-radius:6px;padding:4px 8px;font-size:11px;cursor:pointer;">Delete</button>
        </div>
        <div style="display:grid;grid-template-columns:1fr 1fr;gap:6px;font-size:12px;margin-bottom:10px;">
          <div><span style="color:var(--muted);font-size:10px;font-weight:700;text-transform:uppercase;">Station</span><br><b>${rec.station}</b></div>
          <div><span style="color:var(--muted);font-size:10px;font-weight:700;text-transform:uppercase;">Units</span><br><b>${rec.units}</b></div>
          <div><span style="color:var(--muted);font-size:10px;font-weight:700;text-transform:uppercase;">Contract</span><br><b>${rec.contract}</b></div>
          <div><span style="color:var(--muted);font-size:10px;font-weight:700;text-transform:uppercase;">Staff</span><br><b>${rec.staff}</b></div>
        </div>
        <button onclick="downloadRecord('${rec.id}')" style="width:100%;background:var(--accent);color:#fff;border:none;border-radius:10px;padding:12px;font-size:14px;font-weight:700;cursor:pointer;">⬇ Download Excel</button>
      </div>`).join('');
  }catch(e){document.getElementById('history-list').innerHTML='<div class="sbox s-err">Error loading reports</div>';}
}

function downloadRecord(id){const a=document.createElement('a');a.href=`/download/${id}`;document.body.appendChild(a);a.click();document.body.removeChild(a);}
async function deleteRecord(id){if(!confirm('Delete this report?'))return;await fetch(`/delete/${id}`);loadHistory();}
</script>
</body>
</html>'''

# ─── Email text parser ──────────────────────────────────────────────────────────
def parse_email_text(text):
    result = {}
    for p in [r'ticket\s*[#:]?\s*(\d{4,8})', r'#\s*(\d{4,8})']:
        m = re.search(p, text, re.I)
        if m: result['ticket'] = m.group(1).strip(); break
    for p in [r'station\s*[:\-]?\s*([A-Za-z][A-Za-z0-9\s]+?)(?:\n|,|\.)', r'location\s*[:\-]?\s*([A-Za-z][A-Za-z0-9\s]+?)(?:\n|,|\.)']:
        m = re.search(p, text, re.I)
        if m: result['station'] = m.group(1).strip(); break
    for p in [r'jumlah\s*[:\-]?\s*(\d+)', r'(?:quantity|qty)[^\d]*(\d+)', r'(\d+)\s*(?:biji|pcs|nos?)']:
        m = re.search(p, text, re.I)
        if m: result['totalqty'] = m.group(1)+' Nos'; break
    m = re.search(r'(\d{1,2}[/\-]\d{1,2}[/\-]\d{2,4})', text)
    if m:
        try:
            parts = re.split(r'[/\-]', m.group(1))
            if len(parts)==3:
                d,mo,y = parts
                if len(y)==2: y='20'+y
                result['sitedate'] = f"{y}-{mo.zfill(2)}-{d.zfill(2)}"
        except: pass
    for p in [r'contract\s*(?:no)?\s*[:\-]?\s*([\w\s\/\-]+?)(?:\n|,|\.)', r'(TNB[\s/]?\d+[\/\-]\d+)']:
        m = re.search(p, text, re.I)
        if m:
            c = m.group(1).strip()
            if len(c)>5 and '/' not in c: break
            result['contract'] = c; break
    for p in [r'pic\s*[:\-]?\s*([A-Za-z][A-Za-z\s]+?)(?:\n|,|\.|tel)', r'(?:name|nama)\s*[:\|]?\s*([A-Za-z][A-Za-z\s]+?)(?:\n|,|\||$)']:
        m = re.search(p, text, re.I)
        if m: result['pic'] = m.group(1).strip(); break
    return result

# ─── Email & PDF ───────────────────────────────────────────────────────────────
def send_email(to_addr, subject, body, attachments):
    """Send email via Gmail or Outlook SMTP with attachments."""
    import smtplib
    from email.mime.multipart import MIMEMultipart
    from email.mime.base import MIMEBase
    from email.mime.text import MIMEText
    from email import encoders

    cfg = STATE['email_config']
    addr = cfg['address']
    pwd  = cfg['password']
    if not addr or not pwd:
        raise Exception("Email belum configure. Set EMAIL dan PASSWORD dalam run_mobile.bat")

    msg = MIMEMultipart()
    msg['From']    = addr
    msg['To']      = addr
    msg['Subject'] = subject
    msg.attach(MIMEText(body, 'html'))

    for fname, fdata, ftype in attachments:
        part = MIMEBase('application', 'octet-stream')
        part.set_payload(fdata)
        encoders.encode_base64(part)
        part.add_header('Content-Disposition', f'attachment; filename="{fname}"')
        msg.attach(part)

    # Auto-detect Gmail or Outlook
    if 'gmail.com' in addr.lower():
        smtp_host, smtp_port = 'smtp.gmail.com', 587
    else:
        smtp_host, smtp_port = 'smtp.office365.com', 587

    with smtplib.SMTP(smtp_host, smtp_port) as s:
        s.starttls()
        s.login(addr, pwd)
        s.send_message(msg)
        print(f"  ✓ Email sent via {smtp_host}")

def excel_to_pdf(excel_bytes, filename):
    """Convert Excel to simple PDF using reportlab as fallback."""
    try:
        # Try using LibreOffice if available
        import subprocess, tempfile, os
        with tempfile.TemporaryDirectory() as tmp:
            xlsx_path = os.path.join(tmp, filename)
            with open(xlsx_path, 'wb') as f: f.write(excel_bytes)
            result = subprocess.run(
                ['soffice', '--headless', '--convert-to', 'pdf', '--outdir', tmp, xlsx_path],
                capture_output=True, timeout=30
            )
            pdf_path = xlsx_path.replace('.xlsx', '.pdf')
            if os.path.exists(pdf_path):
                with open(pdf_path, 'rb') as f: return f.read()
    except: pass

    # Fallback: simple PDF with reportlab
    try:
        from reportlab.lib.pagesizes import A4
        from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer
        from reportlab.lib.styles import getSampleStyleSheet
        buf = io.BytesIO()
        doc = SimpleDocTemplate(buf, pagesize=A4)
        styles = getSampleStyleSheet()
        story = [
            Paragraph("SNFOR SDN BHD", styles['Title']),
            Paragraph("Defect Analysis Report", styles['Heading1']),
            Spacer(1, 12),
            Paragraph("Please refer to the attached Excel file for the complete DAR Report.", styles['Normal']),
            Spacer(1, 12),
            Paragraph(f"Generated: {__import__('datetime').datetime.now().strftime('%d/%m/%Y %H:%M')}", styles['Normal']),
        ]
        doc.build(story)
        buf.seek(0)
        return buf.read()
    except Exception as e:
        print(f"  PDF error: {e}")
        return None

# ─── HTTP Handler ───────────────────────────────────────────────────────────────
class Handler(BaseHTTPRequestHandler):
    def log_message(self, *a): pass
    timeout = 300

    def do_GET(self):
        if self.path == '/':
            b = HTML.encode()
            self.send_response(200); self.send_header('Content-Type','text/html; charset=utf-8')
            self.send_header('Content-Length',len(b)); self.end_headers(); self.wfile.write(b)
        elif self.path == '/manifest.json':
            import json as _json
            mf = _json.dumps({"name": "DAR Mobile - SNFOR SDN BHD", "short_name": "DAR Mobile", "description": "Defect Analysis Report Mobile App", "start_url": "/", "display": "standalone", "background_color": "#1A56DB", "theme_color": "#1A56DB", "orientation": "portrait", "icons": [{"src": "/icon-192.png", "sizes": "192x192", "type": "image/png", "purpose": "any maskable"}, {"src": "/icon-512.png", "sizes": "512x512", "type": "image/png", "purpose": "any maskable"}]}).encode()
            self.send_response(200); self.send_header('Content-Type','application/manifest+json')
            self.send_header('Content-Length',len(mf)); self.end_headers(); self.wfile.write(mf)
        elif self.path == '/icon-192.png':
            import base64 as _b64
            ico = _b64.b64decode('iVBORw0KGgoAAAANSUhEUgAAAMAAAADACAIAAADdvvtQAAAQH0lEQVR4nO3deXgUZZ4H8F91dbq7uruSTndOQkhCOEJISBANGHIggoqyZNAkPoiMu8sjf6yj4zOOzvrMuESdZ3lG3XHZ2XmefXSY2ZFZDwTURUFnEJyEI6ADSTgFAknIfXSOvq+q/aM1JE1VpTtvd/rY3+fhjzzdRVd3+pv3rreoWbVXAaHpkoX7DaDohgFCRDBAiAgGCBHBACEiGCBEBAOEiGCAEBEMECKCAUJEMECICAYIEcEAISIYIEQEA4SIYIAQEQwQIoIBQkQwQIgIBggRwQAhIhggRAQDhIhggBARDBAiggFCRDBAiAgGCBHBACEiGCBEBAOEiGCAEBEMECKCAUJEMECICAYIEcEAISIYIEQEA4SIYIAQEQwQIoIBQkQwQIgIBggRwQAhIhggRAQDhIhggBARDBAiggFCRDBAiAgGCBHBACEiGCBEBAOEiGCAEBEMECKCAUJEMECICAYIEcEAISIYIEQEA4SIYIAQEQwQIoIBQkQwQIgIBggRkYf7DcCvnkx+fE2C2LM8Dx6Od7nB4eIsdn7Myg2bPL3D7s5+d2uP8/wNx7edTp4P5vtRKaimt3JYRupPq/SZ9vY+V0AvK/0xJ3J7eJuDt9i5EQvX2u281uU8c9VRf87qdAX1cwZJ+AMkjaJATlNyGhglrdNCxm0HDJs8h7627Ks3NV6yBeWM60q00ukBgJoK9o0PjUE53e3kNMWqKVYtS9NDXqbC+6DZxh06bXnjw6HOAXeIzjs9UV+FJbL0Y6vj99VlHPjl7JUFDPkL1lSyUx5TXclSFPmpAqBlZDWVbP2bWU8+qJvRE08l6gM07o75qj0vZby2LUWtnP6HStPLywvVUx6WmRy3YlEQwhooZRxV90TSMw8nzvypxcROgLw23xu//+WMlMRpVs2PlLMy/4oWfwqqEHm+1rBkrjJcZ/cRawECgMIc5Ud1GSk6ehr/t7rC31g8tFzLKGe2GvuejIKfb04Ky6lvF9GN6Lo/Dr59cMT7M6uWJWrpWQZ5ySJVeYG6dLFUDZKdFrf7xVkbftHpCKTnUpyrXDBb4efBWkb2YIl2X4PJ/9cXM/FjymkqQSPLz1I+uordWCaa5rvzGZ2WHjF7yM9OKGpKIJOV6+h3NV6y/cf+4ZpXuta+cPOLbywSxxdkK/9lS2B/pjWV8QEd739x5T+3hx8a8zScs/7oN311fxwUO4yWQWl+GBpht4uaAPm42O74x9d76t4Z9HCix/zwvoQiv9sKcpqqKtUKPnXyovAAQVmhOt0QwiJ816ERidGmkJ7af9EaIK+3Pxt5dbfon6mMgp/WGvx8qbXL1ImscLPp+bf6h00ClYWMgkfKQ9iU5ng4dl50cMsQHxHfXUS8CRJvHxz5/GvRuuyeYnVOepw/ryNWf/3tiv1Gj+uzU8KnqAlBLTZRr1F02JATL3pnUtQHCABe3zMk9hRFwd/dLVwxTaRn6XuXCg//fHzcBACfnBBuLM/LUBTPU/n3NqdDYkzBKFQozrxYCNDlDqdEUb+qaOqBwY1lrJwW+K48HHzaaAaAxou2vmHhwqA2lANCEg2di+3O0J3Xf7EQIABoOGcVe6porkowHBOJjQqeuGDtH/EAAMfDgZNmwWOqSrVx8pAMCMkoKCsQTv+ohTt7zR6KkwYqRgJ0+rJoCaRSUBlJUh2WhbMVhTnCnbWPj5sFf55Ip6XX3qHx720GZtt63exk4Xf+/tGxgIa4QidGAnRTco56lmSPV6z4cbn5g6dvhebsNXtHv3CnOojTGrQM9CxdXqj+7TOpLz0uPI7Va3S/uS9UawECFRFjCeSkx2S14sszaJloV/xIk3XMMqmr878nzD/6gcBE5uqlakM8PTQ2zVZt3RNJdU/4O+Y5bPI8vqPbZI2MPljMlEA2By8xoqiME22jVCxRi828evtf0o94yWlKYtohiL65Yl/3YueljohoPnvFSIDUShkt/lEkmgtiwz9WB/eXv/mO/VzqcF7pFP7yQjo5z/Nw/LztyX/rrXqp8+ZAYCshQy1GApSglfogYgU+q5bdf6dw+/eLry02h0DsPjkh3JQuyFbmzfF3InYaNIxMqQjP5L+0GAlQVorUcHPXoHATe8PdWpXItyLW5xIbUQSA2gDnYv1HUVCcq/zPp1N/91y62BsOlxgJUMki0eFgm4PvEZkQEKu/Ri3cV83CA0s3elwt1x2CT20sYyWq0aBYV6L5r2fT/FzyNjNipBdWuUR0uLn5ut3tEaiMslLj7looHLsEjaz93dxA30OKjq5coj7SJDqkKWZ8PVC8RjY7KW7tMs3WdQmGeOGZ3bXLNE9VJf7m4+FAzxIisVACFeYoJVYofyXyjYZiHjTQFUU+xizcxXbHzv3GVT/puNAmXM4BwE+q9XP9myGeAbEQoBceFV2zwfNwoFGgNUNR8EgIAnT/XZp4TRB+pUaT54nXenxGocYp4qjtP4yUJa1RH6CnqhJXi0ykA8CRs5a2XoF+74pFzBzJdvf0KOOoDX5M/vujZ8j9r++JrjJYc4dGrP6dYdEdoH/akPjPm0SLH44Hscv/QjdsQ1iLTfTekbHrPaKjPv6vlQupaA1QQbbynZ+l/3yzQaJL8s6fRwV7TIySemh5cMqJ2925QJWdFpyyze3hd+4XnfMqK2BK8sJfCEVNL4xlZDotnZEkL1mkqihU3z3VkvLzbY5XRFa7PliiFZsd+/Fv+/bW+3WhRdFc5cEdmYJP1VSwr+8JzmTnR8dMP35YtMn80xpD7atdQTnRtEV0gAKaZZyorde1ZUe32AyG2NUUbg9/+Iy/nfDm647uIbfgPH91RfwbHxqDsuWDh4Od+407n0oVfHZlAbM8jzklvpRlBkRrFSah5bpjY12XdyHY7dL08jKRi5cbL9kDutLqC5G12LOT5UG88PmjYyaJltBzNfpgnWh6YipAPA9/Ojz2cF1nv8jyUwCorhC9ePnQaeHpCzGHvhY9PojrXL2FkNizKwuYsFylPy52AnT6sr365a6fvd0vOAk6Tqz+4nmQuLpD0CnxEuuhFVqSPR58RHIhFPUBGhj1/OnwaNVLnRu3d065RVDxPNX8DOE58+ZWu8Q1NIIk2kwalezB5UFb5ypdCJUuZqbsUoRORDeiAYDjwe3mXR7e5eZNNm7Uwo2YuZ4hd0e/60av6+w1u+A4oRiJmuVQgMXPd//rtFmsSKupjPezQ+cP6e7YczX66pfD0x2jZtVeDcuJUWyI+ioMhRcGCBHBACEiGCBEBAOEiGCAEBEMECKCAUJEMECICAYIEcEAISIYIEQEA4SIRPpyDh/P1qZWFLMA4OH4rTvaxiwC67lWFmqf25Tm/Xnnnr6/Nk1nTcWvn87MTlf++v3eYy1+LVOs25qxJJf53YGBgydHxR6JSdFaAtEyamWh8KU53oShmRGVARoxeQCgXCgoWoZeukBtdXARsgdlzIuyKszrYpstL0u1MFOVkhjXPzxpRWJpoVZOU/XNphX5WhDf2S7o6naF+fqscInKEojn4ViLmaKgvMi3Fqso1gJAfVNg11egaYvKEggAGppNG8p05UXsvq9u7ZSTpJMvymJGTJ7z10WvD5TJ4J6l8ZVL2ex0BaOQjVo8l9vtnx4fudwhvG/37BRF9arEgrlMvIYes3jOfGv94IhxaNR3+X1ATeb8bOaBFQl5WSqdlna6+Zt9zoZm059PjwnuYxThojVArV2OrgHnnFRFVpqivfe7jS8riliKgmMtJrEbkagUshe3pBfm3rqGQR8vLy3UlhZq3/3L0N6jvrs25ecwT1enjm9Er4+Xr7krfvlizfZd3W09ovv3SKAo+IeHktaX6sYfkdPUwjmqhXNUFcXsK7/vtjoiZf9eP0VrgACgvtm8aY2+oojd3fvdNije/pdEv31bVXJhLuPh+A++NB49YxqzeDKSFY+t1d+Zp3lsraGz39V4YVLd98DyBOOYe9eng01XrTwPRfOZreuTkxLkzz+W9uzODpc74ALjkVWJ60t1did34NhIQ4t5YNilYejCXGbL/YYFmaptVcn/vqcv0NcMr6hsA3k1NJkAoKzouztwZ6cr56QqugddrV3CZUO6Ia6ymAWA3Z8P7T06PDTqdrn5th7Hjt09LdesALBpre8VejwPr/5398nzZpuDszu5Uxcsr/6hm+NuvVRAErR09So9ALz2P73vHTZ29jsdLt445v7rWdMrf+h2e/jyIjZZF2V/0lEcoF6j68pNe7JOnpelglvNZ9Hi5848DUXBiNlzqHFSM4Xn4f0vjQCQmaJI00+68Krxgnm8fvS62e88fs7kfbVA33DJIo0ijrre7Wi66ttE6+hztnY5KArycyLiRpb+i7K8+6hvMi3IVJUXsZfb7eVLWABoaBYN0OwUBQC09zpur3qu3nR4PDxNU5kpil7jrXEBwcKstctRXsRmpga8K3TOLCUA5KQr9/5yHgBQEwYZxn/Ws1H2jURxCQQAx8+ZPRxfWqBdkqs2JMivdtp7hkQvVPVerG4UuqOFh+O9rVdGNekXMiJ0UzfjmBsAGEXAvzotQwMARYFMBjIZUNStf+OmvDNVpImyvPsYNXtartmWLlBvq0qGqYZ/vBHRC22fS8sob7xs9kmdIJ3QXVT18XIAsDkD7i7ZHBwANDSb3vwgylrKEqK7BILvGz3phjiOg+MtUvOmXQNOAMhKU95+f7j5mUqapgCgc2BSiyc3Q+A+Yt4Hb/YFfMeT9l4HABTMZeiI2iqcTNQH6NRFi3faq6XVKr091DeXLTwPOi29bkXCxMcpCmpX6wGgs9/pUwOuWKydM7mtk5GsWFnIAsA33wa8GUPjBYvLzSey8up7BO4ZFaWiuwoDALuT27S91Z8juwdd9c2mymJ2ywOGODn1/ThQ3KY1huL5agB477DvFioUBb/4+1m//3TQ229aMo/Zuj5ZJoNeo0uiuyfGOObe+9XwpjX6R+/Vz0qK+/zUaFuP0+3hE1l5kk5ePE9dkq95dmdHoC8bXlEfoIC89clAcoI8P4fZfJ9h832Ttsn94EvjyfO+TagvTo2uXhb/wua0iQ+abdwb7/Y6pzXbv/eoUaWgflCeWF7Elhf5jiRFyJ28A/L/K0A2B7d9V/fqZWxlMTsnTckoqDGr53K7/bMToxfbBDanunDDdrBxtOYefUEOw6ploxbPmSvWPV8aB2+bC/MTz8Puz4eOtZgfWJ6wOIcxJMgpgGGTe2DE3XTVeurCdPYoCi/cHwgRifpGNAovDBAiggFCRDBAiAgGCBHBACEiGCBEBAOEiGCAEBEMECKCAUJEMECICAYIEcEAISIYIEQEA4SIYIAQEQwQIoIBQkQwQIgIBggRwQAhIhggRAQDhIhggBARDBAiggFCRDBAiAgGCBHBACEiGCBEBAOEiGCAEBEMECKCAUJEMECICAYIEcEAISIYIEQEA4SIYIAQEQwQIoIBQkQwQIgIBggRwQAhIhggRAQDhIhggBARDBAiggFCRDBAiAgGCBHBACEiGCBE5P8AvfMxe01hz48AAAAASUVORK5CYII=')
            self.send_response(200); self.send_header('Content-Type','image/png')
            self.send_header('Content-Length',len(ico)); self.end_headers(); self.wfile.write(ico)
        elif self.path == '/icon-512.png':
            import base64 as _b64
            ico = _b64.b64decode('iVBORw0KGgoAAAANSUhEUgAAAgAAAAIACAIAAAB7GkOtAAAtx0lEQVR4nO3deXxU5d338XNmnySTPYSwhX3fwiabLCq4gFUE5G6L3Xy1dn1au93aqrW9a3u3drNP79an7W1b7aZFq9ZdC6iALIKAgIDsW4CQfSaTWc/zR3yFZCaEmTlXMpP8Pu+Xf0hIzpyQyfmec12/63fp/W59XwMAyGNJ9wkAANKDAAAAoQgAABCKAAAAoQgAABCKAAAAoQgAABCKAAAAoQgAABCKAAAAoQgAABCKAAAAoQgAABCKAAAAoQgAABCKAAAAoQgAABCKAAAAoQgAABCKAAAAoQgAABCKAAAAoQgAABCKAAAAoQgAABCKAAAAoQgAABCKAAAAoQgAABCKAAAAoQgAABCKAAAAoQgAABCKAAAAoQgAABCKAAAAoQgAABCKAAAAoQgAABCKAAAAoQgAABCKAAAAoQgAABCKAAAAoQgAABCKAAAAoQgAABCKAAAAoQgAABCKAAAAoQgAABCKAAAAoQgAABCKAAAAoQgAABCKAAAAoQgAABCKAAAAoQgAABCKAAAAoQgAABCKAAAAoQgAABCKAAAAoQgAABCKAAAAoQgAABCKAAAAoQgAABCKAAAAoQgAABCKAAAAoQgAABCKAAAAoQgAABCKAAAAoQgAABCKAAAAoQgAABCKAAAAoQgAABCKAAAAoQgAABCKAAAAoQgAABCKAAAAoQgAABCKAAAAoQgAABCKAAAAoQgAABCKAAAAoQgAABCKAAAAoQgAABCKAAAAoQgAABCKAAAAoQgAABCKAAAAoQgAABCKAAAAoQgAABCKAAAAoQgAABCKAAAAoQgAABCKAAAAoQgAABCKAAAAoQgAABCKAAAAoQgAABCKAAAAoQgAABCKAAAAoQgAABCKAAAAoQgAABCKAAAAoQgAABCKAAAAoQgAABCKAAAAoQgAABCKAAAAoQgAABCKAAAAoQgAABCKAAAAoQgAABCKAAAAoQgAABCKAAAAoQgAABCKAAAAoQgAABCKAAAAoQgAABCKAAAAoQgAABCKAAAAoQgAABCKAAAAoQgAABCKAAAAoQgAABCKAAAAoQgAABCKAAAAoQgAABCKAAAAoQgAABCKAAAAoQgAABCKAAAAoQgAABCKAAAAoQgAABCKAAAAoQgAABCKAAAAoQgAABCKAAAAoQgAABCKAAAAoQgAABCKAAAAoQgAABCKAAAAoQgAABCKAAAAoQgAABCKAAAAoQgAABCKAAAAoQgAABCKAAAAoQgAABCKAAAAoQgAABCKAAAAoQgAABDKlu4T6LVmjXWv+U5/VUczDC1qaNGoETW0SNSIRLRAyAiGjUDQCIQNfyDq9Ucb/VFvU7TRH631RqsbIjUNkeqGyNnacGV1uDloqDqTHiTLadn1u8FZTsV3OeGIUXHHsZrGiNrDpkbt2ywR4YgRDBuhsBYKG77maJ0vWueN1Hmj5+vCp6rCJ8+HTlaFj1QGZb7lehwCoGfQdc2qa1aL3vInTdM8yXx5TWPkVFX46NnQ4TPBw2dC+08E3j8djES74kwzyJKZ2cqv/pqm2az6srme/32xTvmRewSbVbdZdc2paZpWnGct7+hzIlHt0Jng3mOB3YcDb+5p2n8i2L3niEQRACIUeqyFHuvEoc7Wj/gDxr7jgbcPNm/c69/ynt/r74VpcOv83C468sr5cgMgEVaLNmqAY9QAxy1zPZqmna+LvLG76dm3vK/vagpHeDLIIASAUG6nPnWka+pI1x1L88MRY/N7zS9s8b641Xu+LiNGNszrX2ybNdbdRQefMMQ5eqBj/0lubBPSJ9+6Yp5nxTxPdUPkqQ2Nv3u+7vSFcLpPCprGJDA0TbNZ9bnj3T+4veTt3wz54zfLrp2W/cFQU0+2Yl6u3pXfxcoue7zoxYpyrZ++IX/TL8t/9rk+5aX2dJ8OCAC0YbVoi6ZmP/KNsg0PlX98cZ7L0YNzYOW8pGZJknbLlR4rvz0psVn1VQty1/5k0JduLrBZe/B7rBfgLYwOlJfaf3B7yaZflq++Jq8n/opOG+kaUta1N5h98q3zJ2Z16Uv0bi6HfteHi57/wYD+xQxEpw0BgEsqLbD96NMl/35wYNcNpneRWxd0x/hM97xK7zZ+sPOFHwycNtKV7hMRigDAZQzv71jznf6/+HypJ6tnvFucdv1Ds3K64YWunZadm90z/k0yWXGe9a/f7lcxnAxIA96+SMjK+Z7Xfjxo5pge8Chw3fTs7skqh13/0KyunWkQIttl+fPdZSP6O9J9IuIQAEjUgBLbE/f1v2NpfrpP5DK6sz7n1vkEgBr5OdaHv9LXYe95E049GgGAJFgt2n23Ff/fL5XabRn6i9qnwDavG+dmp450De3i2WY5Rg9y3PUfRek+C1kIACTtlrmex+4qy3Zl4ptnebdXZ66cx1SwMrdfn9fV5VtoKxN/h5H5rpyQ9cS9/XLcGff+WdntYzLL53m6dMWZKDar/vWVhek+C0Ey7hcYPcXk4a4/frPMmUmDtpOGOkcN6O6JxP7FtjnjesDceE9x02wPKwO6DQGA1M0a6/7tV/tmzkqxdLVnoC2EQrquLe/iVdxoRQDAlGumZP/sc33SfRaapmk2q37znO4o/493wxXZmTkj0kMtn0sAdBMetWDW8is97xxq/sNL9ek9jUVTswo81rS8dJbTsmRmzhPrG9Ly6krsOx5Y9M2TnXyCzao7bHputqUo11rexz5qkGPaSNfsse6uKNwc3t/Rr8h2ppqOoV2OAMh0nfxm6rrmtOsuh8Vp1/OyLaUFttICa79i25iBzjHljqFljm6rh/nObcU7DjbvOhLoptfrSHrHYW6d7+nRAXBZ4YgRjhhNgejZmvDeY4EXtmqapuW4Lcvmer54U8GAEsVXkplj3E9taFR7TMQjAHoww9Cag0ZzMKJp2rla7eCpdu3p3U595hj3vAlZCyZnjeziqVG7TX/4zr7X3nWywZeejWUKPdarK9LZmm3mGPfAEvvJqlAaz6H7ef3Rx16tf3x9w72riz91XZ7CI08b5SIAugEDl72WP2Cs29n03ccuLPzaievvPvnHl+u79Oo8qI/9O7cVd93xO7dsrie9c9GSpy6DIePeP1T9/oU6hccc0pfVAN2BABBh95HAtx+pmvHFYz/9R03XxcCqBbmT09TS69YFZi++humdCoW3hbj/0Qsb9/hVHY3tYroHASBIY1P0Z2tqZn7pWBeNVuu69sCnSrp/N7HRgxzjBzsv/3md+sPLZiexy0vtM0bLbWlpGNqPHq9WdTSWAnQPAkCcel/0zt+c/8gDZ6rq1W//O3mY8z8WdvdkrPnN3/0B47//Vn2qymzZifAFAdsPNm/d36zkUDarnlFrDHsrAkCo13c33XD3yb3H1Nft/J9bCrqzG4/Voi0zXTb+ynafrzn6r7fMzjreOCunR++jad6b7zapOpTbydWpy/FPLNeZ6vDN953etFfZuG2LgSX2JTO7b0HW/ElZffLNlv8/s7FR07RnNnlNHsfjtlw3PT2L0TLEW/uUvZ3cTtFR2j0IANGaAtFP/LjynUNqHttbffbGArUH7IT58Z/Gpui6XU2apr17NHCk0mwdZ/d3o8soR88qK4SNpKeiWBYCQDpfc3T1DytPX1C56nLSUGf37B2Wm21ZPC3b5EFe2OoNhj6oAXp2k9lRoHkTs0oL5E5g1vuUTSwFgiRAlyMAoNV5I59/6Gw4YroQso0V3VIUf9Nsj/mpwmc2Xhz5eXqj2VEgi95N33tm8geMQEjBGykS1RqbCIAuRwBA0zTt7YPNv3q6VuEBr5ue3Q0rs8yX3lc3RDbsuThv+f7p4P4TwU4+PxHCR4GUlABcqA9HVd6QoGMEAD7wq2dqK9W13yrwWGeN7dpRoKFl9ikjzNbdP7fZGzPW/IzpUaAR/R2Th5ldl9BDZbssSoL/xHk6wXUHAgAf8AeM//67soU8mqYtnWl2dL5zSorun4kb8zFfC6QJXhCQn6PmknLgpNnnMCSCAMBFT2/0nq1Rduc1f1IXdmdTMtReWR3eeiC2bPH4udDOw2aXR9w8J8duk1jFOGqgmraDu48orkxDhwgAXBSOGH96VVlb/4EldvMV+pcye5y7X5HZYptn3/J22ALI/ChQfo510ZSufQDKTKqqvzapW0+AThAAaOfvaxvMt0VrZX6M/lIUjf90fKF/dlPHwZAUmVPBSh77jlaGjppekIFEEABo53xdZOdhZU/fXRQA2S7LDVeYvb8+djZ0qR1sztaEtx0w+49wVUVWcV56dihLl/kTs8x35dM07VnTPTmQIAIAsV7e5lN1qIqu6Q69dGZOlulGMZ2P85gfBbJZ9WVzBD0E2Kz6N1YVmj9OJKr9bS0B0E0IAMRSOPw6uGu29VAyuhJf/9NWfHloCkSNAv34MyVK8v75zV5pG6ulEQGAWHuPBVStCi4tsClfDjawxG5+pnH/yeCBU50VGl6oj2zaa7ax5bjBzjGDunYzzkzgduo/uaPPqgUKZmXCEePHT6isRUbnCADEag4aBzu9OCbOalG/s8eK+R7ddKZcavq33eeoWBBgvlddJvNkWT51Xd6GX5R/+Co13+Zvn69j+rc7ye1ahU4cqQyNLVezlnVAie34OZW/0itVdNpJ5OL+whbvD28vMVnOv+xKzwN/rVbbZyktrBbNbtNzs63FudbBpfYx5Y7po1wzx7gVLnfYfzL44BM1qo6GRBAA6MAZdc1BywpVvsdmjHaZ3y1256HmRDKp3hd9fXfTNebK+UvyrPMnZf17h7J59S4yttx5+vHhaTyBOm/k9p9UBlU0kkPiGAJCB06rawqkdl8nJSMqTyc8ttP5RHGCVkmaCk6Nrzn6yQcrj6nbSwAJIgDQgZoGZV3dFe7r5HLoS2eZ3W8ramj/eivRy/rLb/uag2bvSRdNzc7L5hftkhp80dt+WKlqM2EkhfclOmD+qtfK7VD2Hrt+Ro7HbfZoW97zJ97vyNccNT9647DrN0laEJCUo5Whpfec2rKfxg/pQQCgA351mzEpfAJQU/6fZG2PologAqADf1vbcO1dJw+fofFn2jAJjA4Ew8qeAJyKqkT6FtqunGC2z0w4Yjy/ObkL+ms7fF5/NMfck0fFcNewfg6udK0OnAre80jVpr3c+KcZTwDogEvduE2zorqO5Vd6LKaj5M13/TWNyU1vBELGK9sV1PCIWhXciSOVoS//z7lF3zjB1T8TEADogMuhbNxG1WiSovGfVJrMKKkFWqEiwHqucMR45W3fbT88M+/O42veaDTfZgNKMASEDrgVBkBAwRPA5GHOEf3N9lQIhoyXtqZyL79+V1OdN5KfY6q1Z1mRbc74rDffNdteoid651Dz135zvvPeG0gLngDQgZJ8ZXcGSgLgVhV9Zv79TlOjP5U7z3DEeDGl5Ihx6wKho0AVw13//smgZ/5rwPUzss238YBCBAA6YH6zrVa+ZrNP+3abftNss+X/mrkOz0pqga6fnmNyMrnn0nVt2kjX779W9tqDg6aP6qptgpAsoW9HdE5hAFSaXlS8eGq2yeEXTdOaAtHXTFT0b9rbVFVvdnGc26kvnakgyXq00QMdT313wL2ri5W3iUUKCAB0YOQAZU2Mzfd2VzL9+/I2n5nBqEhUS7Z+tEPUAmmaZtG1z96Y/9hdZdkurj9pxg8AsTxZliGKNnIJR4zEl912qDjPunCygm1mzY/hmN8jTNO0K0a7B/Xpkk1yepx5E7OevL9/gUfWrpmZhgBArIlDnapm6iprwiYL/pbN8ZgfK6j3RdfvMlt+s+1A8xnTw1m6rq1Q0c66d5gwxPnYXWUK14ojWQQAYs0dr+COu4X5zT2UjJm8sMUbMr222Uimi1wnVsxTsKFNr1Ex3PXQ50vTfRZyEQCIde10Ux3w29p5KGDmy8cMcowbrGBfGiU1PFpi+4hdVnmp/YrRZre07E2WzMz5zNL8dJ+FUCwEQzvlpfZR6maAt79vqsevkm1mq1Ts7tti15HAsbMh8zvdr5zv2fxeZjVC2Hc8sOibJ2M+qOtatsviybIUeqxjBjnGD3FeNTlrWD/1uxx/+yNFm/b49xwzdbuAFBAAaOe2a1TuYbvDRADYrPrNcxWM/5TkWU/8LZ17XcVbOjPnnj9UKVki16UMQ/P6o15/tLI6vPdYYM0bjff/SZs83HXHkvwPqViZ0cpm1R/6Yul1d500P1KHpDAEhItcDl3V7t6aph0/F0q281pbCyZlleT1zhKRHLfl+uk9dUHAzkPNn3vo7C3fOa12A6/RAx1fuKlA4QGRCAIAF62+Js/8kqtW63aaGnjp3T30e/qCgC37/dd/6+TbB1Vu4/XlZQXmN3xGUggAfCAv23LncpW3YM+ZWDmVl21ZNFXZXHQGmjshq0zdcuu0aPBFP/LAmd1HlA3cO+z6fbcVqzoaEkEA4APfXFWk8Pb/Qn1kq4l9/m6a43HYe3OxpEXXll/Zsx8CNE3zNUc//qMz52vNLo9odd307JljKJHqPgQANE3TrqrI+vjiPIUHfGmbz8wSsFU9fIQkESt7xYqw83WRz//yXFTd3O09q4uUHQuXQwBAKyuyPfSFUrWrk9a80ZDy1w7v75g8vPc3jBze31HRK77Nt/b5//fFOlVHqxjuWjytN4/+ZRQCQLpCj/Wv3+pXqLQlyzuHmrcdSH16sHfcGiei10x0//jvNeb7ZLT6+spCFkt3DwJAtLxsy1++1U9h788Wv/lXXcpf2zsGxxPUa6Y6mgLR7//5gqqjjRvsvE7dcnR0ggCQa1g/x/MPDJw4VEGvhbZOnA+9tDX1+p9eUB6TuN5U7PTMJu/OQ8qqQu9czkNAdyAAhFpyRc5z3x8wpEx92fVDT9Wamf7t6QXyyeo1o0Capv3o8RpVhxo32Lm4t0RjJiMAxOmTb/3tV/v+9qt9c7PV//R3Hmp+fH3q0789eolsanrTguc3djdtMVH7G+OrKwtVHQqXQgAIUuixfvujRRt/Wb7kii65yBqG9u1HqgwTFYE3zsqR1h3eZtWX9aI5j5+vqVV1qPGDnZQDdTUCQIRpI10P3tFn86/KP/+hgixnV/3Q/7auYedhU+tCpY3/tLh1vsoGfOn15rtNCvtDfHUFDwFdS8psm0BZTsusce75E91XV2Sb72B8WcfPhb73mKk6kPJS+4xREleBtmx7sLe3NEP++Zqav3yrn5JDTRjiXDQ1+9XtPiVHQzwCoAfTdc1h050O3WXX87KtpQXWvoW2/sX2MYMcYwY5hpQ5rN31gBcKG5/9xdnGJlPbP0reKmvlfE+vCYD1u5p2vN88ZYSaNW53rigkALoOAZDpxpY7Tz+eWe3s493/6AWTTcGEb5a7bI7n+3+uDkd6STf8n6+peexuNQ8Bk4Y6r56S/e8dZECXIABg1lMbGv/4cr3Jg1wx2j2oj4Jxqkdeqr/3D1Xmj5O4z96Yf+9qsz0si/OsCydn9Zpb3bU7m3YeDkwepmaJyVdXFBIAXYRJYJjy2g7fnb8+b/44qqZ/n1axbW9Snt3kNVP41KqXTYD/fI2yNQGThzmvqshSdTS0RQAgdW/t83/mZ2fND1y4nfrSmQoqU09WhbYr3aIkEWeqw2YaH7VaPDVbYTvutHtth0/hVgFfoxyoaxAASNHOQ82f+HFlIKTg7veGGTk5bgVvxac3pt6CwoxnNil47LDb9Jvn9KpFcL94Ut1DwHDXwsk8BKhHACAVb77bdOt/nfH6TZX9tOq54z8tntvsNdP9otXKXrQgQNO0l9/2KSxtYk1AVyAAkLQ1bzTe9t+VvmY1V/+yItuc8Qpu7g6cCu4/ETR/nBRcqI9s3GNqA+QWk4c5R/RX3Jk1vX6u7iFgygjXgkk8BChGACAJ4Yhx/58ufPl/zoXCygoWV1zpsago/396Q3pu/z94dUWjT71sKvilbb731KUyDwHKEQBI1PFzoeX3n/7dC3VqD6uq/P+ZTemZAGjxwlZvUMV0yHJFcZghDEPlQ8DUka75E3kIUIkAwOUZhvaXfzcs/s+TCtu8tKgY7hquYtBjx/vNx8+FzB8nZY1N0XW7FIwC9S20XTmhV13jXtji3X9S3UMALUKVIgBwGfuOB26+79Q3f3te1ZRvW6sWqJr+Teftf4tnFE1B36ro3yRDGIbKcqBpI129LCDTiwDAJZ2qCn/l1+eu7YIb/xYOu/6h2QoudpGo9qyKQkyTXtnuawooyMjrpud4VBTFZo7nN3sPnmImIBP1qvcZVDl4Kvj1/3d+7leO/+P1xmiX9adZPDU7T8WmNJv2NlXVR8wfxyR/wHh1u4JRIJdDv3FWr1oQEDW0h55Stk/AjNGuueMldo3tCgQALgqFjec2ez/ywJmrvn7ib2sbFJb6dEjVboiZMP7TQtVChF62IEDTtGc3NR46zUxAxqEZHLRgyNiwx//iVu+L23y1jd10K12SZ52voqw7GDKe35IpAbBuZ1ODL2p+r80Zo13lpfb0TmurFTW0X/6z9pdfLFVytCtGu+eMd2/co2z7SbEIAKG8/uje44Gt+5s37fVvO+D3B7q7EfGyKz02q4KCx7U7m0zuQ6BQKGy8sNX7HwsV3L+vnOf5yT+UzZ1mgqc3Nn5leeHQMjV7E311ReHGPaeVHEoyAqD3MwytuiFyqip05Gzo8JnQ4TPBfceDRyqDSnpYpkzVPoj/TOv6r3jPbFITACvme366pia9PyO1IlHtoadqHvqCmoeAmWPcs8a639rHQ4ApBECPETW0aNSIRrWooYUjRjBsBIJGIGQ0h4zmQNTbbDQ2Rb3+aENTpKYxWtMYqW6IVNdHKmvClTVhJWuUFBo32DlmkILyf19z9LUM6xS/cU9TVX2kJM9sX8+BJfaZY3rbBe6fGxrvXF6oaoPSr60sXPFdHgJM0fvd+n66zwEAkAZUAQGAUAQAAAhFAACAUAQAAAhFAACAUAQAAAhFAACAUAQAAAhFAACAUAQAAAhFAACAUAQAAAhFAACAUAQAAAhFAACAUAQAAAhFAACAUAQAAAhFAACAUAQAAAhFAACAUAQAAAhFAACAUAQAAAhFAACAUAQAAAhFAACAUAQAAAhFAACAUAQAAAhFAACAUAQAAAhFAACAUAQAAAhFAACAULZ0nwDQfYaUOX/6pYFtP/LK1vqHn67qnld//HvD7Da99Y8HTzbf9ZtTXfRVQCJ4AgAAoXgC6J3u+2S/ySOyOvyrczWhz//0uGGYfYkvrShdOMXT4V+Fwsaq+w6bfQEAXYwnAHFKC+1jBrtNHsTlsMwan63kfACkCwEg0aXu3BM3e3yOy8GbB+jZ+B2WaPb4HKddv/znXdoC0xECIO0IAIncTsvMcTkpf3mfAvu4IWYHkQCkHZPAUkSjmqVN3C+c4nl9Z2Nqh1o4xaO3eX6IOTIuhYlxZBp+caXYdaip7R8nDMsqyksl/nVdW1DRbvxnZ/sjA+gpCAApth/wNfgirX+Mv44naMxgd2mhve1H1u1oMHtyANKBAJAiEtXe2NVuzCe1idyr2n/VwZPNZ6pCps4MQJoQAIKs29EuAPoXO0YOdCV1BKddnzW+3exxzDEB9CAEgCBHzwSOnw22/UiyCwJmjs9xOy++Z0JhY8MuAgDoqagCkmXdjoZP3FDc+sc5Ez2PPH8hFE60L8RVU3Lb/nHrez5fc1Tl+Wma065XjMweO8Q1pMxZWmjPcVscNksoYvibo+dqQyfOBd875n97f5PXH7n8sVJSlGebMSZ7/FD3gD6OAo/N7dQjEa2hKXK6KrjvaPPmvd6T54OXP0qmynJZKkZmjSl3l/d1lBbYs90Wp90SjhhNzdHzdaHjZ4P7jvrf3q/+x4rMRADI8vrOxtuuK7JaPqjizHFbpo/J3vSuN5GvLc63jR/arvx/3XaV078l+bZl8wsWVHji1xg7LbrTbs33WEcNci2anhuJGJv3+Z5cX3usMqDwBEoL7KuvLZo1PiemqtVq0YrzbMV5tknDsz68qHD3Yf+fXrxw9EzSL53evp7D+jtvnlcwY0x223No4bDoDrs132MdOdC1aHpuOGJs2O1ds67mzAVmd3o5AkCWem/knYNN00ZfbOOzcIonwQBYWJHbtvy/tjGsqgBU17Vl8wpWXV0Yf23qkNWqz5mQM3t8zoub6x976UIgZLqznaZdNTX30x8qSWSB9MRh7gc/P/Af62qeWFtjvqdeN8jNtn5ySfH8yYkO99ms+oIKz9yJOU+ur/3HupooDwO9F3MA4sRM204ekZXvsSbyhTFVQ6/vbFRyaXA7Ld/+WL/V1xYlePVvpevaDbPyfvDZAcUpLWhoa/mCgi8u75N4ewyLRVt1deGdq/q2PktlrFGDXD/70sDEr/6tbFZ91dWF93y8H02fejF+tOJse8/XdgDdatHnTbr81WHMYFdZUUz5v4LpX5fDcs/H+00Z1XHn6kQMKXP+16f7p7aorcW8yZ6PLi5K4QvnTsz53LKSlF+3G0wekfXd2/sX5qb+jzN5RNY9nyhLNpvRUzAEJE44YmzY5b1uZl7rRxZOyX12Q13nX7Wwot307+HTgZPnFMyFfmF5nzGDOyhFra4Pv7m7cdt7vqracL0vkuO2FuVZK0Zmz5uU07/EEfPJpYX2u28ru/vhU4nPZrcqybffcVO7/Dt1PvjGrsZ3DjZVN4R9/mhejrVvof2KsdlzJ3rycmIfla6amnvgRPOr2zJxKdzwAc67b+vg2h2OGFv2+t490nTgRHO9N9Loj7odel62bVS5a9rorCvG5ujtv2LsYPcdN5f8as357jt1dBcCQKK1OxraBkB5X8eQfs5OZjUddn32xJjyfwWXvGum5c6ZENuTLhrVntlQ+/fXatpezWsbw7WN4UOnAmvW1Sydnf/RxUWO9sM1Q/s5b7u26JHnLyR7DhUjLz58BEPGn1+pfn5TXduR/Qt14Qt14T1H/H97teYTS4qvmZYbc4SPXV+8bb+vrrGrqpJSk+O23rU69upvGNpLW+rXrKutbQy3/bjXb3j9wdMXgmu3N5T3dXz25j6jBrVL5aum5G57z7dlr687Th3diCEgiQ6dCsTUMna+IGDmuJysNuX/4Yjx5q6E5o07keO2fuz62IGXaFT76d/PPvZS9aXu5Q1D+9fGuvt+f7o5GDv/cMOs/MFlzpTPJxgy7n/k9HMb6y41r9sUiP76qfN/fCE2Y7JdltuXZtxA0O1Li2NGfpqD0R88Wvm7Z6tirv4xjp8N3vf70/F1AZ9aUmK1MhDU2xAAQq1vP4I/b5Knk1/vmPYPb+/3NTaZveG9cU5ejjt2ROV3/6p6a8/lo+XgyeYH/3o25kptsWirripM+Xx+/vjZ/cebL/tpz26oe35TXcwHZ4/PiZkgSa9h/Z3z2zd6ikSNB/96dvuBhG7hQ2HjoX+cO3y63RNhSb5tYUrNo5DJCAChYmp4crOtUy8xE1uUZ5swrN1fmZ/+tVr1a6/Ii/ngniP+l7fUJ3iEdw42xQ9DzRib3acglQvxW3u8W/YlOr7x55erq+vb3UTrunbN9NihoTRauTA2CJ9aX/vOwSRqdlsyICZiF83IoO8RShAAQtU0hGMaRMdM87ZaUNGu+3+DL7LjgNny/ykjs3KzY2////e5qqQO8tjL1TEDQbqupVDvGI1qj75YnfjnB0LGX1+N/fyrpuRmyAhJYa6t7ToPTdNqGsJPvV6b7HFOnQ/GPI2NGOAyU22FDEQAyBVzBz11dJYnq4MFAQvbt394Y2djJGp2+VPMFUrTtP0nmmP6FF1WvTeyOW5acuropCtK3z3SdK42uSWvG9/1+gPtsicvxzpyYOozEArNnhC7knndjsbU1spt2B07HDdpeOoFu8hABIBcW/a1a/lis+pXToqtyRk1yNWvuN2gyloV5f8xLSU0TXv9nVQOuz7uq4b3dyW7cOnNnUlPaAdDRnz2DO+fXGvVLjJlZOw1OsGV3vH2HGmKGQUa1j8jQg6qEAByhcLGxva3eDE3+/EfOVYZMN9+x+209C2MHanfnVJXiX1H/ZFIu0uUxaKV941dKNC5PUf9Kbz03rivGpFkb+2uoOvayPYVnMGQcexsij8yrz8aM9s/sE9y/7bIcASAaDGjQMP6OweWXvwNt9v0Oe3L/5Xc/vcvsccsNWoORs/WpNJ3LBwxTsT15kzqItXUHD2f5PhPi/ggHNov/XfHfQrsbQt2NU07XRU007CoztsuAMwsKkYG4scp2oETzWcuhNoO8iyc4mmdEb1ibHa26+LVJBI13lTR/T9+IvHEudQvUsfPBoe0L/9PaqIyPj8SFN8UOn6dcPeLf7QaXOZ8/HvDWv6/JXcvhu8Hf9Tb/u3Fv+xoSjvHzS1jr8KPU7r177R7CJg/2dM6hXjV1HbjPzsONtV7Fax3zYur/zGzqqDtRseXOn4nvKm+dChsxMysZjktHV40u1Nhbuz3ruua3aa3/Gez6jarbm39z6JbLbrForX8p+vt/uuQI+F+eegRCADp1u9obHv3XeCxtVR6FObaJsaU/29Xs/mXwx77rmsysf1ITDWOpmkORxIXqfgvT+Jrm2OLUN3ONP9CdfUJZH73UySFAJDuQn343SPt5jNbtv1aUOFpW03Y2BR5e7+aVjD2uHr5QDD1Uer4nhDxx++EqZcOxb502jsnO2z8RiMJvF0Qu7HXjLHZ2S5LTHegN3d5wxE1u5+E4o7jTOaePUb8NTf++J0w9dJxjzLxadTNoj1ihxpkDCaBoW3e6/1MoKR19MBu0z+xpDim67KS9p8tgnE3zlmu1G9E4gc9gsnc1JsZM3G3P23DMDWgpET8gq9XttY//HRyS6whB08A0AIhY1P7Rf9Xt5/+PXkuGNMazIz4meT4rnCJi1+9XJfMTHVOR4ufE2G36TE7iDUF0n//Hf9vWxpXFwS0IgCgaZfr76Zk869W1Q2x7YgHlTpSrp8ZHLfsqybu+J0YlOrKpvjVBkpKpEyKX9MQXxgKtCIAoGma9t4x/7lLLMWKRrXXd6oMgFPnQzF3yh2uDU6E1aq3XbnWIr5CvxNZLktqDUTj9x44cukddbrNyfPBmE5NpYV283smo7ciAKBpmmYYHfTVabHzUFPnW4gkq8N1vxOGxXYHSsTYwS5b+5qfSNRItqnc+CGpvPS4uK86dOry2wl0tVDYOFYZ++3Ht94DWhAA+MC69gsCLn58u7Lp31bvHo5tpDMv+TbOmqbNj9ui5NCpQLKlOFdOjm2Bd1kOuz5zXOxV9f0MCABN0+J3fVk4lY1c0DECAB84Xxvadyz2uuxrjm59T/1OsPFLCsaUu5NtNJabbZ01LvbancJihQlDs5IdBZo9PiemfKjeGzl4Mv1DQJqmbYzr4TxigGv6GB4C0AEGB3HRvb873T0v9M77TQ2+SNs9YXRdu31p8f2PnEn8IKsXF8VchQ0jlekKi0X7+PVFD/71bIKf77TrH10cu5vx2h0NEUXrJEw6eT6454g/puH2J24o3nfU7zOx4hq9Ek8ASINIxHhxc+zujxOHZ10zLdFNBycNz7o67pM37/VeqEtlumLW+JwrxiZ6j7z62qL4fnOvbVM/UJayv79WE/ORsiL71z/Sl0YOiEEAID2e31QX3wPujptLpicwYzligOs/V/eNqRyNRrUn1sZe+BJ356q+owddvqH/jXPyl8zOj/ngxne9ldWp9JTuIvuO+TvczOs7n+qX70ll3UNRnu0ji4o+dn3scw96OgIA6eH1R/8UtxOv1aJ/c3XfDy8qtF2in4+ua0tm53/v0/3jO0A8t6ku2fqfthx2/Tu39186O/9SKxLcTsvnlvX55JLimI/7mqPJ7mbcDX77zPn4h6HxQ90/+9KgJbPznYk19fRkWRdUeP5zddnD3yhfsbAgL5sR496GnyjSZu32hskjsua233PGatFXLixcWJH7xq7Gt9/zVdWH672RHLelKM9WMTJr3iTPgI7mig+fDvzllSQ2dm/1zsGm0eWulrkEp13/1NLiRTNy39jZ+M7BpuqGsM8fzc+x9im0zxyXfeVET4cd/x998UJdY/qXgMXw+qMPPHrm+58ZkN2+X0V+jvX2pcWrri54e3/TvqP+Q6cDDb6I1x+JRrUslyXLacnPsQ7q6yjv6xzazzligMvCLWKvRgAgnf7nyXOFHuvYuJr64nzbLfMLbplfkMhBztaEfvhYZSicyhxsVV3ojZ2NX761tPUjA/s4Prq4KH6at0NrdzS8mkmj/20dPxv87iOn7/l4v9y4DRJy3NYFFZ4FcUW0kIZ8RzoFQsb3/1S540AqGwK3OFoZuPd3p5Nq/xDj9Z2NqT09bNzt/c0/M27wp61DpwJ3P3wqE5YoIzMRAEiz5mD0gUfPPPZydbK38IahvfBW/bcePlVdb3ah8pPra3/9z/OJn4BhaE+srfnZ42czpPSzE5XVobt+c+ovr1SbaVV9rjb02MvVf3rxgsITQyZgCAjpZxjaP1+vfXNX4y3zChZM8Vx2W5VIxNi8z/fk+tr4ndlT9tq2hr1H/LddV3TF2JzOO9PtPux/9MULPei2Ohwxnlxf++q2hsUzcq+ellua8Kq3qrrw9v2+re/5dh1qSnujU3QFvd+t76f7HICLnHa9YmT2+KHuwWWO0gJ7TpbVYdPDEcPXHD1fGzpxLrjvqP/t/U1ef1fNu5bk22aMzRk3xDWwj6PAY3M5LJGo0eCLnK4K7jvW/NYeb1LN5jLQkDLn+KHu4QOcZUWO4jyb22Vx2PRAKNrUHPUHolV14RPngsfPBg6dCvT07xSXRQAAgFDMAQCAUAQAAAhFAACAUAQAAAhFAACAUAQAAAhFAACAUAQAAAhFAACAUAQAAAhFAACAUAQAAAhFAACAUAQAAAhFAACAUAQAAAhFAACAUAQAAAhFAACAUAQAAAhFAACAUAQAAAhFAACAUAQAAAhFAACAUAQAAAhFAACAUAQAAAhFAACAUAQAAAhFAACAUAQAAAhFAACAUAQAAAhFAACAUAQAAAhFAACAUAQAAAhFAACAUAQAAAhFAACAUAQAAAhFAACAUAQAAAhFAACAUAQAAAhFAACAUAQAAAhFAACAUAQAAAhFAACAUAQAAAhFAACAUAQAAAhFAACAUAQAAAhFAACAUAQAAAhFAACAUAQAAAhFAACAUAQAAAhFAACAUAQAAAhFAACAUAQAAAhFAACAUAQAAAhFAACAUAQAAAhFAACAUAQAAAhFAACAUAQAAAhFAACAUAQAAAhFAACAUAQAAAhFAACAUAQAAAhFAACAUAQAAAhFAACAUAQAAAhFAACAUAQAAAhFAACAUAQAAAhFAACAUAQAAAhFAACAUAQAAAhFAACAUAQAAAhFAACAUAQAAAhFAACAUAQAAAhFAACAUAQAAAhFAACAUAQAAAhFAACAUAQAAAhFAACAUAQAAAhFAACAUAQAAAhFAACAUAQAAAhFAACAUAQAAAhFAACAUAQAAAhFAACAUAQAAAhFAACAUAQAAAhFAACAUAQAAAhFAACAUAQAAAhFAACAUAQAAAhFAACAUAQAAAhFAACAUAQAAAhFAACAUAQAAAhFAACAUAQAAAhFAACAUAQAAAhFAACAUAQAAAhFAACAUAQAAAhFAACAUAQAAAhFAACAUAQAAAhFAACAUAQAAAhFAACAUAQAAAhFAACAUAQAAAj1/wFds5CI2+oJ1AAAAABJRU5ErkJggg==')
            self.send_response(200); self.send_header('Content-Type','image/png')
            self.send_header('Content-Length',len(ico)); self.end_headers(); self.wfile.write(ico)
        elif self.path == '/ai_status':
            self._json({'active':ai_active(),'name':ai_name()})
        elif self.path == '/template_status':
            saved = STATE['template'] is not None or os.path.exists(TEMPLATE_CACHE)
            self._json({'saved':saved})
        elif self.path == '/get_collage':
            photos_xlsx = STATE.get('last_collage')
            if photos_xlsx:
                self.send_response(200)
                self.send_header('Content-Type','application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
                self.send_header('Content-Length', len(photos_xlsx))
                self.end_headers(); self.wfile.write(photos_xlsx)
            else:
                self.send_response(204); self.end_headers()
        elif self.path == '/history':
            self._json(STATE['history'])
        elif self.path.startswith('/download/'):
            rec_id = self.path.split('/download/')[-1]
            xlsx = STATE['dar_files'].get(rec_id)
            if xlsx:
                rec = next((r for r in STATE['history'] if r['id']==rec_id), {})
                fname = rec.get('filename','DAR_Report.xlsx')
                self.send_response(200)
                self.send_header('Content-Type','application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
                self.send_header('Content-Disposition', f'attachment; filename="{fname}"')
                self.send_header('Content-Length', len(xlsx))
                self.end_headers(); self.wfile.write(xlsx)
            else:
                b = b'File not available (server restarted)'
                self.send_response(404); self.send_header('Content-Type','text/plain')
                self.send_header('Content-Length',len(b)); self.end_headers(); self.wfile.write(b)
        elif self.path.startswith('/delete/'):
            rec_id = self.path.split('/delete/')[-1]
            STATE['history'] = [r for r in STATE['history'] if r['id'] != rec_id]
            STATE['dar_files'].pop(rec_id, None)
            save_history(STATE['history'])
            self._json({'ok':True})

    def do_POST(self):
        path = urlparse(self.path).path
        length = int(self.headers.get('Content-Length',0))
        body = self.rfile.read(length)
        ct = self.headers.get('Content-Type','')

        if path == '/upload/template':
            boundary = ct.split('boundary=')[-1].encode()
            for part in body.split(b'--'+boundary):
                if b'filename=' in part:
                    idx = part.find(b'\r\n\r\n')
                    if idx != -1:
                        data = part[idx+4:].rstrip(b'\r\n--')
                        STATE['template'] = data
                        with open(TEMPLATE_CACHE,'wb') as f: f.write(data)
            self._json({'ok':True})

        elif path == '/upload/new_serial':
            boundary = ct.split('boundary=')[-1].encode()
            for part in body.split(b'--'+boundary):
                if b'filename=' in part:
                    idx = part.find(b'\r\n\r\n')
                    if idx != -1:
                        data = part[idx+4:].rstrip(b'\r\n--')
                        try:
                            import pandas as pd
                            df = pd.read_excel(io.BytesIO(data), header=None)
                            col = df.iloc[:,0].astype(str).str.strip()
                            skip = ['serial','serial number','nan','']
                            serials = [s for s in col.tolist() if s.lower() not in skip]
                            STATE['new_serials'] = serials
                            self._json({'ok':True,'count':len(serials),'serials':serials})
                        except Exception as e:
                            self._json({'ok':False,'error':str(e)})
                        return
            self._json({'ok':False})

        elif path == '/parse_email':
            try:
                result = parse_email_text(json.loads(body).get('text',''))
                self._json(result)
            except Exception as e: self._json({'error':str(e)})

        elif path == '/parse_email_image':
            boundary = ct.split('boundary=')[-1].encode()
            for part in body.split(b'--'+boundary):
                if b'filename=' in part:
                    idx = part.find(b'\r\n\r\n')
                    if idx != -1:
                        img_data = part[idx+4:].rstrip(b'\r\n--')
                        if ai_active():
                            result = extract_from_email_img(img_data)
                            self._json({'ok':True,'fields':result})
                        else:
                            self._json({'ok':False,'error':'No AI key'})
                        return
            self._json({'ok':False})

        elif path == '/upload_tracking':
            # Upload new tracking Excel — parse and return tickets JSON
            try:
                import openpyxl
                boundary = ct.split('boundary=')[-1].encode()
                for part in body.split(b'--'+boundary):
                    if b'filename=' in part and b'name="file"' in part:
                        idx = part.find(b'\r\n\r\n')
                        if idx != -1:
                            data = part[idx+4:].rstrip(b'\r\n--')
                            wb2 = openpyxl.load_workbook(io.BytesIO(data), read_only=True, data_only=True)
                            ws2 = wb2['SMB DAR List - Athena']
                            tickets = []
                            for row in ws2.iter_rows(min_row=4, max_row=300, values_only=True):
                                if not row[1] or str(row[1]).strip() in ['Ticket Number', '']: continue
                                try:
                                    ticket = str(int(float(str(row[1]).strip())))
                                except:
                                    ticket = str(row[1]).strip()
                                def sd(v):
                                    if hasattr(v,'strftime'): return v.strftime('%d/%m/%Y')
                                    return str(v).strip() if v else ''
                                defects = []
                                if row[14]: defects.append('LED Driver')
                                if row[15]: defects.append('LED Module')
                                if row[16]: defects.append('LED Driver')
                                if row[17]: defects.append('SPD')
                                if row[18]: defects.append('Water Ingression')
                                defects = list(dict.fromkeys(defects))
                                tickets.append({
                                    'ticket': ticket,
                                    'location': str(row[2] or '').strip(),
                                    'contract': str(row[3] or '').strip(),
                                    'model': str(row[4] or '').strip(),
                                    'reported': sd(row[5]),
                                    'site_visit': sd(row[7]),
                                    'serial': str(row[9] or '').strip(),
                                    'defects': defects,
                                    'reason': str(row[19] or 'N/A').strip(),
                                    'status': str(row[26] or 'In Progress').strip(),
                                    'closed_date': sd(row[27]),
                                    'do_no': str(row[24] or '').strip(),
                                    'new_serial': str(row[25] or '').strip().replace('\n',''),
                                })
                            self._json({'ok': True, 'tickets': tickets})
                            return
                self._json({'ok': False, 'error': 'No file found'})
            except Exception as e:
                self._json({'ok': False, 'error': str(e)})

        elif path == '/scan_unit':
            # Scan photos for one unit — extract QR + AI data
            boundary = ct.split('boundary=')[-1].encode()
            parts_raw = body.split(b'--'+boundary)
            imgs = {}
            for part in parts_raw:
                if b'name="img__' in part and b'filename=' in part:
                    try:
                        ns = part.find(b'name="img__')+len(b'name="')
                        ne = part.find(b'"', ns)
                        field = part[ns:ne].decode()
                        _, uid, itype = field.split('__',2)
                        idx = part.find(b'\r\n\r\n')
                        if idx != -1: imgs[itype] = part[idx+4:].rstrip(b'\r\n--')
                    except: pass

            extracted = {}
            # 1. QR scan
            for t in ['serial','full','issue']:
                if t in imgs:
                    qr = decode_qr(imgs[t])
                    if qr:
                        extracted['qr_code'] = qr
                        qr_parsed = parse_qr(qr)
                        extracted.update(qr_parsed)
                        print(f"  QR: {qr[:40]}")
                        print(f"  QR parsed: {qr_parsed}")
                        break
            # 2. AI extraction if serial not found
            if not extracted.get('serial') and 'serial' in imgs and ai_active():
                ai_data = extract_from_serial_photo(imgs['serial'])
                for k,v in ai_data.items():
                    if v and not extracted.get(k): extracted[k] = v
            print(f"  Final extracted: {extracted}")
            self._json({'ok':True,'extracted':extracted})

        elif path == '/generate_mobile':
            try:
                boundary = ct.split('boundary=')[-1].encode()
                parts_raw = body.split(b'--'+boundary)
                metadata = None; info = None; photos = {}

                fd_fields = {}
                for part in parts_raw:
                    if b'name="metadata"' in part and b'filename=' not in part:
                        idx = part.find(b'\r\n\r\n')
                        if idx != -1: metadata = json.loads(part[idx+4:].rstrip(b'\r\n--'))
                    elif b'name="info"' in part and b'filename=' not in part:
                        idx = part.find(b'\r\n\r\n')
                        if idx != -1: info = json.loads(part[idx+4:].rstrip(b'\r\n--'))
                    elif b'name="email_addr"' in part and b'filename=' not in part:
                        idx = part.find(b'\r\n\r\n')
                        if idx != -1: fd_fields['email_addr'] = part[idx+4:].rstrip(b'\r\n--').decode()
                    elif b'name="email_pass"' in part and b'filename=' not in part:
                        idx = part.find(b'\r\n\r\n')
                        if idx != -1: fd_fields['email_pass'] = part[idx+4:].rstrip(b'\r\n--').decode()
                    elif b'name="send_email"' in part and b'filename=' not in part:
                        idx = part.find(b'\r\n\r\n')
                        if idx != -1: fd_fields['send_email'] = part[idx+4:].rstrip(b'\r\n--').decode()
                    elif b'photo__' in part and b'filename=' in part:
                        try:
                            # Find name field
                            name_start = part.find(b'name="') + 6
                            name_end = part.find(b'"', name_start)
                            field = part[name_start:name_end].decode()
                            parts_f = field.split('__')
                            if len(parts_f) >= 3:
                                uid = parts_f[1]
                                itype = parts_f[2]
                                # Find data after double CRLF
                                idx = part.find(b'\r\n\r\n')
                                if idx != -1:
                                    data = part[idx+4:].rstrip(b'\r\n--')
                                    if len(data) > 100:  # valid image
                                        if uid not in photos: photos[uid] = {}
                                        photos[uid][itype] = data
                                        print(f"  Photo received: unit={uid} type={itype} size={len(data)}")
                        except Exception as pe:
                            print(f"  Photo parse error: {pe}")

                if not metadata or not info:
                    raise Exception("Missing metadata or info")

                tpl = STATE['template'] or load_template()
                if not tpl: raise Exception("No template")

                # Build units list
                print(f"  Photos received: {list(photos.keys())}")
                print(f"  Metadata units: {[m['id'] for m in metadata]}")
                units_list = []
                for m in metadata:
                    uid = str(m['id'])
                    unit_photos = photos.get(uid, {})
                    print(f"  Unit {uid}: photos={list(unit_photos.keys())} extracted={list(m.get('extracted',{}).keys())}")
                    units_list.append({
                        'folder': f"unit_{m['id']}",
                        'imgs': unit_photos,
                        'extracted': m.get('extracted', {}),
                    })

                block_data = [{'cause':m.get('cause',''),'action':m.get('action',''),
                               'rca':m.get('rca',''),'new_serial':m.get('new_serial','')}
                              for m in metadata]

                new_serials = [m.get('new_serial','') for m in metadata]

                out = generate_dar(tpl, units_list, new_serials, info, block_data)

                # Save to history
                import datetime as _dt
                rec_id = str(int(_dt.datetime.now().timestamp() * 1000))
                contract_clean = info.get('contract','').replace('/','_').replace(' ','_') or 'DAR'
                ticket = info.get('ticket','') or 'NoTicket'
                fname = f"DAR_{contract_clean}_{ticket}.xlsx"
                record = {
                    'id':       rec_id,
                    'ticket':   info.get('ticket','—'),
                    'station':  info.get('station','—'),
                    'contract': info.get('contract','—'),
                    'staff':    info.get('staffname','—'),
                    'date':     _dt.datetime.now().strftime('%d/%m/%Y %H:%M'),
                    'units':    len(metadata),
                    'filename': fname,
                }
                STATE['history'].insert(0, record)  # newest first
                STATE['dar_files'][rec_id] = out    # cache file in memory
                # Keep only last 50 records in memory
                if len(STATE['history']) > 50:
                    old = STATE['history'][50:]
                    STATE['history'] = STATE['history'][:50]
                    for o in old:
                        STATE['dar_files'].pop(o['id'], None)
                save_history(STATE['history'])
                # Sync to Google Sheets in background thread
                threading.Thread(target=sync_to_gsheet, args=(record,), daemon=True).start()

                self.send_response(200)
                self.send_header('Content-Type','application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
                self.send_header('Content-Disposition', f'attachment; filename="DAR_Report.xlsx"')
                self.send_header('Content-Length', len(out))
                self.end_headers(); self.wfile.write(out)

            except Exception as e:
                b = str(e).encode()
                self.send_response(500); self.send_header('Content-Type','text/plain')
                self.send_header('Content-Length',len(b)); self.end_headers(); self.wfile.write(b)

    def _json(self, data, code=200):
        b = json.dumps(data).encode()
        self.send_response(code); self.send_header('Content-Type','application/json')
        self.send_header('Content-Length',len(b)); self.end_headers(); self.wfile.write(b)

# ─── Main ───────────────────────────────────────────────────────────────────────
def start_ngrok(port):
    """Start ngrok tunnel and return public URL."""
    try:
        from pyngrok import ngrok, conf
        # Use authtoken if set
        token = os.environ.get('NGROK_AUTHTOKEN', '')
        if token:
            conf.get_default().auth_token = token
        tunnel = ngrok.connect(port, "http")
        return tunnel.public_url
    except ImportError:
        print("  pyngrok not installed — WiFi only mode")
        return None
    except Exception as e:
        print(f"  ngrok error: {e}")
        return None

def print_qr_terminal(url):
    """Print QR code in terminal for easy scanning."""
    try:
        import qrcode
        qr = qrcode.QRCode(border=1)
        qr.add_data(url)
        qr.make(fit=True)
        qr.print_ascii(invert=True)
    except: pass

def main():
    # Railway uses PORT env var; fallback to 5679 for local
    port = int(os.environ.get('PORT', 5679))
    is_railway = 'RAILWAY_ENVIRONMENT' in os.environ or ('PORT' in os.environ and os.environ.get('PORT') != '5679')
    ip = get_local_ip()

    print("="*55)
    print("   DAR MOBILE v2.0 - SNFOR SDN BHD")
    print("="*55)
    print(f"\n✓ Running on port {port}")
    print(f"  AI: {ai_name()}")
    print(f"  Mode: {'☁ Railway Cloud' if is_railway else '🖥 Local'}")

    if is_railway:
        domain = os.environ.get('RAILWAY_PUBLIC_DOMAIN', '')
        public_url = f"https://{domain}" if domain else None
        if public_url:
            print(f"\n  🌐 PUBLIC URL: {public_url}")
            print("\n  📱 QR Code untuk workers:")
            print_qr_terminal(public_url)
        else:
            print("\n  ⚠ Railway domain belum ready...")
    else:
        use_ngrok = os.environ.get('USE_NGROK', '1') == '1'
        public_url = None
        if use_ngrok:
            print("\n  Starting ngrok tunnel...")
            public_url = start_ngrok(port)
        if public_url:
            print("="*55)
            print("  🌐 PUBLIC URL:")
            print(f"  {public_url}")
            print("="*55)
            print("\n  📱 Scan QR:")
            print_qr_terminal(public_url)
        else:
            print("="*55)
            print(f"  📱 LOCAL: http://{ip}:{port}")
            print("="*55)

    print("\n  Press Ctrl+C to stop.\n")

    server = HTTPServer(('0.0.0.0', port), Handler)
    try: server.serve_forever()
    except KeyboardInterrupt:
        try:
            from pyngrok import ngrok
            ngrok.kill()
        except: pass
        print("\nStopped.")

if __name__ == '__main__':
    main()